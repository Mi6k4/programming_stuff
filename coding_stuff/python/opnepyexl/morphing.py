from openpyxl import Workbook, load_workbook

target_workbook = Workbook()
sheet=target_workbook.active

sheet["A1"]="Номер заказа у партнера"
sheet["B1"]="Номер сертификата" # номер сертификата
sheet["C1"]="Продукт в системе партнера" #назначение платежа
sheet["D1"]="Код продукта в BD"
sheet["E1"]="Дата начала действия" #дата оплаты
sheet["F1"]="Дата окончания действия" # дата окончания сертификата
sheet["G1"]="Стоимость" #сумма
sheet["H1"]="ФИО плательщика" #ФИО
sheet["I1"]="Дата рождения плательщика"
sheet["J1"]="Пол плательщика"
sheet["K1"]="Номер телефона плательщика" #Номер телефона
sheet["L1"]="Адрес электронной почты плательщика"
sheet["M1"]="Серия паспорта плательщика"
sheet["N1"]="Номер паспорта плательщика"
sheet["O1"]="Кем выдан паспорт плательщика"
sheet["P1"]="Дата выдачи паспорта плательщика"
sheet["Q1"]="Адрес плательщика"
sheet["R1"]="Гражданство плательщика"
sheet["S1"]="Город"
sheet["T1"]="Банк"
sheet["U1"]="Наименование ДО" #Офис


morphing_workbook1=load_workbook(filename="sample_for_test.xlsx")
target_sheet=morphing_workbook1.active

for value in target_sheet.iter_cols(min_row=4,min_col=2,values_only=True):
    if value=="Назначение платежа":
        length=morphing_workbook1.max_row()
        for row in range(0,length):
            morphing_workbook1.append([row])
        #for row_value in target_workbook.iter_rows(min_row=2,min_col=3,max_col=3,values_only=True):


target_workbook.save(filename="target_table.xlsx")