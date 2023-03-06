from openpyxl import Workbook, load_workbook

workbook = Workbook()
sheet=workbook.active

sheet["A1"]="hello"
sheet["B1"]="world"

workbook.save(filename="hello_world.xlsx")

workbook1= load_workbook(filename="sample_for_test.xlsx")
print(workbook1.sheetnames)

sheet1= workbook1.active

print(sheet1.title)
for value in sheet1.iter_rows(min_row=1,max_row=2,min_col=1,max_col=3,values_only=True):
    print(value)


for row in sheet.rows:
    print(row)

workbook2=Workbook()
sheet=workbook.active