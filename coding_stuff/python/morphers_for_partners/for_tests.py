import os
import datetime
import yadisk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from mitosheet import *
import pyexcel as p


y = yadisk.YaDisk('406760da5b1345a88999f0acb1ef95bf', '7ecb5577edb643d3a07ce020292db4b6',
                  "AQAEA7qkBXctAAfNlAAXrxr49UN8l0uBHNszaAo")





name = ""
#wb = load_workbook(name,data_only=True)
#wb.save(name+"x")

#base = os.path.splitext(name)[0]
#print(base)
#os.rename(name, base + ".xlsx")

#new_name = name + "x"
#print(new_name)
#os.rename(name,new_name)
#wb = load_workbook(new_name, data_only=True)
wb = load_workbook(name, data_only=True)
#list_of_formated_file_name.append(new_name)


def kotzdorov(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Добавляем новую колонку
    sheet_Name.insert(6, 'new-column-twfs', 0)
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={1: 'ФИО плательщика',
                               3: 'Номер заказа у партнера',
                               4: 'Дата рождения плательщика',
                               'new-column-twfs': 'Пол плательщика',
                               6: 'Дата начала действия',
                               8: 'Продукт в системе партнера',
                               10: 'Номер телефона плательщика',
                               11: 'Адрес электронной почты плательщика',
                               12: 'Номер сертификата'}, inplace=True)
    # Меняем формат пола
    sheet_Name['Пол плательщика'] = IF(TYPE(sheet_Name[5]) != 'NaN',
                                       SUBSTITUTE(SUBSTITUTE(sheet_Name[5], 'Мужской', 'М'), 'Женский', 'Ж'), None)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[(sheet_Name[5].notnull()) & (~sheet_Name[5].str.contains('Пол', na=False))]
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df


def data_placer(file_data_frame, file_name):
    # Присваиваем кастомное имя для результата
    # Шаблон заполнения файла
    shablon = ['Номер заказа у партнера', 'Номер сертификата', 'Продукт в системе партнера', 'Код продукта в BD',
               'Дата начала действия', 'Дата окончания действия', 'Стоимость', 'ФИО плательщика',
               'Дата рождения плательщика', 'Пол плательщика', 'Номер телефона плательщика',
               'Адрес электронной почты плательщика', 'Серия паспорта плательщика', 'Номер паспорта плательщика',
               'Кем выдан паспорт плательщика', 'Дата выдачи паспорта плательщика', 'Адрес плательщика',
               'Гражданство плательщика', 'Город', 'Банк', 'Наименование ДО']
    # Инициализирует новую книгу и активирует ее для дальнейшей работы
    wb = Workbook()
    ws = wb.active
    # Заполняет 1 строку по шаблону
    ws.append(shablon)
    # Находим нужное место для проставления данных по хедерам
    for data_name in list(file_data_frame):
        for number, name in enumerate(shablon):
            if data_name == name:
                # Получаем буквенную координату в книге
                car = get_column_letter(number + 1)
                # Запускам цикл заполнения столбца
                for row_number in range(2, len(file_data_frame[data_name]) + 2):
                    # Проставляем данные в книгу
                    ws[car + str(row_number)] = file_data_frame[data_name][
                        list(file_data_frame[data_name].keys())[row_number - 2]]
                break



    new_file_name = file_name + " morphed.xlsx"
    wb.save(new_file_name)

    return new_file_name

def mts(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Добавляем новый колонки
    sheet_Name.insert(10, 'new-column-4p5a', 0)
    sheet_Name.insert(12, 'new-column-vv6o', 0)
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={7: 'Продукт в системе партнера',
                               8: 'Номер сертификата',
                               'new-column-4p5a': 'Дата начала действия',
                               11: 'Стоимость',
                               'new-column-vv6o': 'ФИО плательщика'}, inplace=True)
    # Применяем формулы
    sheet_Name['ФИО плательщика'] = IF(TYPE(sheet_Name[10]) != 'NaN', PROPER(sheet_Name[10]), None)
    sheet_Name['Дата начала действия'] = IF(TYPE(sheet_Name[9]) != 'NaN',
                                            LEFT(sheet_Name[9], INT(FIND(sheet_Name[9], ' '))), None)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[
        (sheet_Name['ФИО плательщика'].notnull()) & (
            ~sheet_Name['ФИО плательщика'].str.contains('Покупатель', na=False))]
    # Меняем на нужный формат
    sheet_Name['Стоимость'] = to_float_series(sheet_Name['Стоимость'])
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df


def smp_stahovanie(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={0: 'Дата начала действия',
                               2: 'Номер сертификата',
                               3: 'Номер заказа у партнера'}, inplace=True)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[
        (sheet_Name['Номер сертификата'].notnull()) & (
            ~sheet_Name['Номер сертификата'].str.contains('Промокод', na=False))]
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df



data_frame=smp_stahovanie(name)
morphed_file_name=data_placer(data_frame,name)
print(morphed_file_name)