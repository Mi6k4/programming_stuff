import yadisk
import os
from openpyxl import Workbook, load_workbook
import datetime
import yadisk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from mitosheet import *



y = yadisk.YaDisk('406760da5b1345a88999f0acb1ef95bf', '7ecb5577edb643d3a07ce020292db4b6',
                  "AQAEA7qkBXctAAfNlAAXrxr49UN8l0uBHNszaAo")
archive_path="disk:/Account/_Операционное сопровождение/Архив/_____Архив/"
morphed_path="disk:/Account/_Операционное сопровождение/Морфер реестров партнеров/Файлы для загрузки в Thales (после обработки)/"
root_dir="disk:/Account/_Операционное сопровождение/Морфер реестров партнеров/"
now=datetime.datetime.now()






def morpher_list(folder_name):
    list = {
        'ГЭБ (Газэнергопромбанк)': {'function_name': geb},
        #'МТС': {'function_name': mts},
        'ТКБ, ИТБ': {'function_name': tkb_itb},
        'СКБ': {'function_name': skb},
        'ХКС': {'function_name': hks},
        #'КотЗдоров (Деньги Сразу)': {'function_name': kotzdorov},
        #'СМП-страхование': {'function_name': smp_stahovanie},
        'Арсенал Крым': {'function_name': arsenal_krim},
    }
    if folder_name in list:
        function_name = list[folder_name]['function_name']
    else:
        function_name = None
    return function_name



def geb(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Добавляем новый колонки
    sheet_Name.insert(4, 'new-column-yob7', 0)
    sheet_Name.insert(3, 'new-column-2ydk', 0)
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={1: 'tttt',
                               'new-column-2ydk': 'Продукт в системе партнера',
                               'new-column-yob7': 'ФИО плательщика',
                               4: 'Номер телефона плательщика',
                               5: 'Номер сертификата',
                               6: 'Стоимость',
                               7: 'Дата начала действия',
                               8: 'Дата окончания действия',
                               9: 'Наименование ДО'}, inplace=True)
    # Применяем формулы
    sheet_Name['ФИО плательщика'] = IF(TYPE(sheet_Name[3]) != 'NaN', PROPER(sheet_Name[3]), None)
    sheet_Name['Продукт в системе партнера'] = IF(TYPE(sheet_Name[2]) != 'NaN',
                                                  SUBSTITUTE(sheet_Name[2],
                                                             LEFT(sheet_Name[2], INT(FIND(sheet_Name[2], '"') - 1)),
                                                             ''), None)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[
        (sheet_Name['ФИО плательщика'].notnull()) & (~sheet_Name['ФИО плательщика'].str.contains('Фио', na=False))]
    # Меняем на нужный формат
    sheet_Name['Стоимость'] = to_float_series(sheet_Name['Стоимость'])
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df

def tkb_itb(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Добавляем новый колонки
    sheet_Name.insert(8, 'new-column-0kp5', 0)
    sheet_Name.insert(11, 'new-column-2dnm', 0)
    sheet_Name.insert(26, 'new-column-oj55', 0)
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={1: 'Номер сертификата',
                               4: 'Продукт в системе партнера',
                               5: 'Стоимость',
                               'new-column-0kp5': 'Дата начала действия',
                               8: 'Наименование ДО',
                               'new-column-2dnm': 'ФИО плательщика',
                               10: 'Дата рождения плательщика',
                               11: 'Серия паспорта плательщика',
                               12: 'Номер паспорта плательщика',
                               13: 'Кем выдан паспорт плательщика',
                               14: 'Дата выдачи паспорта плательщика',
                               16: 'Адрес плательщика',
                               17: 'Номер телефона плательщика',
                               18: 'Адрес электронной почты плательщика',
                               20: 'Гражданство плательщика',
                               'new-column-oj55': 'Пол плательщика',
                               25: 'Банк'}, inplace=True)
    # Применяем формулы
    sheet_Name['Дата начала действия'] = LEFT(TEXT(sheet_Name[7]), INT(FIND(TEXT(sheet_Name[7]), ' ')))
    sheet_Name['ФИО плательщика'] = PROPER(sheet_Name[9])
    sheet_Name['Пол плательщика'] = IF(TYPE(sheet_Name[23]) != 'NaN',
                                       SUBSTITUTE(SUBSTITUTE(sheet_Name[23], 'Мужской', 'М'), 'Женский', 'Ж'), None)
    # Фильтруем, что-бы убрать лишние строки
    CBRF_10361 = sheet_Name[(sheet_Name['Стоимость'].notnull()) & (
        sheet_Name['Стоимость'].apply(lambda val: all(s not in str(val) for s in ['COST', 'Стоимость'])))]
    # Меняем на нужный формат
    CBRF_10361['Стоимость'] = to_float_series(CBRF_10361['Стоимость'])
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df

def hks(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Добавляем новый колонки
    sheet_Name.insert(11, 'new-column-xrgl', 0)
    sheet_Name.insert(15, 'new-column-z6ia', 0)
    sheet_Name.insert(16, 'new-column-5syh', 0)
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={0: 'Номер заказа у партнера',
                               1: 'Номер сертификата',
                               4: 'Дата начала действия',
                               5: 'Дата окончания действия',
                               'new-column-xrgl': 'ФИО плательщика',
                               11: 'Дата рождения плательщика',
                               12: 'Пол плательщика',
                               'new-column-z6ia': 'Серия паспорта плательщика',
                               'new-column-5syh': 'Номер паспорта плательщика',
                               16: 'Продукт в системе партнера'}, inplace=True)
    # Применяем формулы
    sheet_Name['ФИО плательщика'] = IF(TYPE(sheet_Name[10]) != 'NaN',
                                       SUBSTITUTE(CONCAT(sheet_Name[8], ' ', sheet_Name[9], ' ', sheet_Name[10]), '  ',
                                                  ' '), None)
    sheet_Name['Серия паспорта плательщика'] = IF(TYPE(sheet_Name[13]) != 'NaN', LEFT(sheet_Name[13], 4), None)
    sheet_Name['Номер паспорта плательщика'] = IF(AND(TYPE(sheet_Name['Серия паспорта плательщика']) != 'NaN',
                                                      TYPE(sheet_Name[13]) != 'NaN'),
                                                  SUBSTITUTE(sheet_Name[13], sheet_Name['Серия паспорта плательщика'],
                                                             ''), None)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[
        (sheet_Name[8].notnull()) & (~sheet_Name[8].str.contains('Фамилия Застрахованного', na=False))]
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df

def arsenal_krim(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={0: 'Номер сертификата',
                               2: 'Продукт в системе партнера',
                               3: 'Стоимость',
                               4: 'Дата начала действия',
                               5: 'Дата окончания действия'}, inplace=True)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[(sheet_Name['Стоимость'].notnull()) & (~sheet_Name['Стоимость'].str.contains('Стоимость', na=False))]
    # Меняем на нужный формат
    sheet_Name['Стоимость'] = to_float_series(sheet_Name['Стоимость'])
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df

def skb(file_name):
    # Получаем имя листа
    wb = load_workbook(file_name, data_only=True)
    sheet = wb.sheetnames[0]
    # Инициализируем книгу
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet], skiprows=0, header=None)
    sheet_Name = sheet_df_dictonary[sheet]
    # Добавляем новый колонки
    sheet_Name.insert(4, 'new-column-8awm', 0)
    sheet_Name.insert(3, 'new-column-g96e', 0)
    # Меняем названия прошлых колонок
    sheet_Name.rename(columns={'new-column-g96e': 'Продукт в системе партнера',
                               'new-column-8awm': 'ФИО плательщика',
                               4: 'Номер телефона плательщика',
                               5: 'Номер сертификата',
                               6: 'Стоимость',
                               7: 'Дата начала действия',
                               8: 'Дата окончания действия',
                               9: 'Наименование ДО'}, inplace=True)
    # Применяем формулы
    sheet_Name['ФИО плательщика'] = IF(TYPE(sheet_Name[3]) != 'NaN', PROPER(sheet_Name[3]), None)
    sheet_Name['Продукт в системе партнера'] = IF(TYPE(sheet_Name[2]) != 'NaN', SUBSTITUTE(sheet_Name[2],
                                                                                           LEFT(sheet_Name[2],
                                                                                                INT(FIND(sheet_Name[2],
                                                                                                         '"') - 1)),
                                                                                           ''), None)
    # Фильтруем, что-бы убрать лишние строки
    sheet_Name = sheet_Name[
        (sheet_Name['ФИО плательщика'].notnull()) & (~sheet_Name['ФИО плательщика'].str.contains('Фио', na=False))]
    # Меняем на нужный формат
    sheet_Name['Стоимость'] = to_float_series(sheet_Name['Стоимость'])
    # Оборачиваем все в словарь, для дальнейшей работы с ним
    df = pd.DataFrame.to_dict(sheet_Name)
    # Возвращаем отморфированные данные
    return df


def data_placer(file_path, file_data_frame, file_name):
    # Присваиваем кастомное имя для результата
    # Шаблон заполнения файла
    shablon = ['Номер заказа у партнера', 'Номер сертификата', 'Продукт в системе партнера', 'Код продукта в BD',
               'Дата начала действия', 'Дата окончания действия', 'Стоимость', 'ФИО плательщика',
               'Дата рождения плательщика', 'Пол плательщика', 'Номер телефона плательщика',
               'Адрес электронной почты плательщика', 'Серия паспорта плательщика', 'Номер паспорта плательщика',
               'Кем выдан паспорт плательщика', 'Дата выдачи паспорта плательщика', 'Адрес плательщика',
               'Гражданство плательщика', 'Город', 'Банк', 'Наименование ДО']
    # Инициализирует новую книгу и активирует ее для дальнейшей работы
    wb = Workbook()
    ws = wb.active
    # Заполняет 1 строку по шаблону
    ws.append(shablon)
    # Находим нужное место для проставления данных по хедерам
    for data_name in list(file_data_frame):
        for number, name in enumerate(shablon):
            if data_name == name:
                # Получаем буквенную координату в книге
                car = get_column_letter(number + 1)
                # Запускам цикл заполнения столбца
                for row_number in range(2, len(file_data_frame[data_name]) + 2):
                    # Проставляем данные в книгу
                    ws[car + str(row_number)] = file_data_frame[data_name][
                        list(file_data_frame[data_name].keys())[row_number - 2]]
                break



    new_file_name = file_name + " morphed"
    wb.save(new_file_name)

    return new_file_name





def download_function(root_path):
    list_dir=list(y.listdir(root_path))
    path_list=[]
    list_of_file_path = []
    list_of_file_name = []
    list_of_dirs_name = []
    list_of_formated_file_name=[]

    for dir in list_dir:
        #path_list.append(list_dir[i].path)
        #if "Файлы для загрузки в Thales (после обработки)" not in list_dir[i].name:
        #print(list_dir[i])
        inner_dir=list(y.listdir(dir.path))
        for in_dir in inner_dir:

            if dir.name in in_dir.path and "Файлы для загрузки в Thales (после обработки)" not in in_dir.path:
                list_of_dirs_name.append(dir.name)
                list_of_file_name.append(in_dir.name)
                list_of_file_path.append(in_dir.path)
            #path_list.append(list_dir[i].path)
            #list_of_dirs_name.append(list_dir[i].name)
    #print(path_list)
   # for i in path_list:
    #    inner_dir = list(y.listdir(i))
   #     print(inner_dir)
    #    if "Файлы для загрузки в Thales (после обработки)" not in i:
   #         for j in inner_dir:
    #            list_of_file_path.append(j.path)
    #            list_of_file_name.append(j.name)


    for i in range(len(list_of_file_name)):
        y.download(list_of_file_path[i], list_of_file_name[i])

    for name in list_of_file_name:
        base=os.path.splitext(name)[0]
        os.rename(name,base+".xlsx")
        new_name=base+".xlsx"
        list_of_formated_file_name.append(new_name)

    return list_of_file_name,list_of_file_path,list_of_dirs_name


def upload_and_move_function(morphed_file_name,list_of_file_path,list_of_file_name):
    for name in morphed_file_name:
        #upload_path=morphed_path+str(now)+name
        upload_path = morphed_path + name + str(now)+".xlsx"
        y.upload(name,upload_path)
        os.remove(name)
    for name in list_of_file_name:
        os.remove(name)
    for i in range(len(list_of_file_name)):
        move_path=archive_path+list_of_file_name[i]+" archived "+str(now)
        y.move(list_of_file_path[i],move_path)



def main():

    list_of_file_name,list_of_file_path,list_dir_name=download_function(root_dir)
    morphed_file_name=[]
    print(list_of_file_name)
    print(list_of_file_path)
    print(list_dir_name)

    for i in range(len(list_of_file_name)):
        morpher=morpher_list(list_dir_name[i])
        if morpher != None:
            data_frame=morpher(list_of_file_name[i])
            new_file_name= data_placer(list_of_file_path[i],data_frame,list_of_file_name[i])
            morphed_file_name.append(new_file_name)
            print(morphed_file_name)
    upload_and_move_function(morphed_file_name,list_of_file_path,list_of_file_name)



main()
#data_frame = geb("GB.XLSX")
#file_path, new_file_name = data_placer(list_of_file_path[1], data_frame, list_of_file_name[1])
#print(file_path)
#print(new_file_name)
