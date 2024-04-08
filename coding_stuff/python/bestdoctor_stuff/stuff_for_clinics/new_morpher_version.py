from openpyxl import load_workbook, Workbook
from datetime import datetime
import re
import traceback
import pyexcel as p
from openpyxl.utils import get_column_letter
import os
import pandas as pd
from mitosheet import TYPE, IF, PROPER, AND, CONCAT, to_int_series, FIND, SUBSTITUTE, LEFT, INT, CLEAN, to_float_series, FLOAT, FILLNAN   # noqa: E501
import yadisk
import psycopg2
from typing import List

# flake8: noqa

ya_disk = yadisk.YaDisk('406760da5b1345a88999f0acb1ef95bf', '7ecb5577edb643d3a07ce020292db4b6',
                        'AQAEA7qkBXctAAfNlAAXrxr49UN8l0uBHNszaAo')
gp_conn = 'postgresql://zeppelin:R63v5NspNsSEem@c-c9qbht031ah0gtrlftmj.rw.mdb.yandexcloud.net:5432/warehouse'  # noqa: E501

headers = {
    'policy_number': ['Номер полиса', '№ полиса', 'Страховой полис', 'Полис', 'ID', ''''№ полиса или № по базе''',  # noqa
                      'POLICY', '№ полиса ДМС', '№ полиса или № по базе', '№ полиса (карточки)', '№\nполиса',  # noqa: E501
                      'Данные полиса', 'Номер полиса ', 'Полис №', 'ID карта пациента', '№Полиса', 'Номер \nполиса',  # noqa: E501
                      '№ полиса ДМС', '№ Полиса ДМС', 'Номер полиса \nДМС\n', 'ПОЛИС', '№ Страховой полис',  # noqa: E501
                      'Номер полиса ДМС', 'Номер страхового полиса', '№ Полиса', 'КОД УСЛУГИ'],
    'fio': ['ФИО', 'Пациент', 'ФИО пациента', 'ФИО застрахованного', 'Ф.И.О.', 'гнитно', '\nФИО',
            ''''Ф.И.О., возраст (лет)''', 'Фамилия имя Отчество', 'ФИО                   пациента',  # noqa
            'Пациент (ФИО)', 'Ф.И.О., возраст (лет) ', 'Ф.И.О. застрахованных', 'Фамилия И.О. пациента',  # noqa: E501
            'Фамилия, имя, отчество', 'ФИО \nзастрахованных (по алфавиту)\n', 'Фамилия, Имя, отчество',  # noqa: E501
            'Застрахованного', 'Фамилия И. О. № полиса', 'Фамилия И.О. \n№ полиса', 'ФИО \nзастрахованного',  # noqa: E501
            'ФИО\nпациента', 'ФИО Застрахованного', 'Ф.И.О. ', 'ФИО застрахованного (полностью)',
            'Фамилия Имя Отчество застрахованного', 'ФИО застрахован. ', 'Ф.И.О. пациента', 'Фамилия Имя Отчество',  # noqa: E501
            'ФИО Застрахованного (в одну ячейку; делить на три раздельных ячейки не надо)', 'ФИО Пациента',  # noqa: E501
            ],
    'guarantee_letter': ['№ ГП', 'ГП', '№ гарантийного письма ', 'Номер ГП', 'номер гарантийного письма ',  # noqa: E501
                         '№ и дата гарантийного письма', 'Номер и дата гарантийного письма', 'Дата и номер ГП',  # noqa: E501
                         'дата/номер гарантийного письма ', '№ гарантийного письма', 'Гарантийное письмо',  # noqa: E501
                         'Гарантийное \nписьмо \nномер', 'НОМЕР ГП', 'Номер гарантийного письма'],
    'first_name': ['Имя', 'P_NAME', 'Имя Застрахованного', 'Имя ', 'Номер гп'],
    'last_name': ['Фамилия', 'P_FAMILY', 'Фамилия Застрахованного', 'Фамилия '],
    'middle_name': ['Отчество', 'P_PATRONIMIC', 'Отчество Застрахованного ', 'Отчество '],
    'date': ['Дата', 'Дата услуги', 'Дата начала оказания услуги', ''''Дата предоставления услуги''',  # noqa
             'Дата обращения', 'Дата и время\nначала', 'Начало', 'Дата окачания услуги', 'Дата оказания услуг',  # noqa: E501
             'Дата обр.', 'DATA', 'Дата приема', 'Дата оказания услуги', 'дата', 'Дата предоставления услуги',  # noqa: E501
             'Дата взятия пробы', 'Дата оказания медицинской услуги', 'Дата\nначала услуги', 'дата оказания',  # noqa: E501
             'Дата оказания услуги ', 'Дата оказания медицинской услуги ', 'оказания', 'Дата оказания',  # noqa: E501
             'ДАТА ОКАЗАНИЯ УСЛУГИ ', 'Дата поступления (госпитализация)', 'Дата начала и окончания госпитализации',  # noqa: E501
             'Дата \nоказания \nуслуги', 'дата оказания услуги', 'ДАТА ОКАЗАНИЯ УСЛУГИ', 'Дата начала оказания услуги'],  # noqa: E501
    'end_date': ['Дата окончания оказания услуги', 'Дата выполнения', 'Дата выезда', 'Дата окончания', 'Дата выписки',  # noqa: E501
                 'Дата и время\nвозвращения', 'Дата окончания услуги'],
    'doctor_name': ['Доктор', 'Врач', 'Ф.И.О. доктора ', 'Исполнитель', 'ФИО врача, оказавшего услугу', 'ФИО врача',  # noqa: E501
                    '\nФИО врача', 'Врач                              (ФИО)', 'ФИО доктора', 'Ф.И.О. Врача',  # noqa: E501
                    'Код врача (или ФИО)', 'ФИО лечащего врача', 'ВРАЧ', 'Врач (ФИО)'],
    'doctor_last_name': ['D_FAMILY'],
    'doctor_first_name': ['D_NAME'],
    'doctor_middle_name': ['D_PATRONIMIC'],
    'tooth_number': ['№ зуба', 'Зуб №', 'Номер зуба (для стоматологических услуг)', 'Зуб', 'Зуб - номер', '№ Зуба',  # noqa: E501
                     'Номер зуба', '№ зуба (по международной нумерации)', '№ зуба', '№ ЗУБА',
                     '№ зуба (по междун. классиф.)'],
    'mkb': ['МКБ-10', 'Код по МКБ-10', 'Код по МКБ', 'Код МКБ', 'Код МКБ-10', 'Код диагноза по МКБ10', 'МКБ',  # noqa: E501
            'Шифр заболевания по МКБ-10', 'Код диагноза по МКБ-Х или развернутый диагноз', 'Код диагноза по МКБ-Х',  # noqa: E501
            'Диагноза по МКБ-Х', 'МКБ10', 'Код диагноза по МКБ10 (не менее четырех знаков)', 'Код диагноза\n (МКБ-10)',  # noqa: E501
            'Код диагноза по МКБ-Х или развернутый диагноз ', 'Код\nдиагн.', 'Кол.\nУслуг', 'КОД ДИАГНОЗА ПО МКБ-10',  # noqa: E501
            'Код диагноза по МКБ 10  и стомат. диагноз', 'Код диагноза по МКБ-10', 'Код диагноза по МКБ 10 и диагноз', ],  # noqa
    'service_code': ['Код услуги', 'Код услуги по прейск.ЛПУ', 'Код', ''''Код услуги''', 'SRV_CODE', 'Код Услуги ',  # noqa
                     'Код мед. услуги по прейскуранту', 'Код услуги по Прейскуранту', 'Код\nуслуги', 'Код  услуги',  # noqa: E501
                     'Код МЗ', 'КОД УСЛУГИ', 'Код услуги по прейскуранту', 'Код\n услуги', 'Код услуги '  # noqa: E501
                     'Код услуги по прейскуранту ЛПУ', 'Код медицинской услуги', 'Код услуги (Прейскурант)',  # noqa: E501
                     'Код услуги по прейск', 'Код услуги ', 'Код Услуги'],
    'service_name': ['Наименование услуги', 'Название услуги', 'Название мед.услуги', '\nНаименование услуги',  # noqa: E501
                     'Наименование медицинской услуги', ''''Наименование услуги''',  # noqa
                     'Наименование выполненной услуги', 'Полное наименование услуги по прейскуранту',  # noqa: E501
                     'Услуга', 'SRV_NAME', 'Оказанные услуги', 'Наименование мед. услуги', 'Анализ / Профиль',  # noqa: E501
                     'Наименование медицинской услуги или ее кода по Прейскуранту', 'наименование услуги',  # noqa: E501
                     'Наименование услуги (код по Перечню услуг)', 'Название услуг', 'Наименование \nуслуги',  # noqa: E501
                     'Наименование медицинской услуги по Прейскуранту', 'Код и название мед. услуги',  # noqa: E501
                     'Наименование услуги подменная', 'Наименование медицинской услуги по Прейскуранту',  # noqa: E501
                     'оказанных услуг', 'Выполненные исследования', 'НАИМЕНОВАНИЕ УСЛУГИ', 'Наименование услуги '],  # noqa: E501
    'service_price': ['Цена, руб.', 'Стоимость', 'Стоимость (руб.)', 'Цена', 'Цена (руб)', 'Цена услуги', 'Цена услуги ',  # noqa: E501
                      ''''Стоимость услуги''', 'SRV_COST', 'Цена по прейскуранту', 'Цена 1-й услуги, руб.',  # noqa
                      'Стоимость мед. услуги (руб.)', 'Стоимость услуги', 'Стоимость 1 услуги руб.', 'Цена ', 'Сумма',  # noqa: E501
                      'Цена услуги по прейскуранту  ', 'Стоимость \nуслуги\n', 'Стоим. за ед.', 'Стоим. за\nед.',  # noqa: E501
                      'Общая стоим.', 'ЦЕНА', 'Стоимость\nруб.', 'Цена     услуги', 'Цена услуги ', 'Цена\nуслуги',  # noqa: E501
                      'Цена за единицу', 'Цена, \nруб.', 'цена*кол-во (стоимость)'],
    'service_amount': ['Кол-во', 'Пр.', 'Количество', ''''Кол-во услуг''', 'Кол-во услуг', 'SRV_TOTAL', 'кол-во услуг',  # noqa
                       'Кол-во мед. услуг', 'Количество услуг', 'К-во услуг', 'Кол-во', 'Кол-во услуг  ',  # noqa: E501
                       'Количество \nуслуг\n', 'услуг', 'Кол. услуг', 'Кол. Услуг', 'КОЛ.  УСЛУГ',  # noqa: E501
                       'Количество оказанных услуг', 'Кол-во ', 'кол-во', 'Кол- во', 'КОЛ-ВО', 'Кол-во оказанных услуг'],  # noqa: E501
    'total_price': ['Сумма, руб.', 'Сумма к оплате', 'Стоимость (руб.)', 'Стоимость', ''''Общая стоимость''',  # noqa
                    'Сумма рублей', 'Стоимость услуг по прейскуранту без учета скидки (руб)', 'SRV_SUMM',  # noqa: E501
                    'Начислено', 'Сумма к оплате, руб.', 'Полная стоимость', 'Ст-ть услуг (руб.)', 'стоимость',  # noqa: E501
                    'Итого (руб.)', 'Общая стоимость', 'Общая сумма руб.', 'Общая Сумма', 'Ст-ть',  # noqa: E501
                    'Стоимость руб.', 'услуги,', 'СТОИМОСТЬ', 'Стоимость услуг (Цена*Количество)',  # noqa: E501
                    'Цена (руб) * кол-во услуг', 'Стоимость, руб.', 'Полная стоимость'],  # noqa: E501
    'payment_type': ['Тип оплаты'],
    'discount_sice': ['% Скидка', 'Франшиза', 'Спецсект (коэф.)', 'Коэф', 'Скидка, %', 'Скидка или надбавка, %',  # noqa: E501
                      'Скидка', 'СКИДКА'],
    'diagnosis': ['Диагноз', ''''Диагноз''', 'SRV_DIAG', 'Код диагноза', 'Диагноз клинический (текст)',  # noqa
                  'Диагноз клинический', 'Код диагн.', 'ДИАГНОЗ (текст)', 'Диагноз,\n№ зуба', 'Диагноз (по МКБ 10)',  # noqa: E501
                  'Диагноз, номер зуба', 'Диагноз (код по МКБ)', 'Диагноз или код (по МКБ-10)', 'Диaгноз'],  # noqa: E501
    'clinic_code': ['Код клиники', 'Код филиала клиники (точки)'],
    'clinic_name': ['Клиника-Исполнитель (при Сети Клиник)', 'Наименование филиала клиники', 'Наименование филиала клиники (точки)'],  # noqa: E501
    'doctor_code': ['Код врача', 'Код врача, оказавшего услугу ', 'Код врача'],  # noqa: E501
    'doctor_speciality': ['Специальность доктора', 'Специальность врача назначившего обследование/лечение',  # noqa: E501
                          'Специаль-ность врача', 'Специальность врача'],
    'doctor_speciality_2': ['Специальность направившего врача', 'Специальность врача оказавшего услугу  ', 'Специальность отправителя'],  # noqa: E501
    'branch_code': ['Номер филиала', 'Код отделения'],
    'branch_name': ['Отделение ', 'Филиал', 'Учреждение', 'Отделение', 'Подразделение'
                    'Отделение (поликлинический прием, лаборатория, функционалой диагностики, физиотерапия, ПНД, процедурный кабинет, дневной стационар, стоматология)',  # noqa: E501
                    'Отделение (поликлинический прием, лаборатория, функционалой диагностики, физиотерапия, ПНД, процедурный кабинет, дневной стационар, стоматология)',  # noqa: E501
                    'Отделение / Специальность', 'Наименование филиала', 'Наименование отделения'],
    'number_disease_history': ['№ истории болезни', 'Номер медицинской карты /истории болезни', 'Номер мед.карты'],  # noqa: E501
    'price-list_id': ['ID прейскуранта'],
    'service_type': ['Тип обслуживания / Вид помощи', 'Код вида обращения'],
    'medical_service': ['Медицинская услуга'],
}
shablon = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги', 'Дата окончания оказания услуги',  # noqa: E501
           'Код услуги', 'Наименование услуги', 'Код МКБ-10', 'Диагноз', 'Номер зуба (для стоматологических услуг)',  # noqa: E501
           'Цена услуги', 'Количество', 'Скидка, %', 'Код филиала клиники (точки)',
           'Наименование филиала клиники (точки)', '№ ГП', 'Код врача', 'Врач (ФИО)', 'Специальность врача',  # noqa: E501
           'Специальность направившего врача', 'Код отделения', 'Наименование отделения', '№ истории болезни',  # noqa: E501
           'ID Прейскуранта', 'Тип оплаты', 'Тип обслуживания / Вид помощи', 'Медицинская услуга']
LOCAL_HEADERS = {
    'policy_number': ['Страховой полис'],
    'fio': ['ФИО пациента'],
    'guarantee_letter': ['№ ГП'],
    'date': ['Дата начала оказания услуги'],
    'end_date': ['Дата окончания оказания услуги'],
    'doctor_name': ['Врач (ФИО)'],
    'tooth_number': ['Номер зуба (для стоматологических услуг)'],
    'mkb': ['Код МКБ-10', 'МКБ10'],
    'service_code': ['Код услуги'],
    'service_name': ['Наименование услуги'],
    'service_price': ['Цена услуги'],
    'service_amount': ['Количество'],
    'total_price': ['Сумма, руб.'],
    'discount_sice': ['Скидка, %'],
    'payment_type': ['Тип оплаты'],
    'diagnosis': ['Диагноз'],
    'clinic_code': ['Код филиала клиники (точки)'],
    'clinic_name': ['Наименование филиала клиники (точки)'],
    'doctor_code': ['Код врача'],
    'doctor_speciality': ['Специальность врача'],
    'doctor_speciality_2': ['Специальность направившего врача'],
    'branch_code': ['Код отделения'],
    'branch_name': ['Наименование отделения'],
    'number_disease_history': ['№ истории болезни'],
    'price-list_id': ['ID Прейскуранта'],
    'service_type': ['Тип обслуживания / Вид помощи'],
    'medical_service': ['Медицинская услуга'],
}


class DwsConn:
    def __init__(self, conn_string):
        self.conn_string = conn_string

    def select(self, query: str) -> List[tuple]:
        with psycopg2.connect(self.conn_string) as conn:
            with conn.cursor() as cursor:
                cursor.execute(query)
                return cursor.fetchall()

    def execute(self, query: str) -> None:
        with psycopg2.connect(self.conn_string) as conn:
            with conn.cursor() as cursor:
                cursor.execute(query)
                conn.commit()


def data_pandas_rzd(file_name, file_path):  # noqa: CFQ001
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if sheet_name == 'Счет, с детализацией по услугам':
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,  # noqa: E501
                                           header=None)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        # Added columns
        Sheet_Name_Here.insert(1, 'new-column-t4si', 0)
        Sheet_Name_Here.insert(2, 'new-column-28w5', 0)
        Sheet_Name_Here.insert(3, 'new-column-lwtq', 0)
        Sheet_Name_Here.insert(4, 'new-column-kvwo', 0)
        Sheet_Name_Here.insert(5, 'new-column-cy5g', 0)
        Sheet_Name_Here.insert(6, 'new-column-upvc', 0)
        # Renamed columns
        Sheet_Name_Here.rename(columns={2: 'Кол-во',
                                        4: 'Дата услуги',
                                        5: 'Код врача',
                                        8: '№ зуба',
                                        'new-column-28w5': 'ФИО',
                                        'new-column-t4si': 'Номер полиса',
                                        'new-column-lwtq': 'Код услуги',
                                        'new-column-kvwo': 'Наименование услуги',
                                        'new-column-cy5g': 'Цена',
                                        'new-column-upvc': 'МКБ'}, inplace=True)
        # Set formula
        Sheet_Name_Here['ФИО'] = IF(FIND(Sheet_Name_Here[1], 'Клиент') > 0,
                                    SUBSTITUTE(Sheet_Name_Here[1], LEFT(Sheet_Name_Here[1],
                                                                        INT(FIND(Sheet_Name_Here[1], ':') + 2)), ''),  # noqa: E501
                                    None)
        Sheet_Name_Here['Номер полиса'] = IF(FIND(Sheet_Name_Here['Дата услуги'], 'Полис') > 0,
                                             SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Дата услуги']),  # noqa: E501
                                                                   ':', ''), ' ', ''), None)
        Sheet_Name_Here['Код услуги'] = IF(FIND(Sheet_Name_Here[1], '.') > 0, LEFT(Sheet_Name_Here[1],  # noqa: E501
                                                                                   INT(FIND(Sheet_Name_Here[1],  # noqa: E501
                                                                                            ' ')) - 1),  # noqa: E501
                                           None)
        Sheet_Name_Here['Наименование услуги'] = SUBSTITUTE(Sheet_Name_Here[1], Sheet_Name_Here['Код услуги'],  # noqa: E501
                                                            '')
        Sheet_Name_Here['МКБ'] = IF(AND(TYPE(Sheet_Name_Here[6]) != 'NaN', TYPE(Sheet_Name_Here[7]) != 'NaN'),  # noqa: E501
                                    # noqa: E501
                                    CONCAT(Sheet_Name_Here[6], ' ', Sheet_Name_Here[7]),
                                    Sheet_Name_Here[6])
        # Filled NaN values columns
        columns_to_fill_nan = ['ФИО', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(
            method='ffill')
        # Filtered
        Sheet_Name_Here = Sheet_Name_Here[
            (Sheet_Name_Here['Кол-во'].notnull()) & (
                ~Sheet_Name_Here['Кол-во'].str.contains('К-во', na=False))]  # noqa: E501
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата услуги'].notnull()]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here[3].notnull()]
        # Changed dtype
        Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
        # Set formula
        Sheet_Name_Here['Цена'] = INT(SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here[3], ' ', ''), ',', '.')) / Sheet_Name_Here[  # noqa: E501
            'Кол-во']  # noqa: E501
        # Changed dtype
        Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,  # noqa: E501
                                           header=None)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        # Renamed columns
        Sheet_Name_Here.rename(columns={1: 'ФИО',
                                        3: 'Номер полиса',
                                        4: 'Код услуги',
                                        5: 'Наименование услуги',
                                        6: 'Врач',
                                        7: 'Кол-во',
                                        8: 'Цена',
                                        9: 'Дата',
                                        10: 'МКБ'}, inplace=True)
        # Filtered
        Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull()) & (
            ~Sheet_Name_Here['Кол-во'].str.contains('Кол-во мед. услуг', na=False))]

        # Changed dtype
        Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
        Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def sphera_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,  # noqa: E501
                                       header=None)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(7, 'new-column-kexs', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={0: 'ФИО',
                                    1: 'Номер полиса',
                                    'new-column-kexs': 'Кол-во',
                                    2: 'Дата начала оказания услуги',
                                    3: 'Дата окончания оказания услуги',
                                    10: 'Цена услуги'}, inplace=True)
    # Set formula
    Sheet_Name_Here['Кол-во'] = IF(TYPE(Sheet_Name_Here[6]) != 'NaN', 1, None)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Дата окончания оказания услуги'].notnull()) & (  # noqa: E501
        ~Sheet_Name_Here['Дата окончания оказания услуги'].str.contains('Окончание', na=False))]
    # Changed dtype
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def data_national_medical_service(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    Sheet_Name_Here.rename(columns={'Дата талона': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Сумма в счете': 'Цена'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def data_pandas_k_31(file_name, file_path):  # noqa: CFQ001
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,  # noqa: E501
                                       header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    Sheet_Name_Here.insert(2, 'new-column-fl6j', 0)

    # Added column new-column-8gcf
    Sheet_Name_Here.insert(21, 'new-column-8gcf', 0)

    # Set formula of new-column-fl6j
    Sheet_Name_Here['new-column-fl6j'] = IF(TYPE(Sheet_Name_Here[1]) != 'NaN', PROPER(Sheet_Name_Here[1]),  # noqa: E501
                                            None)  # noqa: E501

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-fl6j': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={2: 'Номер полиса'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={4: 'Дата услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={5: 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={6: 'Наименование услуги'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={7: 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={8: 'Цена'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={14: 'МКБ'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={15: 'Диагноз'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={16: '№ зуба'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'new-column-8gcf': 'Врач'}, inplace=True)

    # Renamed columns Код врача
    Sheet_Name_Here.rename(columns={18: 'Код врача'}, inplace=True)

    # Renamed columns Клиника-Исполнитель (при Сети Клиник)
    Sheet_Name_Here.rename(columns={21: 'Клиника-Исполнитель (при Сети Клиник)'}, inplace=True)

    # Renamed columns Отделение / Специальность
    Sheet_Name_Here.rename(columns={17: 'Отделение / Специальность'}, inplace=True)

    # Set formula of Врач
    Sheet_Name_Here['Врач'] = IF(TYPE(Sheet_Name_Here[19]) != 'NaN', PROPER(Sheet_Name_Here[19]), None)  # noqa: E501

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull()) & (
        ~Sheet_Name_Here['Кол-во'].str.contains('Кол-во', na=False))]  # noqa: E501

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

    # Changed Номер полиса to dtype int

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)  # Возможно надо условием чистить номер полиса


def gbuz_eramishanceva(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2,  # noqa: E501
                                       converters={'Цена': str})  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)
    Sheet_Name_Here.insert(2, 'new-column-p6i7', 0)
    Sheet_Name_Here.insert(2, 'new-column-yd2l', 0)
    Sheet_Name_Here.rename(columns={'new-column-yd2l': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-p6i7': 'Дата окончания услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата': 'Пропуск'}, inplace=True)
    Sheet_Name_Here['Дата услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', LEFT(Sheet_Name_Here['Пропуск'],  # noqa: E501
                                                                                        INT(FIND(
                                                                                            Sheet_Name_Here['Пропуск'],  # noqa: E501
                                                                                            # noqa: E501
                                                                                            '-') - 1)), None)  # noqa: E501
    Sheet_Name_Here['Дата окончания услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                                  SUBSTITUTE(Sheet_Name_Here['Пропуск'],
                                                             LEFT(Sheet_Name_Here['Пропуск'],
                                                                  INT(FIND(Sheet_Name_Here['Пропуск'], '-'))), ''),  # noqa: E501
                                                  # noqa: E501
                                                  None)
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Фамилия И.О. пациента': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Стоимость': 'Пропуск2'}, inplace=True)
    columns_to_fill_nan = ['Пропуск', 'Дата услуги', 'Дата окончания услуги', 'ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = Sheet_Name_Here['Цена'].fillna(0).astype('float')
    Sheet_Name_Here['Номер полиса'] = Sheet_Name_Here['Номер полиса'].fillna(0).astype('int')
    Sheet_Name_Here.insert(8, 'new-column-8z8r', 0)
    Sheet_Name_Here['new-column-8z8r'] = FLOAT(Sheet_Name_Here['Цена']) * 1
    Sheet_Name_Here.rename(columns={'Цена': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-8z8r': 'Цена'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def guta_clinic(file_name, file_path):  # noqa: CCR001, CFQ001, C901
    stop = False
    patient = False
    policy = None
    fio = None
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    for row_number in range(4, sheet.max_row + 1):
        if stop:
            break
        for column_number in range(1, sheet.max_column + 1):
            inform_value = sheet.cell(row=row_number, column=column_number).value
            if 'Общество с ограниченной ответственностью «ГУТА-КЛИНИК»' in str(inform_value):
                stop = True
                break
            if 'Итого по пациенту' in str(inform_value):
                patient = False
                break
            if sheet.cell(row=row_number, column=14).value is None:
                policy = sheet.cell(row=row_number, column=column_number).value.split(' ', 4)[0]
                fio = sheet.cell(row=row_number, column=column_number).value.split(' ', 4)[-1]
                patient = True
                break
            if patient and column_number == 1:
                data['fio'].append(fio)
                data['policy_number'].append(policy)
                data['end_date'].append(None)
            elif column_number == 2:
                data['date'].append(inform_value)
            elif column_number == 3:
                data['service_code'].append(inform_value)
            elif column_number == 6:
                data['service_name'].append(inform_value)
            elif column_number == 11:
                data['tooth_number'].append(inform_value)
            elif column_number == 12:
                data['diagnosis'].append(inform_value)
            elif column_number == 13:
                data['service_price'].append(inform_value)
            elif column_number == 14:
                data['service_amount'].append(inform_value)
            elif column_number == 15:
                discount = str(int(sheet.cell(row=row_number, column=column_number).value) / int(sheet.cell(row=row_number, column=13).value) * 100 * -1) + '%'  # noqa: E501
                data['discount_sice'].append(discount)
            elif column_number == 17:
                break
    return filling_book(data, file_name, file_path)


def chaika(file_name, file_path):  # noqa: CFQ001
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5,  # noqa: E501
                                       converters={'Цена услуги': str, 'Количество': str})  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Added column new-column-u22a
    Sheet_Name_Here.insert(4, 'new-column-u22a', 0)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Пациент': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Франшиза': 'Пропуск123'}, inplace=True)
    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-u22a': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', PROPER(Sheet_Name_Here['Пропуск']),  # noqa: E501
                                None)  # noqa: E501

    # Renamed columns Пропуск2
    Sheet_Name_Here.rename(columns={'Цена услуги': 'Пропуск2'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Количество': 'Кол-во'}, inplace=True)

    # Added column new-column-47q2
    Sheet_Name_Here.insert(15, 'new-column-47q2', 0)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-47q2': 'ГП'}, inplace=True)

    # Renamed columns Пропуск3
    Sheet_Name_Here.rename(columns={'№ ГП': 'Пропуск3'}, inplace=True)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = IF(TYPE(Sheet_Name_Here['Пропуск3']) != 'NaN',
                               SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск3']), '№', ''), ' ', ''), None)  # noqa: E501

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'Доктор': 'Врач'}, inplace=True)

    # Renamed columns Специальность врача
    Sheet_Name_Here.rename(columns={'Специальность доктора': 'Специальность врача'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'МКБ-10': 'МКБ'}, inplace=True)

    # Added column new-column-o8uy
    Sheet_Name_Here.insert(14, 'new-column-o8uy', 0)

    # Renamed columns Скидка
    Sheet_Name_Here.rename(columns={'new-column-o8uy': 'Скидка'}, inplace=True)

    # Set formula of Скидка
    Sheet_Name_Here['Скидка'] = IF(FIND(Sheet_Name_Here['Кол-во'], 'Скидка, %') > 0, Sheet_Name_Here['К оплате'],  # noqa: E501
                                   None)  # noqa: E501

    # Filled NaN values in 1 columns in Sheet1
    columns_to_fill_nan = ['Скидка']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='bfill')  # noqa: E501

    # Added column new-column-1jy9
    Sheet_Name_Here.insert(11, 'new-column-1jy9', 0)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-1jy9': 'Цена'}, inplace=True)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN', TYPE(Sheet_Name_Here['Скидка']) != 'NaN'),
        FLOAT(Sheet_Name_Here['Пропуск2']) - (
                    FLOAT(Sheet_Name_Here['Пропуск2']) / 100 * INT(Sheet_Name_Here['Скидка'])), None)  # noqa: E501, E126

    # Filtered Цена
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    Sheet_Name_Here['Скидка'] = [int(float(x.replace(',', '.'))) for x in Sheet_Name_Here['Скидка']]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vek_21(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Ст-ть': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-во ': 'Кол-во'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medgard_samara_toliati(file_name, file_path):  # noqa: CCR001, CFQ001, C901
    stop = False
    patient = False
    policy = None
    fio = None
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    for row_number in range(8, sheet.max_row + 1):
        if stop:
            break
        for column_number in range(2, sheet.max_column + 1):
            inform_value = sheet.cell(row=row_number, column=column_number).value
            if 'ООО "Бестдоктор"' in str(inform_value):
                break
            if 'В ИТОГЕ' in str(inform_value):
                number_of_discount = sheet.cell(row=row_number + 1, column=2).value.split('В ИТОГЕ К ОПЛАТЕ СО СКИДКОЙ ', 2)[-1].replace('%', '')  # noqa: E501
                data['discount_sice'] = [number_of_discount for x in range(0, len(data['fio']))]
                data['']
                stop = True
                break
            if 'Итого по пациенту' in str(sheet.cell(row=row_number, column=2).value):
                patient = False
                break
            if sheet.cell(row=row_number, column=7).value is None:
                fio_1 = inform_value.split('  ', 2)[1]
                fio = fio_1.split(' ', 3)[0] + ' ' + fio_1.split(' ', 3)[1] + ' ' + fio_1.split(' ', 3)[2]  # noqa: E501
                policy = re.sub('^0{2}|^0{3}|^0{4}|^0{5}', '', fio_1.split(' ', 3)[-1])
                patient = True
                break
            if patient and column_number == 3:
                data['fio'].append(fio)
                data['policy_number'].append(policy)
                data['date'].append(str(inform_value).replace(' 0:00:00', ''))
            elif column_number == 4:
                data['service_code'].append(inform_value)
            elif column_number == 5:
                data['service_name'].append(inform_value)
            elif column_number == 6:
                data['service_price'].append(inform_value)
            elif column_number == 7:
                data['service_amount'].append(inform_value)
            elif column_number == 11:
                data['mkb'].append(inform_value)
    return filling_book(data, file_name, file_path)


def medgard_nevskii(file_name, file_path):  # noqa: CCR001, CFQ001, C901
    start = False
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    for row_number in range(10, sheet.max_row + 1):
        for column_number in range(2, sheet.max_column + 1):
            inform_value = sheet.cell(row=row_number, column=column_number).value
            if 'ИТОГО' in str(sheet.cell(row=row_number, column=1).value):
                start = False
                break
            if 'ФИО Застрахованного (полностью)' in str(sheet.cell(row=row_number, column=2).value):
                start = True
                doctor = None
                fio = None
                patient_policy = None
                date = None
                pre_service_value = None
                service_name = None
                break
            if start:
                if column_number == 2:
                    if sheet.cell(row=row_number, column=column_number - 1).value is not None:
                        tooth_number = None
                        if inform_value is None:
                            last_name = ''
                        else:
                            last_name = inform_value
                        if sheet.cell(row=row_number + 1, column=column_number).value is None:
                            first_name = ''
                        else:
                            first_name = sheet.cell(row=row_number + 1, column=column_number).value
                        if sheet.cell(row=row_number + 2, column=column_number).value is None:
                            middle_name = ''
                        else:
                            middle_name = sheet.cell(row=row_number + 2, column=column_number).value
                        fio = last_name + ' ' + first_name + ' ' + middle_name
                        continue
                elif column_number == 3:
                    if inform_value is not None:
                        patient_policy = inform_value
                        continue
                elif column_number == 4:
                    if inform_value is not None:
                        date = inform_value
                        continue
                elif column_number == 5:
                    if inform_value is not None:
                        if sheet.cell(row=row_number, column=column_number - 1).value is not None:
                            for i in range(0, 100):
                                if sheet.cell(row=row_number + i, column=column_number).value is not None:  # noqa: E501
                                    if tooth_number is None:
                                        tooth_number = str(
                                            sheet.cell(row=row_number + i, column=column_number).value)  # noqa: E501
                                    else:
                                        tooth_number += ', '
                                        tooth_number += str(
                                            sheet.cell(row=row_number + i, column=column_number).value)  # noqa: E501
                                else:
                                    break
                        elif sheet.cell(row=row_number, column=8).value is not None and tooth_number is None:  # noqa: E501
                            for i in range(0, 100):
                                if sheet.cell(row=row_number + i, column=column_number).value is not None:  # noqa: E501
                                    if tooth_number is None:
                                        tooth_number = str(
                                            sheet.cell(row=row_number + i, column=column_number).value)  # noqa: E501
                                    else:
                                        tooth_number += ', '
                                        tooth_number += str(
                                            sheet.cell(row=row_number + i, column=column_number).value)  # noqa: E501
                                else:
                                    break

                elif column_number == 7:
                    if sheet.cell(row=row_number, column=8).value is not None:
                        data['mkb'].append(inform_value)
                        continue
                    if doctor is None:
                        for i in range(0, 100):
                            if sheet.cell(row=row_number + i, column=column_number).value is not None and 'Врач:' in str(sheet.cell(row=row_number + i, column=column_number).value):  # noqa: E501
                                doctor = str(
                                    sheet.cell(row=row_number + i, column=column_number).value.replace('Врач:',  # noqa: E501
                                                                                                       ''))  # noqa: E501
                                break
                elif column_number == 8:
                    if inform_value is not None:
                        data['service_code'].append(inform_value)
                        continue
                elif column_number == 9:
                    if sheet.cell(row=row_number, column=column_number - 1).value is not None:
                        if pre_service_value is not None:
                            data['service_name'].append(service_name)
                        pre_service_value = inform_value
                        service_name = pre_service_value
                    else:
                        if inform_value is None and pre_service_value is not None:
                            data['service_name'].append(service_name)
                            pre_service_value = inform_value
                            continue
                        elif str(pre_service_value) in str(service_name) and pre_service_value is not None:  # noqa: E501
                            pre_service_value += inform_value
                            service_name = pre_service_value
                        else:
                            pre_service_value = inform_value
                elif column_number == 10:
                    if inform_value is not None:
                        data['service_amount'].append(inform_value)
                elif column_number == 11:
                    if inform_value is not None and str(inform_value) != ' ':
                        data['service_price'].append(inform_value)
                        data['tooth_number'].append(tooth_number)
                        data['fio'].append(fio)
                        data['policy_number'].append(patient_policy)
                        data['date'].append(date)
                        data['doctor_name'].append(doctor)

    return filling_book(data, file_name, file_path)


def smart_clinic(file_name, file_path):  # noqa: CFQ001
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7,  # noqa: E501
                                       converters={'Скидка, %': str})  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'ФИО пациента': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'ID карта пациента': 'Номер полиса'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Цена\nуслуги': 'Пропуск'}, inplace=True)

    # Added column new-column-qoop
    Sheet_Name_Here.insert(8, 'new-column-qoop', 0)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-qoop': 'Цена'}, inplace=True)

    # Renamed columns Пропуск22
    Sheet_Name_Here.rename(columns={'Стоимость': 'Пропуск22'}, inplace=True)

    # Renamed columns Пропускк3333
    Sheet_Name_Here.rename(columns={'Стоимость со скидкой': 'Пропускк3333'}, inplace=True)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'№ ГП': 'ГП'}, inplace=True)

    # Renamed columns Скидка
    Sheet_Name_Here.rename(columns={'Скидка, %': 'Скидка'}, inplace=True)

    # Added column new-column-2592
    Sheet_Name_Here.insert(9, 'new-column-2592', 0)

    # Set formula of new-column-2592
    Sheet_Name_Here['new-column-2592'] = IF(TYPE(Sheet_Name_Here['Скидка']) != 'NaN',
                                            FLOAT(Sheet_Name_Here['Пропуск']) - FLOAT(  # noqa: E501
                                                (FLOAT(Sheet_Name_Here['Пропуск']) / 100 * FLOAT(
                                                    Sheet_Name_Here['Скидка']))), None)

    # Added column new-column-y330
    Sheet_Name_Here.insert(9, 'new-column-y330', 0)

    # Set formula of new-column-y330
    Sheet_Name_Here['new-column-y330'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                            FLOAT(SUBSTITUTE(Sheet_Name_Here['Пропуск'], ' ', '')),  # noqa: E501
                                            None)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = FLOAT(FILLNAN(Sheet_Name_Here['new-column-2592'], Sheet_Name_Here['new-column-y330']))  # noqa: E501

    # Filled NaN values in 3 columns in Лист1
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def udina(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        columns_to_fill_nan = ['№', 'Фамилия, имя, отчество', '№ полиса ДМС', 'Номер карты']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:  # noqa: B001, E722
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def data_pandas_son_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)  # noqa: E501
    Sheet1 = sheet_df_dictonary[sheet_name]
    Sheet1 = Sheet1[Sheet1['Кол-во'].notnull()]
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата']
    Sheet1[columns_to_fill_nan] = Sheet1[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet1)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def alfasrahovanie_moscow(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№п/п'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def alfasrahovanie_rostov_na_donu(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№п/п'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    repair_dict = df.to_dict()
    repair_dict['Страховой полис'] = {x: re.sub('^0{3}|^0{4}|^0{2}|^0{5}', '', str(value)).replace(' ', '') for x, value  # noqa: E501
                                      # noqa: E501
                                      in zip(repair_dict['Страховой полис'].keys(),
                                             repair_dict['Страховой полис'].values())}
    data = pd.DataFrame(repair_dict)
    data.to_excel(file_name)
    return look_data(file_name, file_path)


def sadoko_kst(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    try:
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        repair_dict = df.to_dict()
        repair_dict['Страховой полис'] = {x: re.sub('^0{3}|^0{4}|^0{2}|^0{5}', '', str(value)).replace(' ', '') for  # noqa: E501
                                          x, value  # noqa: E501
                                          in zip(repair_dict['Страховой полис'].keys(),
                                                 repair_dict['Страховой полис'].values())}
        data = pd.DataFrame(repair_dict)
        data.to_excel(file_name)
        return look_data(file_name, file_path)
    except:  # noqa: B001, E722
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги по прейскуранту  '].notnull()]  # noqa: E501
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def sadoko_crs(file_name, file_path):  # барать нан в гп
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    repair_dict = df.to_dict()
    repair_dict['Франшиза'] = {x: int(price) / int(discounted_price) * 100 for x, price, discounted_price in  # noqa: E501
                               zip(repair_dict['Цена услуги по прейскуранту  '].keys(),
                                   repair_dict['Цена услуги по прейскуранту  '].values(),
                                   repair_dict['Цена услуги с учетом скидки'].values())}
    repair_dict['номер гарантийного письма '] = {x: str(value).split('"', 2)[0] for x, value in
                                                 zip(repair_dict['номер гарантийного письма '].keys(),  # noqa: E501
                                                     repair_dict['номер гарантийного письма '].values())}  # noqa: E501
    data = pd.DataFrame(repair_dict)
    data.to_excel(file_name)
    return look_data(file_name, file_path)


def vivea(file_name, file_path):  # noqa: CCR001, CFQ001, C901
    stop = False
    patient = False
    fio = None
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    data = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    for row_number in range(12, sheet.max_row + 1):
        if stop:
            break
        for column_number in range(1, sheet.max_column + 1):
            inform_value = sheet.cell(row=row_number, column=column_number).value
            if sheet.cell(row=row_number, column=8).value is not None and sheet.cell(row=row_number,
                                                                                     column=1).value is None:  # noqa: E501
                patient = False
                break
            elif sheet.cell(row=row_number, column=8).value is None and sheet.cell(row=row_number,
                                                                                   column=1).value is None:  # noqa: E501
                stop = True
                break
            if not patient:
                if column_number == 3:
                    fio = str(inform_value)
                    patient = True
                    break
            else:
                if column_number == 1:
                    data['date'].append(str(inform_value))
                    data['fio'].append(fio)
                elif column_number == 2:
                    data['service_code'].append(inform_value)
                elif column_number == 3:
                    data['service_name'].append(inform_value)
                elif column_number == 7:
                    data['service_amount'].append(inform_value)
                elif column_number == 8:
                    data['service_price'].append(inform_value)
    return filling_book(data, file_name, file_path)  # Внести в зепелин


def litfond(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,  # noqa: E501
                                       header=None)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns
    Sheet_Name_Here.rename(columns={0: 'Номер филиала',
                                    2: 'Номер полиса',
                                    1: 'Номер мед.карты',
                                    3: 'ФИО',
                                    5: 'Дата услуги',
                                    6: 'МКБ',
                                    7: '№ зуба',
                                    8: 'Код услуги',
                                    9: 'Кол-во',
                                    10: 'Наименование услуги',
                                    11: 'Цена',
                                    15: 'Врач'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull()) & (
        ~Sheet_Name_Here['Кол-во'].str.contains('Кол-во', na=False))]  # noqa: E501

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def zhemchuzhina_podolia(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Стоимость услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def dalven(file_name, file_path):  # noqa: CFQ001
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]

        # Renamed columns ПропускЗУБ
        Sheet_Name_Here.rename(columns={'№ зуба': 'ПропускЗУБ'}, inplace=True)

        # Added column new-column-iia1
        Sheet_Name_Here.insert(1, 'new-column-iia1', 0)

        # Added column new-column-508q
        Sheet_Name_Here.insert(1, 'new-column-508q', 0)

        # Renamed columns № зуба
        Sheet_Name_Here.rename(columns={'new-column-508q': '№ зуба'}, inplace=True)

        # Renamed columns ФИО
        Sheet_Name_Here.rename(columns={'new-column-iia1': 'ФИО'}, inplace=True)

        # Renamed columns ПропускДАТА
        Sheet_Name_Here.rename(columns={'Диагноз': 'ПропускДАТА'}, inplace=True)

        # Added column new-column-qbse
        Sheet_Name_Here.insert(4, 'new-column-qbse', 0)

        # Added column new-column-4h5p
        Sheet_Name_Here.insert(4, 'new-column-4h5p', 0)

        # Renamed columns МКБ
        Sheet_Name_Here.rename(columns={'new-column-4h5p': 'МКБ'}, inplace=True)

        # Renamed columns Дата услуги
        Sheet_Name_Here.rename(columns={'new-column-qbse': 'Дата услуги'}, inplace=True)

        # Renamed columns Код услуги
        Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

        # Added column new-column-ekdj
        Sheet_Name_Here.insert(8, 'new-column-ekdj', 0)

        # Added column new-column-wiuf
        Sheet_Name_Here.insert(8, 'new-column-wiuf', 0)

        # Added column new-column-l2er
        Sheet_Name_Here.insert(8, 'new-column-l2er', 0)

        # Renamed columns Наименование услуги
        Sheet_Name_Here.rename(columns={'new-column-l2er': 'Наименование услуги'}, inplace=True)

        # Renamed columns ГП
        Sheet_Name_Here.rename(columns={'new-column-wiuf': 'ГП'}, inplace=True)

        # Renamed columns Номер полиса
        Sheet_Name_Here.rename(columns={'new-column-ekdj': 'Номер полиса'}, inplace=True)

        # Added column new-column-3zfo
        Sheet_Name_Here.insert(12, 'new-column-3zfo', 0)

        # Renamed columns Цена
        Sheet_Name_Here.rename(columns={'new-column-3zfo': 'Цена'}, inplace=True)

        # Renamed columns ПропускСУММА
        Sheet_Name_Here.rename(columns={'Сумма руб': 'ПропускСУММА'}, inplace=True)

        # Renamed columns ПропускУСЛУГА
        Sheet_Name_Here.rename(columns={'Название работы': 'ПропускУСЛУГА'}, inplace=True)

        # Set formula of № зуба
        Sheet_Name_Here['№ зуба'] = IF(AND(TYPE(Sheet_Name_Here['ПропускЗУБ']) != 'NaN',
                                           SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['ПропускЗУБ']), ' ', ''), '.',  # noqa: E501
                                                      # noqa: E501
                                                      '') != ''), Sheet_Name_Here['ПропускЗУБ'], None)  # noqa: E501

        # Set formula of ФИО
        Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['ПропускЗУБ']) != 'NaN',
                                        SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['ПропускЗУБ']), ' ', ''), '.',  # noqa: E501
                                                   # noqa: E501
                                                   '') == ''), Sheet_Name_Here['ПропускЗУБ'], None)

        # Set formula of Дата услуги
        Sheet_Name_Here['Дата услуги'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускДАТА']) != 'NaN', TYPE(Sheet_Name_Here['ПропускУСЛУГА']) == 'NaN'),  # noqa: E501
            # noqa: E501
            Sheet_Name_Here['ПропускДАТА'], None)

        # Set formula of МКБ
        Sheet_Name_Here['МКБ'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускДАТА']) != 'NaN', TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN'),  # noqa: E501
            Sheet_Name_Here['ПропускДАТА'], None)

        # Added column new-column-9ofx
        Sheet_Name_Here.insert(10, 'new-column-9ofx', 0)

        # Renamed columns ПереходныйПолис
        Sheet_Name_Here.rename(columns={'new-column-9ofx': 'ПереходныйПолис'}, inplace=True)

        # Set formula of ГП
        Sheet_Name_Here['ГП'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'ГП') > 0),  # noqa: E501
            # noqa: E501
            SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], LEFT(Sheet_Name_Here['ПропускУСЛУГА'],  # noqa: E501
                                                                         INT(FIND(Sheet_Name_Here['ПропускУСЛУГА'],  # noqa: E501
                                                                                  'ГП №') + 4)), ''), ')', ''),  # noqa: E501
            None)  # noqa: E501

        # Set formula of Наименование услуги
        Sheet_Name_Here['Наименование услуги'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'СП') <= 0,  # noqa: E501
                # noqa: E501
                FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'ГП') <= 0), Sheet_Name_Here['ПропускУСЛУГА'], None)  # noqa: E501

        # Set formula of ПереходныйПолис
        Sheet_Name_Here['ПереходныйПолис'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'ГП') > 0),  # noqa: E501
            SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'],
                       LEFT(Sheet_Name_Here['ПропускУСЛУГА'], INT(FIND(Sheet_Name_Here['ПропускУСЛУГА'], ' ('))), ''),  # noqa: E501
            '')  # noqa: E501

        # Set formula of Номер полиса
        Sheet_Name_Here['Номер полиса'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'СП')),  # noqa: E501
            FLOAT(  # noqa: E501
                SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], Sheet_Name_Here['ПереходныйПолис'], ''),  # noqa: E501
                           LEFT(SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], Sheet_Name_Here['ПереходныйПолис'], ''),  # noqa: E501
                                # noqa: E501
                                INT(FIND(
                                    SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], Sheet_Name_Here['ПереходныйПолис'],  # noqa: E501
                                               ''),  # noqa: E501
                                    ':'))), '')), None)

        # Set formula of Цена
        Sheet_Name_Here['Цена'] = IF(
            AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['ПропускСУММА']) != 'NaN'),  # noqa: E501
            FLOAT(FLOAT(Sheet_Name_Here['ПропускСУММА']) / INT(Sheet_Name_Here['Кол-во'])), None)

        # Added column new-column-18ir
        Sheet_Name_Here.insert(10, 'new-column-18ir', 0)

        # Renamed columns ПереходныйГП
        Sheet_Name_Here.rename(columns={'ГП': 'ПереходныйГП'}, inplace=True)

        # Renamed columns ГП
        Sheet_Name_Here.rename(columns={'new-column-18ir': 'ГП'}, inplace=True)

        # Added column new-column-w4qe
        Sheet_Name_Here.insert(10, 'new-column-w4qe', 0)

        # Renamed columns STOP_GP
        Sheet_Name_Here.rename(columns={'new-column-w4qe': 'STOP_GP'}, inplace=True)

        # Set formula of STOP_GP
        Sheet_Name_Here['STOP_GP'] = IF(
            AND(TYPE(Sheet_Name_Here['ПереходныйГП']) == 'object', FIND(Sheet_Name_Here['ПропускЗУБ'], 'Итого') > 0),  # noqa: E501
            '',  # noqa: E501
            None)

        # Set formula of ГП
        Sheet_Name_Here['ГП'] = FILLNAN(Sheet_Name_Here['STOP_GP'], Sheet_Name_Here['ПереходныйГП'])

        # Filled NaN values in 4 columns in Sheet_Name_Here
        columns_to_fill_nan = ['ФИО', 'Дата услуги', 'ГП', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501

        # Filtered Кол-во
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

        # Filtered Цена
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]

        # Changed Кол-во to dtype int
        Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

        # Set formula of Кол-во
        Sheet_Name_Here['Цена'] = IF(
            AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['ПропускСУММА']) != 'NaN'),  # noqa: E501
            FLOAT(FLOAT(Sheet_Name_Here['ПропускСУММА']) / INT(Sheet_Name_Here['Кол-во'])), None)
        Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'ФИО']
        Sheet_Name_Here_columns.insert(0, 'ФИО')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    except:  # noqa: B001, E722
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=16)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]

        # Renamed columns ПропускЗУБ
        Sheet_Name_Here.rename(columns={'№ зуба': 'ПропускЗУБ'}, inplace=True)

        # Added column new-column-iia1
        Sheet_Name_Here.insert(1, 'new-column-iia1', 0)

        # Added column new-column-508q
        Sheet_Name_Here.insert(1, 'new-column-508q', 0)

        # Renamed columns № зуба
        Sheet_Name_Here.rename(columns={'new-column-508q': '№ зуба'}, inplace=True)

        # Renamed columns ФИО
        Sheet_Name_Here.rename(columns={'new-column-iia1': 'ФИО'}, inplace=True)

        # Renamed columns ПропускДАТА
        Sheet_Name_Here.rename(columns={'Диагноз': 'ПропускДАТА'}, inplace=True)

        # Added column new-column-qbse
        Sheet_Name_Here.insert(4, 'new-column-qbse', 0)

        # Added column new-column-4h5p
        Sheet_Name_Here.insert(4, 'new-column-4h5p', 0)

        # Renamed columns МКБ
        Sheet_Name_Here.rename(columns={'new-column-4h5p': 'МКБ'}, inplace=True)

        # Renamed columns Дата услуги
        Sheet_Name_Here.rename(columns={'new-column-qbse': 'Дата услуги'}, inplace=True)

        # Renamed columns Код услуги
        Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

        # Added column new-column-ekdj
        Sheet_Name_Here.insert(8, 'new-column-ekdj', 0)

        # Added column new-column-wiuf
        Sheet_Name_Here.insert(8, 'new-column-wiuf', 0)

        # Added column new-column-l2er
        Sheet_Name_Here.insert(8, 'new-column-l2er', 0)

        # Renamed columns Наименование услуги
        Sheet_Name_Here.rename(columns={'new-column-l2er': 'Наименование услуги'}, inplace=True)

        # Renamed columns ГП
        Sheet_Name_Here.rename(columns={'new-column-wiuf': 'ГП'}, inplace=True)

        # Renamed columns Номер полиса
        Sheet_Name_Here.rename(columns={'new-column-ekdj': 'Номер полиса'}, inplace=True)

        # Added column new-column-3zfo
        Sheet_Name_Here.insert(12, 'new-column-3zfo', 0)

        # Renamed columns Цена
        Sheet_Name_Here.rename(columns={'new-column-3zfo': 'Цена'}, inplace=True)

        # Renamed columns ПропускСУММА
        Sheet_Name_Here.rename(columns={'Сумма руб': 'ПропускСУММА'}, inplace=True)

        # Renamed columns ПропускУСЛУГА
        Sheet_Name_Here.rename(columns={'Название работы': 'ПропускУСЛУГА'}, inplace=True)

        # Set formula of № зуба
        Sheet_Name_Here['№ зуба'] = IF(AND(TYPE(Sheet_Name_Here['ПропускЗУБ']) != 'NaN',
                                           SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['ПропускЗУБ']), ' ', ''), '.',  # noqa: E501
                                                      # noqa: E501
                                                      '') != ''), Sheet_Name_Here['ПропускЗУБ'], None)  # noqa: E501

        # Set formula of ФИО
        Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['ПропускЗУБ']) != 'NaN',
                                        SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['ПропускЗУБ']), ' ', ''), '.',  # noqa: E501
                                                   # noqa: E501
                                                   '') == ''), Sheet_Name_Here['ПропускЗУБ'], None)
        # Set formula of Дата услуги
        Sheet_Name_Here['Дата услуги'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускДАТА']) != 'NaN', TYPE(Sheet_Name_Here['ПропускУСЛУГА']) == 'NaN'),  # noqa: E501
            Sheet_Name_Here['ПропускДАТА'], None)

        # Set formula of МКБ
        Sheet_Name_Here['МКБ'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускДАТА']) != 'NaN', TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN'),  # noqa: E501
            Sheet_Name_Here['ПропускДАТА'], None)

        # Added column new-column-9ofx
        Sheet_Name_Here.insert(10, 'new-column-9ofx', 0)

        # Renamed columns ПереходныйПолис
        Sheet_Name_Here.rename(columns={'new-column-9ofx': 'ПереходныйПолис'}, inplace=True)

        # Set formula of ГП
        Sheet_Name_Here['ГП'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'ГП') > 0),  # noqa: E501
            # noqa: E501
            SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], LEFT(Sheet_Name_Here['ПропускУСЛУГА'],  # noqa: E501
                                                                         INT(FIND(Sheet_Name_Here['ПропускУСЛУГА'],  # noqa: E501
                                                                                  # noqa: E501
                                                                                  'ГП №') + 4)), ''), ')', ''), None)  # noqa: E501

        # Set formula of Наименование услуги
        Sheet_Name_Here['Наименование услуги'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'СП') <= 0,  # noqa: E501
                FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'ГП') <= 0), Sheet_Name_Here['ПропускУСЛУГА'], None)  # noqa: E501

        # Set formula of ПереходныйПолис
        Sheet_Name_Here['ПереходныйПолис'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'ГП') > 0),  # noqa: E501
            SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'],
                       LEFT(Sheet_Name_Here['ПропускУСЛУГА'], INT(FIND(Sheet_Name_Here['ПропускУСЛУГА'], ' ('))), ''),  # noqa: E501
            # noqa: E501
            0)

        # Set formula of Номер полиса
        Sheet_Name_Here['Номер полиса'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускУСЛУГА']) != 'NaN', FIND(Sheet_Name_Here['ПропускУСЛУГА'], 'СП')),  # noqa: E501
            FLOAT(  # noqa: E501
                SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], Sheet_Name_Here['ПереходныйПолис'], ''),  # noqa: E501
                           LEFT(SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], Sheet_Name_Here['ПереходныйПолис'], ''),  # noqa: E501
                                INT(FIND(
                                    SUBSTITUTE(Sheet_Name_Here['ПропускУСЛУГА'], Sheet_Name_Here['ПереходныйПолис'],  # noqa: E501
                                               ''),
                                    ':'))), '')), None)

        # Set formula of Цена
        Sheet_Name_Here['Цена'] = IF(
            AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['ПропускСУММА']) != 'NaN'),  # noqa: E501
            FLOAT(FLOAT(Sheet_Name_Here['ПропускСУММА']) / INT(Sheet_Name_Here['Кол-во'])), None)

        # Added column new-column-18ir
        Sheet_Name_Here.insert(10, 'new-column-18ir', 0)

        # Renamed columns ПереходныйГП
        Sheet_Name_Here.rename(columns={'ГП': 'ПереходныйГП'}, inplace=True)

        # Renamed columns ГП
        Sheet_Name_Here.rename(columns={'new-column-18ir': 'ГП'}, inplace=True)

        # Added column new-column-w4qe
        Sheet_Name_Here.insert(10, 'new-column-w4qe', 0)

        # Renamed columns STOP_GP
        Sheet_Name_Here.rename(columns={'new-column-w4qe': 'STOP_GP'}, inplace=True)

        # Set formula of STOP_GP
        Sheet_Name_Here['STOP_GP'] = IF(
            AND(TYPE(Sheet_Name_Here['ПереходныйГП']) == 'object', FIND(Sheet_Name_Here['ПропускЗУБ'], 'Итого') > 0),  # noqa: E501
            # noqa: E501
            '',
            None)

        # Set formula of ГП
        Sheet_Name_Here['ГП'] = FILLNAN(Sheet_Name_Here['STOP_GP'], Sheet_Name_Here['ПереходныйГП'])

        # Filled NaN values in 4 columns in Sheet_Name_Here
        columns_to_fill_nan = ['ФИО', 'Дата услуги', 'ГП', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501

        # Filtered Кол-во
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

        # Filtered Цена
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]

        # Changed Кол-во to dtype int
        Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

        # Set formula of Кол-во
        Sheet_Name_Here['Цена'] = IF(
            AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['ПропускСУММА']) != 'NaN'),  # noqa: E501
            FLOAT(FLOAT(Sheet_Name_Here['ПропускСУММА']) / INT(Sheet_Name_Here['Кол-во'])), None)
        Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'ФИО']
        Sheet_Name_Here_columns.insert(0, 'ФИО')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def sk_test(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3, converters={'№ зуба': str, 'Код услуги': str})  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Фамилия': 'ПропускФ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Имя': 'ПропускИ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Отчество': 'ПропускО'}, inplace=True)
    Sheet_Name_Here.insert(3, 'new-column-cvx9', 0)
    Sheet_Name_Here.rename(columns={'new-column-cvx9': 'ФИО'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(
        AND(TYPE(Sheet_Name_Here['ПропускФ']) != 'NaN', TYPE(Sheet_Name_Here['ПропускИ']) != 'NaN', TYPE(Sheet_Name_Here['ПропускО']) != 'NaN'),  # noqa: E501
        PROPER(CONCAT(Sheet_Name_Here['ПропускФ'], Sheet_Name_Here['ПропускИ'], Sheet_Name_Here['ПропускО'])), None)  # noqa: E501
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата оказания': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Диагноз МКБ': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Кол-во'}, inplace=True)
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ilinskaya_bolnica(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stil_dent(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество \nуслуг\n'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_name = wb.sheetnames[1]
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во оказанных услуг'].notnull()]
        columns_to_fill_nan = ['ФИО', '№.1']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501
        Sheet_Name_Here.rename(columns={'№.1': 'Полис'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Поверх': 'Поверхность зуба'}, inplace=True)
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def allergomed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Застрахованного': 'Пропуск'}, inplace=True)

    # Added column new-column-ztow
    Sheet_Name_Here.insert(3, 'new-column-ztow', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-ztow': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', PROPER(Sheet_Name_Here['Пропуск']), None)  # noqa: E501

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'оказанных услуг': 'Наименование услуги'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'оказания': 'Дата услуги'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'услуг': 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'услуги,': 'Цена'}, inplace=True)

    # Added column new-column-hjo7
    Sheet_Name_Here.insert(10, 'new-column-hjo7', 0)

    # Renamed columns ПРОПУСКСКИДКА
    Sheet_Name_Here.rename(columns={'new-column-hjo7': 'ПРОПУСКСКИДКА'}, inplace=True)

    # Added column new-column-doot
    Sheet_Name_Here.insert(11, 'new-column-doot', 0)

    # Renamed columns Скидка
    Sheet_Name_Here.rename(columns={'new-column-doot': 'Скидка'}, inplace=True)

    # Set formula of ПРОПУСКСКИДКА
    Sheet_Name_Here['ПРОПУСКСКИДКА'] = IF(
        AND(TYPE(Sheet_Name_Here['№ полиса']) != 'NaN', FIND(Sheet_Name_Here['№ полиса'], 'Скидка')), SUBSTITUTE(  # noqa: E501
            SUBSTITUTE(Sheet_Name_Here['№ полиса'],
                       LEFT(Sheet_Name_Here['№ полиса'], INT(FIND(Sheet_Name_Here['№ полиса'], '='))), ''), ' руб.',  # noqa: E501
            ''), None)

    # Filled NaN values in 1 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ПРОПУСКСКИДКА']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='bfill')  # noqa: E501

    # Set formula of Скидка
    Sheet_Name_Here['Скидка'] = IF(Sheet_Name_Here['ПРОПУСКСКИДКА'] == '0', None, Sheet_Name_Here['ПРОПУСКСКИДКА'])  # noqa: E501

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

    # Added column new-column-49l5
    Sheet_Name_Here.insert(5, 'new-column-49l5', 0)

    # Renamed columns ПропускМКБ
    Sheet_Name_Here.rename(columns={'МКБ': 'ПропускМКБ'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'new-column-49l5': 'МКБ'}, inplace=True)

    # Set formula of МКБ
    Sheet_Name_Here['МКБ'] = IF(AND(TYPE(Sheet_Name_Here['ПропускМКБ']) != 'NaN', Sheet_Name_Here['ПропускМКБ'] != ' '),  # noqa: E501
                                Sheet_Name_Here['ПропускМКБ'], None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def magazin_nedvizhemosti(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=25, header=None)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={2: 'Номер полиса'}, inplace=True)

    # Added column new-column-7v3v
    Sheet_Name_Here.insert(6, 'new-column-7v3v', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-7v3v': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here[5]) != 'NaN', PROPER(Sheet_Name_Here[5]), None)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={7: 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={8: 'Наименование услуги'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={9: 'Кол-во'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={14: 'Врач'}, inplace=True)

    # Renamed columns Код врача
    Sheet_Name_Here.rename(columns={13: 'Код врача'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={12: 'Дата услуги'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={10: 'Цена'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Кол. услуг', na=False))]  # noqa: E501

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    # Changed Номер полиса to dtype float
    Sheet_Name_Here['Номер полиса'] = to_float_series(Sheet_Name_Here['Номер полиса'])

    # Changed Код врача to dtype float
    Sheet_Name_Here['Код врача'] = to_float_series(Sheet_Name_Here['Код врача'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def data_pandas_stomat_tagansk(file_name, file_path):  # добавить взепелин
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)  # noqa: E501
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')  # noqa: E501
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,  # noqa: E501
                                           header=None)
        Sheet1 = sheet_df_dictonary[sheet_name]
        Sheet1 = Sheet1[Sheet1[8].notnull()]
        Sheet1 = Sheet1[Sheet1[7].notnull()]
        columns_to_fill_nan = [0, 1, 3]
        Sheet1[columns_to_fill_nan] = Sheet1[columns_to_fill_nan].fillna(method='ffill')
        df = pd.DataFrame.from_dict(Sheet1)
        try:
            repair_dict = df.to_dict()
            for x in repair_dict['Наименование услуги']:
                if 'Итого со скидкой' in x:
                    discount_size = x.replace('Итого со скидкой ', '', x.replace('%:', '', x))
                    break
                else:
                    discount_size = None
                    continue
            repair_dict['Скидка'] = {x: discount_size for x in repair_dict['Количество'].keys()}
            df = pd.DataFrame(repair_dict)
            df = df[df['Количество'].notnull()]
        except:
            pass
        df.to_excel(file_name)
        return look_data(file_name, file_path)  # добавить взепелин


def data_pandas_azino(file_name, file_path):  # добавить взепелин
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)  # noqa: E501
    list_1 = sheet_df_dictonary[sheet_name]
    list_1 = list_1[list_1[8].notnull()]
    columns_to_fill_nan = [0, 1, 2, 3]
    list_1[columns_to_fill_nan] = list_1[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(list_1)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medroscontract(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def data_pandas_kredo_experto(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4,
                                       converters={'Цена (руб)': str})
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    Sheet_Name_Here.rename(columns={'Филиал': 'Номер филиала'}, inplace=True)

    # Renamed columns Номер мед.карты
    Sheet_Name_Here.rename(columns={'Код пациента': 'Номер мед.карты'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Страховой полис': 'Номер полиса'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'ФИО пациента': 'ФИО'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код диагноза': 'МКБ'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Стоимость (руб)': 'Пропуск'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'Доктор': 'Врач'}, inplace=True)

    # Renamed columns Пропуск2
    Sheet_Name_Here.rename(columns={'Программа прикрепления': 'Пропуск2'}, inplace=True)

    # Renamed columns Пропуск3
    Sheet_Name_Here.rename(columns={'Место работы': 'Пропуск3'}, inplace=True)

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = Sheet_Name_Here['Цена'].astype('float')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def davidovskogo(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(3, 'new-column-3x9u', 0)

    # Renamed columns ПропускФИО
    Sheet_Name_Here.rename(columns={'ФИО': 'ПропускФИО'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-3x9u': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис': 'Номер полиса'}, inplace=True)

    # Added column new-column-i8j4
    Sheet_Name_Here.insert(5, 'new-column-i8j4', 0)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'new-column-i8j4': 'Дата услуги'}, inplace=True)

    # Added column new-column-yt9z
    Sheet_Name_Here.insert(6, 'new-column-yt9z', 0)

    # Renamed columns Дата окончания услуги
    Sheet_Name_Here.rename(columns={'new-column-yt9z': 'Дата окончания услуги'}, inplace=True)

    # Renamed columns ПропускДАТА
    Sheet_Name_Here.rename(columns={'Дата начала и окончания госпитализации': 'ПропускДАТА'}, inplace=True)

    # Renamed columns ПропускДАТАРЕЗЕРВ
    Sheet_Name_Here.rename(columns={'Дата оказания': 'ПропускДАТАРЕЗЕРВ'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)

    # Renamed columns Наименование филиала клиники
    Sheet_Name_Here.rename(columns={'Отделение': 'Наименование филиала клиники'}, inplace=True)

    # Renamed columns ПропускВРАЧ
    Sheet_Name_Here.rename(columns={'ФИО лечащего врача': 'ПропускВРАЧ'}, inplace=True)

    # Added column new-column-1hwh
    Sheet_Name_Here.insert(13, 'new-column-1hwh', 0)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'new-column-1hwh': 'Врач'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Количество': 'Кол-во'}, inplace=True)

    # Renamed columns ПропускЦЕНА
    Sheet_Name_Here.rename(columns={'Цена услуги': 'ПропускЦЕНА'}, inplace=True)

    # Added column new-column-aha0
    Sheet_Name_Here.insert(15, 'new-column-aha0', 0)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-aha0': 'Цена'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['ПропускФИО']) != 'NaN', PROPER(Sheet_Name_Here['ПропускФИО']),
                                None)

    # Set formula of Дата услуги
    Sheet_Name_Here['Дата услуги'] = IF(TYPE(Sheet_Name_Here['ПропускДАТА']) == 'NaN',
                                        TEXT(Sheet_Name_Here['ПропускДАТАРЕЗЕРВ']),
                                        LEFT(Sheet_Name_Here['ПропускДАТА'],
                                             INT(FIND(Sheet_Name_Here['ПропускДАТА'], '-') - 2)))

    # Set formula of Дата окончания услуги
    Sheet_Name_Here['Дата окончания услуги'] = IF(TYPE(Sheet_Name_Here['ПропускДАТА']) != 'NaN',
                                                  SUBSTITUTE(Sheet_Name_Here['ПропускДАТА'],
                                                             LEFT(Sheet_Name_Here['ПропускДАТА'],
                                                                  INT(FIND(Sheet_Name_Here['ПропускДАТА'],
                                                                           '-'))), ''), None)

    # Set formula of Врач
    Sheet_Name_Here['Врач'] = IF(TYPE(Sheet_Name_Here['ПропускВРАЧ']) != 'NaN', PROPER(Sheet_Name_Here['ПропускВРАЧ']),
                                 None)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['ПропускЦЕНА']) != 'NaN', Sheet_Name_Here['Скидка'] != '0.0'),
                                 Sheet_Name_Here['ПропускЦЕНА'] - (
                                             Sheet_Name_Here['ПропускЦЕНА'] / 100 * Sheet_Name_Here['Скидка']),
                                 Sheet_Name_Here['ПропускЦЕНА'])

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_52_kdc_min_ob_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if 'п.x' in file_name:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол. Услуг'].notnull()]
        columns_to_fill_nan = ['Фамилия И. О. № полиса', '№']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Выполненные исследования'].str.contains('7', na=False)]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        try:
            repair_dict = df.to_dict()
            repair_dict['ФИО'] = {x: str(value).split('           ', 2)[0] for x, value in
                                  zip(repair_dict['Фамилия И. О. № полиса'].keys(),
                                      repair_dict['Фамилия И. О. № полиса'].values())}
            repair_dict['№ полиса'] = {x: int(str(value).split('           ', 2)[-1]) for x, value in
                                       zip(repair_dict['Фамилия И. О. № полиса'].keys(),
                                           repair_dict['Фамилия И. О. № полиса'].values())}
            repair_dict.pop('Фамилия И. О. № полиса')
            data = pd.DataFrame(repair_dict)
            data.to_excel(file_name)
        except:
            df.to_excel(file_name)
        return look_data(file_name, file_path)
    if 'c.x' in file_name:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол. Услуг'].notnull()]
        Sheet_Name_Here = Sheet_Name_Here[
            (Sheet_Name_Here['Стоим. за ед.'].notnull()) & (Sheet_Name_Here['Стоим. за ед.'] != 10)]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        repair_dict = df.to_dict()
        repair_dict['ФИО'] = {x: str(value).split('           ', 2)[0] for x, value in
                              zip(repair_dict['Фамилия И. О. № полиса'].keys(),
                                  repair_dict['Фамилия И. О. № полиса'].values())}
        repair_dict['№ полиса'] = {x: int(str(value).split('           ', 2)[-1]) for x, value in
                                   zip(repair_dict['Фамилия И. О. № полиса'].keys(),
                                       repair_dict['Фамилия И. О. № полиса'].values())}
        repair_dict.pop('Фамилия И. О. № полиса')
        data = pd.DataFrame(repair_dict)
        data.to_excel(file_name)
        return look_data(file_name, file_path)
    elif 'k.x' in file_name:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        columns_to_fill_nan = ['Фамилия И.О. \n№ полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол.\nУслуг'].notnull()]
        Sheet_Name_Here = Sheet_Name_Here[
            Sheet_Name_Here['Врач'].apply(lambda val: all(s not in str(val) for s in ['Итого:   ', '8']))]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        try:
            repair_dict = df.to_dict()
            repair_dict['ФИО'] = {x: str(value).split('           ', 2)[0] for x, value in
                                  zip(repair_dict['Фамилия И. О. № полиса'].keys(),
                                      repair_dict['Фамилия И. О. № полиса'].values())}
            repair_dict['№ полиса'] = {x: int(str(value).split('           ', 2)[-1]) for x, value in
                                       zip(repair_dict['Фамилия И. О. № полиса'].keys(),
                                           repair_dict['Фамилия И. О. № полиса'].values())}
            repair_dict.pop('Фамилия И. О. № полиса')
            data = pd.DataFrame(repair_dict)
            data.to_excel(file_name)
        except:
            df.to_excel(file_name)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        columns_to_fill_nan = ['Фамилия И.О. \n№ полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол.\nУслуг'].notnull()]
        Sheet_Name_Here = Sheet_Name_Here[
            Sheet_Name_Here['Врач'].apply(lambda val: all(s not in str(val) for s in ['Итого:   ', '8']))]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        repair_dict = df.to_dict()
        repair_dict['ФИО'] = {x: str(value).split('\n', 2)[0] for x, value in
                              zip(repair_dict['Фамилия И.О. \n№ полиса'].keys(),
                                  repair_dict['Фамилия И.О. \n№ полиса'].values())}
        repair_dict['№ полиса'] = {x: int(str(value).split('\n', 2)[-1]) for x, value in
                                   zip(repair_dict['Фамилия И.О. \n№ полиса'].keys(),
                                       repair_dict['Фамилия И.О. \n№ полиса'].values())}
        repair_dict.pop('Фамилия И.О. \n№ полиса')
        data = pd.DataFrame(repair_dict)
        data.to_excel(file_name)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def praktica_32(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Дата ': 'Дата рождения'}, inplace=True)
    Sheet_Name_Here.rename(columns={'ID карта': 'ID карта пациента'}, inplace=True)
    Sheet_Name_Here.rename(columns={'дата оказания': 'дата оказания услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Диагноз': 'Диагноз (код по МКБ)'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Дата рождения'].str.contains('рождения', na=False)]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def avismed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=13)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def aibolit(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Номер мед.карты
    Sheet_Name_Here.rename(columns={1: 'Номер мед.карты'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={2: 'Номер полиса'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={3: 'ФИО'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={4: 'Дата услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={5: 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={6: 'Наименование услуги'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={8: 'Кол-во'}, inplace=True)

    # Added column new-column-7fva
    Sheet_Name_Here.insert(10, 'new-column-7fva', 0)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-7fva': 'Цена'}, inplace=True)

    # Added column new-column-ypkl
    Sheet_Name_Here.insert(10, 'new-column-ypkl', 0)

    # Renamed columns Скидка
    Sheet_Name_Here.rename(columns={'new-column-ypkl': 'Скидка'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={11: 'МКБ'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={12: 'Диагноз'}, inplace=True)

    # Filtered ФИО
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['ФИО'].notnull()) & (~Sheet_Name_Here['ФИО'].str.contains('ФИО', na=False))]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    # Changed 10 to dtype float
    Sheet_Name_Here[10] = to_float_series(Sheet_Name_Here[10])

    # Set formula of Скидка
    Sheet_Name_Here['Скидка'] = IF(TYPE(Sheet_Name_Here[9]) != 'NaN', TEXT(FLOAT(Sheet_Name_Here[9]) * 100), None)

    # Changed 9 to dtype float
    Sheet_Name_Here[9] = to_float_series(Sheet_Name_Here[9])

    # Set formula of 9
    Sheet_Name_Here['Скидка'] = IF(TYPE(Sheet_Name_Here[9]) != 'NaN', TEXT(FLOAT(Sheet_Name_Here[9]) * 100), None)

    # Deleted columns Скидка
    Sheet_Name_Here.drop(['Скидка'], axis=1, inplace=True)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = Sheet_Name_Here[10] / Sheet_Name_Here['Кол-во']



def allegro(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def buyanova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во\nуслуг'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def v_nadezhnih_rukah(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.drop(['№'], axis=1, inplace=True)
        Sheet_Name_Here.rename(columns={'Номер и дата гарантийного письма': 'ГП'}, inplace=True)
        Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Фамилия Имя Отчество застрахованного': 'ФИО'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Дата выписки': 'Дата окончания услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'МКБ'}, inplace=True)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
        Sheet_Name_Here.insert(11, 'new-column-8wwq', 0)
        Sheet_Name_Here.rename(columns={'new-column-8wwq': 'Цена'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Стоимость': 'Пропуск'}, inplace=True)
        Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                     SUBSTITUTE(Sheet_Name_Here['Пропуск'], ' ', ''), None)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
        Sheet_Name_Here['Номер полиса'] = to_int_series(Sheet_Name_Here['Номер полиса'])
        columns_to_fill_nan = ['ФИО']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.drop(['№'], axis=1, inplace=True)
        Sheet_Name_Here.rename(columns={'Номер и дата гарантийного письма': 'ГП'}, inplace=True)
        Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Фамилия Имя Отчество застрахованного': 'ФИО'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Дата выписки': 'Дата окончания услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'МКБ'}, inplace=True)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
        Sheet_Name_Here.insert(11, 'new-column-8wwq', 0)
        Sheet_Name_Here.rename(columns={'new-column-8wwq': 'Цена'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Стоимость': 'Пропуск'}, inplace=True)
        Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                     SUBSTITUTE(Sheet_Name_Here['Пропуск'], ' ', ''), None)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
        Sheet_Name_Here['Номер полиса'] = to_int_series(Sheet_Name_Here['Номер полиса'])
        columns_to_fill_nan = ['ФИО']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def veronika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=12)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-uadk
    Sheet_Name_Here.insert(1, 'new-column-uadk', 0)

    # Added column new-column-kq2d
    Sheet_Name_Here.insert(1, 'new-column-kq2d', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-kq2d': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'new-column-uadk': 'Номер полиса'}, inplace=True)

    # Added column new-column-n4al
    Sheet_Name_Here.insert(3, 'new-column-n4al', 0)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-n4al': 'ГП'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код\nработы': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название работы': 'Наименование услуги'}, inplace=True)

    # Added column new-column-8fwv
    Sheet_Name_Here.insert(20, 'new-column-8fwv', 0)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-8fwv': 'Цена'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={'№\nзуба': '№ зуба'}, inplace=True)

    # Added column new-column-8wlq
    Sheet_Name_Here.insert(4, 'new-column-8wlq', 0)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'new-column-8wlq': 'Дата услуги'}, inplace=True)

    # Added column new-column-m7xt
    Sheet_Name_Here.insert(5, 'new-column-m7xt', 0)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'new-column-m7xt': 'Врач'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(
        AND(FIND(Sheet_Name_Here['№ зуба'], '.') <= 0, SUBSTITUTE(CLEAN(Sheet_Name_Here['№ зуба']), ' ', '') == '',
            TYPE(Sheet_Name_Here['№ зуба']) != 'NaN'), PROPER(Sheet_Name_Here['№ зуба']), None)

    # Set formula of Дата услуги
    Sheet_Name_Here['Дата услуги'] = IF(
        AND(TYPE(Sheet_Name_Here['№ зуба']) != 'NaN', FIND(Sheet_Name_Here['№ зуба'], '  ') > 0), SUBSTITUTE(
            SUBSTITUTE(Sheet_Name_Here['№ зуба'],
                       LEFT(Sheet_Name_Here['№ зуба'], INT(FIND(Sheet_Name_Here['№ зуба'], '  '))), ''), ' ', ''), None)

    # Set formula of Врач
    Sheet_Name_Here['Врач'] = IF(
        AND(TYPE(Sheet_Name_Here['№ зуба']) != 'NaN', FIND(Sheet_Name_Here['№ зуба'], '  ') > 0),
        LEFT(Sheet_Name_Here['№ зуба'], INT(FIND(Sheet_Name_Here['№ зуба'], '  '))), None)

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = IF(
        AND(TYPE(Sheet_Name_Here['Наименование услуги']) != 'NaN',
            FIND(Sheet_Name_Here['Наименование услуги'], '№ СП:') > 0,
            FIND(Sheet_Name_Here['Наименование услуги'], 'ГП') <= 0),
        FLOAT(SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['Наименование услуги'],
                                    LEFT(Sheet_Name_Here[
                                             'Наименование услуги'],
                                         INT(FIND(Sheet_Name_Here[
                                                      'Наименование услуги'],
                                                  ':'))), ''), ' ',
                         '')), None)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = IF(AND(TYPE(Sheet_Name_Here['Наименование услуги']) != 'NaN',
                                   FIND(Sheet_Name_Here['Наименование услуги'], 'ГП') > 0),
                               SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['Наименование услуги'], CLEAN(
                                   LEFT(Sheet_Name_Here['Наименование услуги'],
                                        INT(FIND(Sheet_Name_Here['Наименование услуги'], 'ГП')))), ''), ' ',
                                          ''), None)

    # Added column new-column-fim9
    Sheet_Name_Here.insert(4, 'new-column-fim9', 0)

    # Renamed columns ГП0000
    Sheet_Name_Here.rename(columns={'ГП': 'ГП0000'}, inplace=True)

    # Added column new-column-7w3j
    Sheet_Name_Here.insert(5, 'new-column-7w3j', 0)

    # Renamed columns ГП00001
    Sheet_Name_Here.rename(columns={'new-column-fim9': 'ГП00001'}, inplace=True)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-7w3j': 'ГП'}, inplace=True)

    # Set formula of ГП00001
    Sheet_Name_Here['ГП00001'] = IF(
        AND(TYPE(Sheet_Name_Here['№ зуба']) != 'NaN', FIND(Sheet_Name_Here['№ зуба'], 'Итого') > 0,
            TYPE(Sheet_Name_Here['ГП0000']) == 'NaN'), '',
        None)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = FILLNAN(Sheet_Name_Here['ГП00001'], Sheet_Name_Here['ГП0000'])

    # Renamed columns Summ
    Sheet_Name_Here.rename(columns={'Сумма\nруб': 'Summ'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-\nво': 'Кол-во'}, inplace=True)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['Summ']) != 'NaN', TYPE(Sheet_Name_Here['Кол-во']) != 'NaN'),
                                 Sheet_Name_Here['Summ'] / INT(Sheet_Name_Here['Кол-во']), None)

    # Filled NaN values in 5 columns in Лист1
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'ГП', 'Дата услуги', 'Врач']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Set formula of Кол-во
    Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['Summ']) != 'NaN', TYPE(Sheet_Name_Here['Кол-во']) != 'NaN'),
                                 Sheet_Name_Here['Summ'] / INT(Sheet_Name_Here['Кол-во']), None)

    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'ФИО']
    Sheet_Name_Here_columns.insert(0, 'ФИО')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def intan(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    columns_to_fill_nan = ['Ф.И.О. ', 'Номер полиса ДМС', '№ и дата гарантийного письма', 'Дата оказания услуги',
                           'Врач']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def zdorovor_pokolenie(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№ п/п'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def zvezda(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-3khn
    Sheet_Name_Here.insert(3, 'new-column-3khn', 0)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Фамилия Имя Отчество': 'Пропуск'}, inplace=True)

    # Renamed columns Номер полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полис'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-3khn': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', Sheet_Name_Here['Пропуск'] != ' '),
                                Sheet_Name_Here['Пропуск'], None)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование медицинской услуги по Прейскуранту': 'Наименование услуги'},
                           inplace=True)

    # Renamed columns Пропкус
    Sheet_Name_Here.rename(columns={'Сумма': 'Пропкус'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ 10': 'Диагноз'}, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Номер полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Итого', na=False))]

    # Filtered Цена
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def zao_medi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п.п.
    Sheet_Name_Here.drop(['№ п.п.'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Клиника': 'Наименование филиала клиники (точки)'}, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Added column new-column-f9pq
    Sheet_Name_Here.insert(3, 'new-column-f9pq', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-f9pq': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['Фамилия'], ' ', Sheet_Name_Here['Имя'], ' ',
                                             Sheet_Name_Here['Отчество'])

    # Deleted columns Фамилия, Имя, Отчество
    Sheet_Name_Here.drop(['Фамилия', 'Имя', 'Отчество'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def granti_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if 'Контрагент' in str(wb[sheet_name].cell(3, 1).value):
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def gk_medecina(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["""'Кол-во услуг"""].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vorohobova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во"].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vitonika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во"].notnull()]
    columns_to_fill_nan = ['ФИО', '(ID) Пациента', 'Дата']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ermishanceva(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во"].notnull()]
    columns_to_fill_nan = ['Дата', 'Фамилия И.О. пациента', '№ полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def jms(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=11)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во"].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def dentiform(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    columns_to_fill_nan = ['ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    repair_dict = df.to_dict()

    # Доставать номер зуба из наименования услуги (первый столбец с фио)
    data = pd.DataFrame(repair_dict)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во"].notnull()]
    data = pd.DataFrame.from_dict(Sheet_Name_Here)
    data.to_excel(file_name)
    return look_data(file_name, file_path)


def dentika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Застрахованный': 'Пропуск'}, inplace=True)

    # Added column new-column-5ups
    Sheet_Name_Here.insert(3, 'new-column-5ups', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-5ups': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', PROPER(Sheet_Name_Here['Пропуск']), None)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена, руб.': 'Цена'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Кол-во'}, inplace=True)

    # Added column new-column-x1nh
    Sheet_Name_Here.insert(12, 'new-column-x1nh', 0)

    # Added column new-column-bbdz
    Sheet_Name_Here.insert(12, 'new-column-bbdz', 0)

    # Added column new-column-kuuy
    Sheet_Name_Here.insert(12, 'new-column-kuuy', 0)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'new-column-kuuy': 'Диагноз'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'new-column-bbdz': 'МКБ'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={'new-column-x1nh': '№ зуба'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'Код врача (или ФИО)': 'Врач'}, inplace=True)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'№ гарантийного письма': 'ГП'}, inplace=True)

    # Renamed columns Дата ГП
    Sheet_Name_Here.rename(columns={'Дата гарантийного письма': 'Дата ГП'}, inplace=True)

    # Renamed columns Пропуск1
    Sheet_Name_Here.rename(columns={'Диагноз или код (по МКБ-10)': 'Пропуск1'}, inplace=True)

    Sheet_Name_Here.rename(columns={'Стоимость, руб.': 'Пропуск000'}, inplace=True)
    # Set formula of МКБ
    Sheet_Name_Here['МКБ'] = IF(TYPE(Sheet_Name_Here['Пропуск1']) != 'NaN',
                                LEFT(Sheet_Name_Here['Пропуск1'], INT(FIND(Sheet_Name_Here['Пропуск1'], ' '))),
                                None)

    # Set formula of № зуба
    Sheet_Name_Here['№ зуба'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск1']) != 'NaN', FIND(Sheet_Name_Here['Пропуск1'], 'зуб')),
        SUBSTITUTE(LEFT(Sheet_Name_Here['Пропуск1'], INT(FIND(Sheet_Name_Here['Пропуск1'], '-') - 1)),
                   CONCAT(Sheet_Name_Here['МКБ'], 'зуб'), ''), None)

    # Set formula of Диагноз
    Sheet_Name_Here['Диагноз'] = IF(TYPE(Sheet_Name_Here['Пропуск1']) != 'NaN', SUBSTITUTE(
        SUBSTITUTE(Sheet_Name_Here['Пропуск1'],
                   LEFT(Sheet_Name_Here['Пропуск1'], INT(FIND(Sheet_Name_Here['Пропуск1'], '-'))), ''),
        Sheet_Name_Here['МКБ'],
        ''), None)

    # Filled NaN values in 3 columns in Лист1
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Номер мед.карты']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Итого', na=False))]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def diamond(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Номер ID карты': 'Страховой полис',
                                    'ДАТА ОКАЗАНИЯ УСЛУГИ': 'Дата начала оказания услуги',
                                    'КОД УСЛУГИ': 'Код услуги',
                                    'НАИМЕНОВАНИЕ УСЛУГИ': 'Наименование услуги',
                                    'КОЛ.  УСЛУГ': 'Количество',
                                    'ЦЕНА': 'Цена услуги',
                                    'ФИО врача / код врача': 'Врач (ФИО)',
                                    'Диагноз МКБ-10': 'Код МКБ-10',
                                    'Номер зуба': 'Номер зуба (для стоматологических услуг)',
                                    '№ г/п дата': '№ ГП'}, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата начала оказания услуги'].notnull()]
    # Deleted columns
    Sheet_Name_Here.drop(['Unnamed: 0', 'СТОИ-МОСТЬ'], axis=1, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def kamelia_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    for i in range(0, len(wb.sheetnames)):
        sheet_name = wb.sheetnames[i]
        try:
            if '№' in sheet_name.cell('B1').value:
                sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
                Sheet_Name_Here = sheet_df_dictonary[sheet_name]
                Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Фамилия'].str.contains('4', na=False)]
                df = pd.DataFrame.from_dict(Sheet_Name_Here)
                df.to_excel(file_name)
                look_data(file_name, file_path)
            else:
                sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=17)
                Sheet_Name_Here = sheet_df_dictonary[sheet_name]
                Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Фамилия'].str.contains('4', na=False)]
                # columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
                # Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
                # columns_to_fill_nan = ['Дата окончания оказания услуги']
                # Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='bfill')
                # df = pd.DataFrame.from_dict(Sheet_Name_Here)
                # repair_dict = df.to_dict()
                # for x in repair_dict['Наименование услуги']:
                #     if 'Итого со скидкой' in x:
                #         discount_size = x.replace('Итого со скидкой ', '', x.replace('%:', '', x))
                #         break
                #     else:
                #         discount_size = None
                #         continue
                # repair_dict['Скидка'] = {x: discount_size for x in repair_dict['Количество'].keys()}
                # df = pd.DataFrame(repair_dict)
                # df = df[df['Количество'].notnull()]
                df = pd.DataFrame.from_dict(Sheet_Name_Here)
                df.to_excel(file_name)
                look_data(file_name, file_path)
        except:
            sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=17)
            Sheet_Name_Here = sheet_df_dictonary[sheet_name]
            Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Фамилия'].str.contains('4', na=False)]
            df = pd.DataFrame.from_dict(Sheet_Name_Here)
            df.to_excel(file_name)
            look_data(file_name, file_path)
    return True


def lazur(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if isinstance(wb[sheet_name].cell(3, 1).value, str):
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['КОЛ-ВО'].notnull()]
        columns_to_fill_nan = ['Ф.И.О.', 'НОМЕР КАРТЫ', 'ДАТА ОКАЗАНИЯ УСЛУГИ', 'ВРАЧ']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        # Надо номера зуба костылить
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Количество"].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def lrc_lomonosovskiy(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Количество"].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medas_samara(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Код услуги"].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medas_perm(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["% Скидка"].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mgu_lomonosova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    try:
        for i in range(0, len(wb.worksheets)):
            sheet_name = wb.sheetnames[i]
            sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
            Sheet_Name_Here = sheet_df_dictonary[sheet_name]
            Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во услуг  "].notnull()]
            Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Наименование услуги'].str.contains(7, na=False)]
            df = pd.DataFrame.from_dict(Sheet_Name_Here)
            df.to_excel(file_name)
            look_data(file_name, file_path)
    except:
        sheet_name = wb.sheetnames[0]
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["Кол-во услуг  "].notnull()]
        Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Наименование услуги'].str.contains('7', na=False)]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def nemeckaya_clinika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'рождения': 'Дата рождения'}, inplace=True)
    Sheet_Name_Here.rename(columns={'оказания': 'Дата оказания услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'услуг': 'Кол-во услуг'}, inplace=True)
    Sheet_Name_Here.rename(columns={'единицу': 'Цена за единицу'}, inplace=True)
    Sheet_Name_Here.rename(columns={' ': ' Цена для СК'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата рождения'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nauka(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер и дата гарантийного письма': '№ ГП'}, inplace=True)

    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата выписки': 'Дата окончания оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость по прейскуранту': 'Цена услуги'}, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Changed Дата окончания оказания услуги to dtype datetime
    Sheet_Name_Here['Дата окончания оказания услуги'] = pd.to_datetime(
        Sheet_Name_Here['Дата окончания оказания услуги'], unit='s',
        errors='coerce')

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in Таблица_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def data_pandas_24mrt(file_name, file_path):
    data_list = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    next_service = False
    next_patient = False
    new_service = False
    need_to_break = False
    guarantee_letter = None
    temp_fio = None
    policy_number = None
    temp_date = None
    temp_doctor = None
    for row_number in range(1, sheet.max_row + 2):
        if need_to_break:
            break
        for column_number in range(1, sheet.max_column + 1):
            inform_value = sheet.cell(row=row_number, column=column_number).value
            if inform_value is not None:
                if inform_value == 'Место оказания услуги':
                    next_patient = True
                    break
                if 'Итого по пациенту' in str(inform_value):
                    next_service = False
                    next_patient = True
                    break
                if 'Итого по Реестру' in str(inform_value):
                    need_to_break = True
                    break
            if next_service:
                if column_number == 3:
                    if new_service:
                        data_list['fio'].append(temp_fio)
                        data_list['policy_number'].append(policy_number)
                        data_list['guarantee_letter'].append(guarantee_letter)
                        data_list['date'].append(temp_date)
                        data_list['doctor_name'].append(temp_doctor)
                        continue
                elif column_number == 10:
                    data_list['service_code'].append(inform_value)
                    continue
                elif column_number == 11:
                    data_list['mkb'].append(inform_value)
                    continue
                elif column_number == 12:
                    data_list['service_name'].append(inform_value)
                    continue
                elif column_number == 13:
                    data_list['service_amount'].append(inform_value)
                    continue
                elif column_number == 14:
                    data_list['service_price'].append(inform_value)
                elif column_number == 14:
                    data_list['total_price'].append(inform_value)
                    new_service = True
                    break
            if next_patient:
                if column_number == 2:
                    data_list['fio'].append(inform_value)
                    temp_fio = inform_value
                    continue
                elif column_number == 3:
                    policy_number = re.sub("^0{3}|^0{4}|^0{2}|^0{5}", '',
                                           inform_value.replace('ООО "БестДоктор" Полис №/', ''))
                    data_list['policy_number'].append(policy_number)
                    continue
                elif column_number == 4:
                    guarantee_letter = inform_value.split(' ', 2)[0]
                    data_list['guarantee_letter'].append(guarantee_letter)
                    continue
                elif column_number == 7:
                    data_list['date'].append(inform_value)
                    temp_date = inform_value
                    continue
                elif column_number == 9:
                    data_list['doctor_name'].append(inform_value)
                    temp_doctor = inform_value
                    next_patient = False
                    next_service = True
                    break
    return filling_book(data_list, file_name, file_path)


def medtim(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here["'Кол-во услуг"].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medicinskiy_cent_atlas(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'ПОЛИС': 'Номер полиса'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'КОД УСЛУГИ': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'НАИМЕНОВАНИЕ УСЛУГИ': 'Наименование услуги'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'КОЛ.  УСЛУГ': 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'ЦЕНА': 'Цена'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'СТОИМОСТЬ': 'Пропуск'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'ДАТА ОКАЗАНИЯ УСЛУГИ ': 'Дата услуги'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'ДИАГНОЗ (текст)': 'Диагноз'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Changed Номер полиса to dtype float
    Sheet_Name_Here['Номер полиса'] = to_float_series(Sheet_Name_Here['Номер полиса'])

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def pirogova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Кол-во услуг': '2'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 27': 'Кол-во услуг'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    Sheet_Name_Here.rename(columns={'Наименование услуги': '1'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 15': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Стоимость, руб.': '3'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 29': 'Стоимость, руб.'}, inplace=True)
    columns_to_fill_nan = ['Фамилия Имя Отчество застрахованного']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def policlinicheskiy_komplex(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    Sheet_Name_Here['Кол-во услуг'] = Sheet_Name_Here['Кол-во услуг'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = Sheet_Name_Here['Цена'].fillna(0).astype('int')
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')
    Sheet_Name_Here.rename(columns={'ID прейскуранта': 'Пропуск'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def semeiniy_med_centr(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if isinstance(wb[sheet_name].cell(6, 1).value, str):
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.rename(columns={'Кол': 'Кол-во'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код.1': '№ зуба'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Уда': 'Уда-ть'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Cумма': 'Cумма к оплате'}, inplace=True)
        Sheet_Name_Here.insert(1, 'new-column-2lcz', 0)
        Sheet_Name_Here.rename(columns={'new-column-2lcz': 'ФИО'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Фамилия': 'Доктор'}, inplace=True)
        Sheet_Name_Here['ФИО'] = IF(SUBSTITUTE(Sheet_Name_Here['№'], '', '') == ' ', None,
                                    Sheet_Name_Here['№'])
        columns_to_fill_nan = ['ФИО']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull()) & (
            ~Sheet_Name_Here['Кол-во'].str.contains('во', na=False))]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№'].apply(
            lambda val: all(s not in str(val) for s in ['Итого по пациенту', 'В ИТОГЕ']))]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        repair_dict = df.to_dict()
        repair_dict['Полис'] = {x: str(y).split(' ', -1)[-1]
                                for x, y in zip(list(repair_dict['ФИО'].keys()), list(repair_dict['ФИО'].values()))}
        repair_dict['ФИО'] = {x: str(y).split(' ', 2)[-1].split(' ', -1)[0] + ' ' +
                                 str(y).split(' ', 2)[-1].split(' ', -1)[1] + ' ' +
                                 str(y).split(' ', 2)[-1].split(' ', -1)[2]
                              for x, y in zip(list(repair_dict['ФИО'].keys()), list(repair_dict['ФИО'].values()))}
        data = pd.DataFrame(repair_dict)
        data.to_excel(file_name)
        return look_data(file_name, file_path)
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)

def city_clinik(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Диагноз \n(код по МКБ-10)': 'Диагноз'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Стоимость, руб': 'Цена'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Кол-во'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Общая сумма, руб': 'Пропуск'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def sk_unit(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['\nФИО'].str.contains('ИТОГО', na=False)]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def son_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    columns_to_fill_nan = ['№', 'ФИО', 'Номер полиса', 'Дата']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomat_poli_9_kazan(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг  '].notnull()]
    columns_to_fill_nan = ['Номер медицинской карты /истории болезни', 'Фамилия Застрахованного', 'Имя Застрахованного',
                           'Отчество Застрахованного ', '№', 'Дата рождения', '№ Полиса ДМС', 'Дата оказания услуги ']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def super_medic(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(2, 'new-column-6hgg', 0)
    Sheet_Name_Here.rename(columns={'new-column-6hgg': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Ф.И.О., возраст (лет) ': 'возраст'}, inplace=True)
    Sheet_Name_Here['ФИО'] = LEFT(Sheet_Name_Here['возраст'], INT(FIND(Sheet_Name_Here['возраст'], '(') - 2))
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    Sheet_Name_Here.drop(['№\nп/п'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Unnamed: 6'], axis=1, inplace=True)
    Sheet_Name_Here['Кол-во услуг'] = Sheet_Name_Here['Кол-во услуг'].fillna(0).astype('int')
    Sheet_Name_Here['Стоимость услуги'] = to_int_series(Sheet_Name_Here['Стоимость услуги'])
    Sheet_Name_Here['Общая стоимость'] = to_int_series(Sheet_Name_Here['Общая стоимость'])
    Sheet_Name_Here['№ полиса или № по базе'] = Sheet_Name_Here['№ полиса или № по базе'].fillna(0).astype('int')
    Sheet_Name_Here.rename(columns={'№ полиса или № по базе': 'Номер полиса '}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата предоставления услуги': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Кол-во'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ugmk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    columns_to_fill_nan = ['Номер полиса', 'Ф.И.О. пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fnkc_fmba(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    columns_to_fill_nan = ['№ п/п', 'Фамилия Имя Отчество застрахованного', '№ полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def formula_ulibki(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    columns_to_fill_nan = ['№ п/п', 'ФИО пациента', '№ индивидуальной карты пациента', 'Дата (период) оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def forum_internationla_technolog(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def centr_semeynoy_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис\nДМС': 'Номер полиса'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код\nМКВ': 'МКБ'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата\nобращения': 'Дата услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код\nуслуги': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуга': 'Наименование услуги'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Сумма': 'Пропуск'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'Врач, фамилия': 'Врач'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Номер полиса', 'ФИО', 'Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_ckb_s_poliklinikoy(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if 'Стационарный' in wb[sheet_name].cell(1, 3).value:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    else:
        return  # Ручная часть


def crmt_Novoslobotskaya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={1: 'ФИО',
                                    2: 'Номер полиса',
                                    3: 'Дата услуги',
                                    4: 'Код услуги',
                                    5: 'МКБ',
                                    6: 'Наименование услуги',
                                    7: 'Кол-во',
                                    8: 'Цена'}, inplace=True)
    # Filled NaN values in columns
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Кол-во', na=False))]
    # Changed dtype
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])
    Sheet_Name_Here['Номер полиса'] = to_float_series(Sheet_Name_Here['Номер полиса'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def csvmp(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl',
                                       sheet_name=[sheet_name[0]],
                                       skiprows=6, converters={'Код услуги': str})
    Sheet_Name_Here = sheet_df_dictonary[sheet_name[0]]
    Sheet_Name_Here.rename(columns={'№ Гарантийного письма': 'ГП'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата гарантийного письмо': 'Дата ГП'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Фамилия врача': 'Врач'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Цена услуги': 'Цена'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'№ индидуальной карты (ID) пациента': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')
    Sheet_Name_Here['Номер полиса'] = Sheet_Name_Here['Номер полиса'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def estetika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(1, 'new-column-arme', 0)
    Sheet_Name_Here.insert(1, 'new-column-e7o4', 0)
    Sheet_Name_Here.rename(columns={'new-column-e7o4': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-arme': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={2: 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={3: 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={5: 'Стоимость услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={6: 'Кол-во'}, inplace=True)
    Sheet_Name_Here.insert(7, 'new-column-vnp9', 0)
    Sheet_Name_Here.insert(7, 'new-column-w5xl', 0)
    Sheet_Name_Here.rename(columns={'new-column-w5xl': 'Диагноз'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-vnp9': 'Номер зуба'}, inplace=True)
    Sheet_Name_Here.rename(columns={1: 'Пропуск'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = IF(AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск']), ' ', '') != ''),
                               Sheet_Name_Here['Пропуск'], None)
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск']), ' ', '') != ''), None,
                      Sheet_Name_Here['Пропуск'])
    Sheet_Name_Here.rename(columns={0: 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={4: 'Пропуск2'}, inplace=True)
    Sheet_Name_Here['Номер зуба'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN', SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск2']), ' ', ''), '.', '') != ''),
        CLEAN(Sheet_Name_Here['Пропуск2']), None)
    Sheet_Name_Here.insert(8, 'new-column-1hj3', 0)
    Sheet_Name_Here.rename(columns={'new-column-1hj3': 'Пропуск3'}, inplace=True)
    Sheet_Name_Here['Пропуск3'] = IF(AND(TYPE(Sheet_Name_Here['Номер зуба']) != 'NaN',
                               SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск2']), ' ', ''), '.', '') != ''),
                           SUBSTITUTE(Sheet_Name_Here['Пропуск2'], Sheet_Name_Here['Номер зуба'], ''), Sheet_Name_Here['Пропуск2'])
    Sheet_Name_Here['Диагноз'] = IF(Sheet_Name_Here['Пропуск3'] == '', None, Sheet_Name_Here['Пропуск3'])
    columns_to_fill_nan = ['ФИО', 'Дата услуги', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Стоимость услуги'] = Sheet_Name_Here['Стоимость услуги'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def data_hospital_centrosoyuz_rf(file_name, file_path):
    data_list = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    next_patient = False
    next_service = False
    need_to_copy_paste = False
    next_policy_number = False
    need_to_break = False
    temp_fio = None
    policy_number = None
    date = None
    for row_number in range(30, sheet.max_row + 2):
        if need_to_break:
            break
        for column_number in range(1, sheet.max_column + 1):
            inform_value = sheet.cell(row=row_number, column=column_number).value
            if inform_value is not None:
                if 'Страховая программа:' in str(inform_value):
                    next_patient = True
                    break
                if 'Итого:' in str(sheet.cell(row=row_number, column=6).value):
                    next_patient = True
                    need_to_copy_paste = False
                    next_service = False
                    next_policy_number = False
                    break
                if 'Всего:' in str(sheet.cell(row=row_number, column=6).value):
                    need_to_break = True
                    break
                if next_patient:
                    if column_number == 2:
                        temp_fio = inform_value
                        data_list['fio'].append(temp_fio)
                        next_policy_number = True
                        next_patient = False
                        break
                if next_policy_number:
                    if column_number == 4:
                        policy_number = re.sub('^0{2}|^0{3}|^0{4}|^0{5}', '', str(inform_value))
                        data_list['policy_number'].append(policy_number)
                        next_policy_number = False
                        break
                if column_number == 2 and str(inform_value) == 'Дата':
                    next_service = True
                    break
                if need_to_copy_paste:
                    data_list['fio'].append(temp_fio)
                    data_list['policy_number'].append(policy_number)
                    need_to_copy_paste = False
                if next_service:
                    if column_number == 2:
                        if inform_value is None:
                            data_list['date'].append(date)
                        else:
                            date = inform_value
                            data_list['date'].append(date)
                        continue
                    if column_number == 3:
                        data_list['mkb'].append(inform_value)
                        continue
                    if column_number == 4:
                        data_list['service_name'].append(inform_value)
                        continue
                    if column_number == 5:
                        data_list['service_code'].append(inform_value)
                        continue
                    if column_number == 6:
                        data_list['service_amount'].append(inform_value)
                        continue
                    if column_number == 7:
                        data_list['service_price'].append(inform_value)
                        need_to_copy_paste = True
                        continue
    return clear_data(data_list, file_name, file_path)


def data_pandas_izumrud(file_name, file_path):
    data_list = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    check_fio = None
    policy_check = None
    patient_check = False
    need_to_skip = False
    need_to_break = False
    need_to_copy_paste = False
    done = False
    for row_number in range(1, (sheet.max_row)):
        if need_to_break:
            need_to_break = False
            break
        if need_to_skip:
            need_to_skip = False
            continue
        for column_number in range(2, sheet.max_column):
            inform = sheet.cell(row=row_number, column=column_number)
            inform_value = inform.value
            if not done:
                try:
                    if 'Итого оказано услуг:' in inform_value:
                        need_to_break = True
                        break
                    if 'Пациент:' in inform_value and column_number == 2:
                        check_fio = str(inform_value).replace('Пациент: ', '')
                        patient_check = True
                        data_list['fio'].append(str(check_fio))
                        break
                    if 'Полис:' in inform_value and column_number == 2 and patient_check:
                        policy_check = re.sub('^0{3}', '', str(inform_value).replace('Полис: ', '').replace('л', ''))
                        data_list['policy_number'].append(policy_check)
                        break
                    if inform_value == 'Код услуги':
                        done = True
                        break
                except:
                    continue
            else:
                if column_number == 2 and inform_value == 'Итого:':
                    done = False
                    patient_check = False
                    need_to_copy_paste = False
                    break
                if need_to_copy_paste:
                    need_to_copy_paste = False
                    data_list['fio'].append(str(check_fio))
                    data_list['policy_number'].append(str(policy_check))
                if column_number == 2:
                    data_list['service_code'].append(str(inform_value))
                    continue
                if column_number == 3:
                    data_list['service_name'].append(str(inform_value))
                    continue
                if column_number == 6:
                    data_list['service_amount'].append(str(inform_value))
                    continue
                if column_number == 7:
                    data_list['service_price'].append(str(inform_value))
                    continue
                if column_number == 8:
                    data_list['date'].append(str(inform_value))
                    need_to_copy_paste = True
                    break
    return clear_data(data_list, file_name, file_path)


def center_sovremennoy_medecini(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО застрахован.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Диагноз, номер зуба': 'Диагноз'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Наименование услуги по прейскуранту ': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns №п/п
    Sheet_Name_Here.drop(['№п/п'], axis=1, inplace=True)

    # Deleted columns Поверхности
    Sheet_Name_Here.drop(['Поверхности'], axis=1, inplace=True)

    # Deleted columns Ст-ть услуг (руб.)
    Sheet_Name_Here.drop(['Ст-ть услуг (руб.)'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def set_semeynih_med_centrov(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_path, engine='openpyxl', sheet_name=[sheet_name], skiprows=5, header=None)
    list_1 = sheet_df_dictonary[sheet_name]
    list_1 = list_1[list_1[1].notnull()]
    df = pd.DataFrame.from_dict(list_1)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_nikio_im_l_i_sverzhevckogo(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол- во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def apex(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Unnamed: 17'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def dems(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Filtered Код по прайсу
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код по прайсу'].notnull()]

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Дата рождения']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Полис to dtype str
    Sheet_Name_Here['Полис'] = Sheet_Name_Here['Полис'].astype('str')

    # Deleted columns Код услуги
    Sheet_Name_Here.drop(['Код услуги'], axis=1, inplace=True)

    # Changed кол. услуг to dtype int
    Sheet_Name_Here['кол. услуг'] = to_int_series(Sheet_Name_Here['кол. услуг'])

    # Deleted columns     стоимость
    Sheet_Name_Here.drop(['    стоимость'], axis=1, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата ': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код по прайсу': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'кол. услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'цена': 'Цена услуги'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def kds(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество услуг'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def klinika_dobrogo_stomatologa(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество услуг'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mnogoprofilnii_medecinskii_center_dialain(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Медцентр': 'Наименование филиала клиники'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def poliklinika_2_vita_medicus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def rma(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Время  (кол-во)'].notnull()]
    columns_to_fill_nan = ['№ п/п', 'Дата', 'ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def sogaz_medservis(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата оказания услуг': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'ФИО Пациента': 'ФИО'}, inplace=True)
    Sheet_Name_Here.drop(['Дата рождения Пациента'], axis=1, inplace=True)
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.drop(['Заказчик'], axis=1, inplace=True)
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here['Номер полиса'] = Sheet_Name_Here['Номер полиса'].fillna(0).astype('int')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги '].notnull()]
    columns_to_fill_nan = ['Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here['Дата услуги'] = pd.to_datetime(Sheet_Name_Here['Дата услуги'],
                                                         infer_datetime_format=True, errors='coerce')
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = Sheet_Name_Here['Цена'].fillna(0).astype('int')
    Sheet_Name_Here['Стоимость \n(в руб.)'] = Sheet_Name_Here['Стоимость \n(в руб.)'].fillna(0).astype('int')
    Sheet_Name_Here.insert(7, 'new-column-9zal', 0)
    Sheet_Name_Here.rename(columns={'Кол-во': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-9zal': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Стоимость \n(в руб.)': 'Итого'}, inplace=True)
    Sheet_Name_Here['Кол-во'] = INT(Sheet_Name_Here['Итого'] / Sheet_Name_Here['Цена'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatkomplex_olimpia_perm(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatbiznes_company(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена (руб)'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def efa(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if isinstance(wb[sheet_name].cell(2, 1).value, str):
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
        columns_to_fill_nan = ['Дата ', 'Пациент', 'Полис']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена, руб.'].notnull()]
        columns_to_fill_nan = ['№', '№ полиса', 'Застрахованный', 'Номер мед.карты']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def botkinskaya_33(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vale_dental(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[
            (Sheet_Name_Here['Кол-во услуг'].notnull()) & (~Sheet_Name_Here['Кол-во услуг'].str.contains('Кол-во услуг', na=False))]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def guten_tag(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    Sheet_Name_Here.rename(columns={'Дата приема': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = Sheet_Name_Here['Цена'].fillna(0).astype('int')
    Sheet_Name_Here.rename(columns={'Номер ID-карты': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = Sheet_Name_Here['Номер полиса'].fillna(0).astype('int')
    Sheet_Name_Here.rename(columns={'Артикул': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.insert(7, 'new-column-g3b8', 0)
    Sheet_Name_Here.rename(columns={'new-column-g3b8': 'МКБ'}, inplace=True)
    Sheet_Name_Here.insert(8, 'new-column-tddv', 0)
    Sheet_Name_Here.rename(columns={'Диагноз': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-tddv': 'Диагноз'}, inplace=True)
    Sheet_Name_Here['МКБ'] = IF ( LEFT ( Sheet_Name_Here['Пропуск'], INT ( FIND ( Sheet_Name_Here['Пропуск'], ' ', ) )) != 'nan' , LEFT ( Sheet_Name_Here['Пропуск'], INT ( FIND ( Sheet_Name_Here['Пропуск'], ' ', ) )) ,None )
    Sheet_Name_Here['Диагноз'] = IF ( SUBSTITUTE ( Sheet_Name_Here['Пропуск'], Sheet_Name_Here['МКБ'], '') != 'nan' , SUBSTITUTE(Sheet_Name_Here['Пропуск'], Sheet_Name_Here['МКБ'], ''), None )
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def doctor_ryadom(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.insert(2, 'new-column-3uxs', 0)
        Sheet_Name_Here.rename(columns={'ГП': 'Пропуск'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-3uxs': 'ГП'}, inplace=True)
        Sheet_Name_Here['ГП'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                 LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], ' '))), None)
        Sheet_Name_Here.rename(columns={'Исполнитель': 'Врач'}, inplace=True)
        Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)
        Sheet_Name_Here.rename(columns={'Зуб': 'Номер зуба'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)
        Sheet_Name_Here.insert(5, 'new-column-fjsi', 0)
        Sheet_Name_Here.rename(columns={'new-column-fjsi': 'Специальность врача'}, inplace=True)
        Sheet_Name_Here['Специальность врача'] = IF(
            AND(TYPE(Sheet_Name_Here['Врач']) != 'NaN', INT(FIND(Sheet_Name_Here['Врач'], '_')) > 0),
            SUBSTITUTE(Sheet_Name_Here['Врач'], LEFT(Sheet_Name_Here['Врач'], INT(FIND(Sheet_Name_Here['Врач'], '_'))), ''), None)
        Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата услуги'}, inplace=True)
        Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
        Sheet_Name_Here.rename(columns={'Филиал': 'Наименование филиала клиники'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Врач': 'Пропуск2'}, inplace=True)
        Sheet_Name_Here.insert(5, 'new-column-z5vk', 0)
        Sheet_Name_Here.rename(columns={'new-column-z5vk': 'Врач'}, inplace=True)
        Sheet_Name_Here['Врач'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск2'],
                                                                                        SUBSTITUTE(
                                                                                            Sheet_Name_Here['Пропуск2'],
                                                                                            LEFT(Sheet_Name_Here['Пропуск2'],
                                                                                                 INT(FIND(Sheet_Name_Here[
                                                                                                              'Пропуск2'],
                                                                                                          '_') - 1)), ''),
                                                                                        ''), None)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.insert(2, 'new-column-3uxs', 0)
        Sheet_Name_Here.rename(columns={'ГП': 'Пропуск'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-3uxs': 'ГП'}, inplace=True)
        Sheet_Name_Here['ГП'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                   LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], ' '))), None)
        Sheet_Name_Here.rename(columns={'Исполнитель': 'Врач'}, inplace=True)
        Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)
        Sheet_Name_Here.rename(columns={'Зуб': 'Номер зуба'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)
        Sheet_Name_Here.insert(5, 'new-column-fjsi', 0)
        Sheet_Name_Here.rename(columns={'new-column-fjsi': 'Специальность врача'}, inplace=True)
        Sheet_Name_Here['Специальность врача'] = IF(
            AND(TYPE(Sheet_Name_Here['Врач']) != 'NaN', INT(FIND(Sheet_Name_Here['Врач'], '_')) > 0),
            SUBSTITUTE(Sheet_Name_Here['Врач'], LEFT(Sheet_Name_Here['Врач'], INT(FIND(Sheet_Name_Here['Врач'], '_'))),
                       ''), None)
        Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата услуги'}, inplace=True)
        Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
        Sheet_Name_Here.rename(columns={'Филиал': 'Наименование филиала клиники'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Врач': 'Пропуск2'}, inplace=True)
        Sheet_Name_Here.insert(5, 'new-column-z5vk', 0)
        Sheet_Name_Here.rename(columns={'new-column-z5vk': 'Врач'}, inplace=True)
        Sheet_Name_Here['Врач'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск2'],
                                                                                            SUBSTITUTE(
                                                                                                Sheet_Name_Here[
                                                                                                    'Пропуск2'],
                                                                                                LEFT(Sheet_Name_Here[
                                                                                                         'Пропуск2'],
                                                                                                     INT(FIND(
                                                                                                         Sheet_Name_Here[
                                                                                                             'Пропуск2'],
                                                                                                         '_') - 1)),
                                                                                                ''),
                                                                                            ''), None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def evropeiskiy_medecinskiy_center_egmk_zdorovie(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    if isinstance(wb[sheet_name].cell(6, 1).value, str):
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    else:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    columns_to_fill_nan = ['Номер полиса', 'Ф.И.О. пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(
        method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def osnova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15, converters={'Код услуги': str})
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.insert(3, 'new-column-9d3u', 0)
        Sheet_Name_Here.rename(columns={'new-column-9d3u': 'ФИО'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Фамилия': 'ПропускФ'}, inplace=True)
        Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['ПропускФ']) != 'NaN', FIND(Sheet_Name_Here['№ п/п'], '.') > 0), Sheet_Name_Here['ПропускФ'],
                            None)
        Sheet_Name_Here.rename(columns={'Имя': 'ПропускИ'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Отчество': 'ПропускО'}, inplace=True)
        Sheet_Name_Here.insert(7, 'new-column-3koq', 0)
        Sheet_Name_Here.rename(columns={'Серия и номер полиса': 'ПропускП'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-3koq': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.insert(8, 'new-column-hqif', 0)
        Sheet_Name_Here.insert(8, 'new-column-6c5r', 0)
        Sheet_Name_Here['new-column-hqif'] = IF(TYPE(Sheet_Name_Here['ПропускП']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['ПропускП'],
                                                                                       LEFT(Sheet_Name_Here['ПропускП'],
                                                                                            INT(FIND(Sheet_Name_Here['ПропускП'],
                                                                                                     '№'))), ''), None)
        Sheet_Name_Here['new-column-6c5r'] = IF(TYPE(Sheet_Name_Here['new-column-hqif']) != 'NaN', FLOAT(Sheet_Name_Here['new-column-hqif']), None)
        Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['new-column-6c5r']) != 'NaN', TEXT(INT(Sheet_Name_Here['new-column-6c5r'])),
                                     Sheet_Name_Here['new-column-hqif'])
        columns_to_fill_nan = ['ФИО', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
        Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
        Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,
                                           converters={'Код услуги': str})
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.insert(3, 'new-column-9d3u', 0)
        Sheet_Name_Here.rename(columns={'new-column-9d3u': 'ФИО'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Фамилия': 'ПропускФ'}, inplace=True)
        Sheet_Name_Here['ФИО'] = IF(
            AND(TYPE(Sheet_Name_Here['ПропускФ']) != 'NaN', FIND(Sheet_Name_Here['№ п/п'], '.') > 0),
            Sheet_Name_Here['ПропускФ'],
            None)
        Sheet_Name_Here.rename(columns={'Имя': 'ПропускИ'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Отчество': 'ПропускО'}, inplace=True)
        Sheet_Name_Here.insert(7, 'new-column-3koq', 0)
        Sheet_Name_Here.rename(columns={'Серия и номер полиса': 'ПропускП'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-3koq': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.insert(8, 'new-column-hqif', 0)
        Sheet_Name_Here.insert(8, 'new-column-6c5r', 0)
        Sheet_Name_Here['new-column-hqif'] = IF(TYPE(Sheet_Name_Here['ПропускП']) != 'NaN',
                                                SUBSTITUTE(Sheet_Name_Here['ПропускП'],
                                                           LEFT(Sheet_Name_Here['ПропускП'],
                                                                INT(FIND(Sheet_Name_Here['ПропускП'],
                                                                         '№'))), ''), None)
        Sheet_Name_Here['new-column-6c5r'] = IF(TYPE(Sheet_Name_Here['new-column-hqif']) != 'NaN',
                                                FLOAT(Sheet_Name_Here['new-column-hqif']), None)
        Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['new-column-6c5r']) != 'NaN',
                                             TEXT(INT(Sheet_Name_Here['new-column-6c5r'])),
                                             Sheet_Name_Here['new-column-hqif'])
        columns_to_fill_nan = ['ФИО', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
        Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
        Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def klinika_nemeckoy_stomatologii_guten_tag(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def medecina_alfastrahovaniya_perm(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)


    return look_data(file_name, file_path)
def medecina_alfastrahovaniya_samara(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['№'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medecina_alfastrahovaniya_yaroslavl(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mir_ulibok(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=14)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['\nФИО'].apply(lambda val: all(s not in str(val) for s in ['ИТОГО', '3']))]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def panaceya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'наименование услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'дата оказания': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'диагноз': 'МКБ'}, inplace=True)

    # Added column new-column-n34l
    Sheet_Name_Here.insert(8, 'new-column-n34l', 0)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-n34l': 'Цена'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'стоимость': 'Пропуск'}, inplace=True)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', TYPE(Sheet_Name_Here['кол-во']) != 'NaN'),
        Sheet_Name_Here['Пропуск'] / INT(Sheet_Name_Here['кол-во']), None)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Номер полиса', 'ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['кол-во'].notnull()]

    # Changed кол-во to dtype int
    Sheet_Name_Here['кол-во'] = Sheet_Name_Here['кол-во'].fillna(0).astype('int')

    # Set formula of кол-во
    Sheet_Name_Here['Цена'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', TYPE(Sheet_Name_Here['кол-во']) != 'NaN'),
        Sheet_Name_Here['Пропуск'] / INT(Sheet_Name_Here['кол-во']), None)

    # Added column new-column-6edq
    Sheet_Name_Here.insert(4, 'new-column-6edq', 0)

    # Added column new-column-dlha
    Sheet_Name_Here.insert(4, 'new-column-dlha', 0)

    # Renamed columns Пропуск2
    Sheet_Name_Here.rename(columns={'Наименование услуги': 'Пропуск2'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'new-column-dlha': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'new-column-6edq': 'Код услуги'}, inplace=True)

    # Set formula of Наименование услуги
    Sheet_Name_Here['Наименование услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN',
                                                SUBSTITUTE(Sheet_Name_Here['Пропуск2'],
                                                           LEFT(Sheet_Name_Here['Пропуск2'],
                                                                INT(FIND(Sheet_Name_Here['Пропуск2'], ' '))), ''), None)

    # Set formula of Код услуги
    Sheet_Name_Here['Код услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN',
                                       SUBSTITUTE(Sheet_Name_Here['Пропуск2'], Sheet_Name_Here['Наименование услуги'],
                                                  ''), None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def perviy_doctor(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Стоимость услуги'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Сумма'].notnull()]
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def prioritet(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Номер полиса'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomat_studia_vash_doctor(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatologicheskay_poliklinika_dento_smail(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        columns_to_fill_nan = ['ФИО                   пациента', 'Дата рождения', 'Номер', 'Дата оказания услуги']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
        Sheet_Name_Here.rename(columns={'Номер': 'Номер полиса'}, inplace=True)
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.rename(columns={'Unnamed: 0': 'ФИО'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 1': 'дата рождения'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Пациента': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 3': 'Дата'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 4': 'Код услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 5': 'Наименование услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 6': 'МКБ-10'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 7': 'Диагноз'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 8': 'Цена, руб.'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Кол-во'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 10': 'Сумма, руб.'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 11': '№ ГП'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 12': 'Специальность доктора'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Unnamed: 13': 'Доктор'}, inplace=True)
        df = pd.DataFrame.from_dict(Sheet_Name_Here)
        df.to_excel(file_name)
        return look_data(file_name, file_path)


def stomatologicheskay_poliklinika_vizavi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['\nФИО'].str.contains('ИТОГО', na=False)]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatologicheskay_poliklinika_9_goroda_kazani(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    for i in range(0, len(wb.sheetnames)):
        try:
            sheet_name = wb.sheetnames[i]
            sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
            Sheet_Name_Here = sheet_df_dictonary[sheet_name]
            columns_to_fill_nan = ['Номер медицинской карты /истории болезни', 'Фамилия Застрахованного',
                                   'Имя Застрахованного', 'Отчество Застрахованного ', 'Дата рождения',
                                   '№ Полиса ДМС', 'Дата оказания услуги ']
            Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
            Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг  '].notnull()]
            df = pd.DataFrame.from_dict(Sheet_Name_Here)
            df.to_excel(file_name)
            look_data(file_name, file_path)
        except:
            sheet_name = wb.sheetnames[i]
            sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,
                                               header=None)
            Sheet_Name_Here = sheet_df_dictonary[sheet_name]
            columns_to_fill_nan = [1, 2, 3, 4, 5, 6, 7]
            Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
            Sheet_Name_Here.rename(columns={8: 'Код услуги'}, inplace=True)
            Sheet_Name_Here.rename(columns={7: 'Дата оказания услуги '}, inplace=True)
            Sheet_Name_Here.rename(columns={6: '№ Полиса ДМС'}, inplace=True)
            Sheet_Name_Here.rename(columns={5: 'Дата рождения'}, inplace=True)
            Sheet_Name_Here.rename(columns={4: 'Отчество Застрахованного '}, inplace=True)
            Sheet_Name_Here.rename(columns={3: 'Имя Застрахованного'}, inplace=True)
            Sheet_Name_Here.rename(columns={2: 'Фамилия Застрахованного'}, inplace=True)
            Sheet_Name_Here.rename(columns={1: 'Номер медицинской карты /истории болезни'}, inplace=True)
            Sheet_Name_Here.rename(columns={0: '№'}, inplace=True)
            Sheet_Name_Here.rename(columns={9: 'Наименование услуги'}, inplace=True)
            Sheet_Name_Here.rename(columns={10: 'Стоимость услуг (Цена*Количество)'}, inplace=True)
            Sheet_Name_Here.rename(columns={11: 'Кол-во услуг  '}, inplace=True)
            Sheet_Name_Here.rename(columns={12: 'Цена услуги по прейскуранту  '}, inplace=True)
            Sheet_Name_Here.rename(columns={13: 'Код диагноза по МКБ10 (не менее четырех знаков)'}, inplace=True)
            Sheet_Name_Here.rename(columns={14: 'Зуб - номер'}, inplace=True)
            Sheet_Name_Here.rename(columns={15: 'Специальность врача оказавшего услугу  '}, inplace=True)
            Sheet_Name_Here.rename(columns={16: 'ФИО врача, оказавшего услугу'}, inplace=True)
            Sheet_Name_Here.rename(columns={17: 'Клиника-Исполнитель (при Сети Клиник)'}, inplace=True)
            Sheet_Name_Here.rename(columns={18: 'дата/номер гарантийного письма '}, inplace=True)
            Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во услуг  '].notnull()) & (~Sheet_Name_Here['Кол-во услуг  '].str.contains('Кол-во услуг  ', na=False))]
            df = pd.DataFrame.from_dict(Sheet_Name_Here)
            df.to_excel(file_name)
    return True


def stomatologicheskiy_center(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(6, 'new-column-sa9g', 0)
    Sheet_Name_Here.insert(6, 'new-column-hjev', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={1: 'ФИО',
                                    2: 'Номер полиса',
                                    3: 'Дата услуги',
                                    4: '№ Зуба',
                                    'new-column-hjev': 'МКБ',
                                    'new-column-sa9g': 'Диагноз',
                                    5: 'Пропуск',
                                    6: 'Код услуги',
                                    7: 'Наименование услуги',
                                    8: 'Врач',
                                    9: 'Кол-во',
                                    10: 'Цена',
                                    11: 'Пропуск000'}, inplace=True)
    # Set formula
    Sheet_Name_Here['МКБ'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], ' '))),
                      None)
    Sheet_Name_Here['Диагноз'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск'], Sheet_Name_Here['МКБ'], ''), None)
    # Filled NaN values in columns
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull())
                                      & (Sheet_Name_Here['Кол-во'].apply(lambda val: all(s not in str(val) for s in ['всего', 'Количество оказанных услуг', 'итого'])))]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    # Changed dtype
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatoshka(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-jocb
    Sheet_Name_Here.insert(7, 'new-column-jocb', 0)

    # Added column new-column-l9bc
    Sheet_Name_Here.insert(7, 'new-column-l9bc', 0)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Диагноз, номер зуба': 'Пропуск'}, inplace=True)

    # Set formula of new-column-l9bc
    Sheet_Name_Here['new-column-l9bc'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск'], SUBSTITUTE(
        SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск']), '(', ''), ')', ''), ' ', ''), ''), None)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'new-column-l9bc': 'Диагноз'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={'new-column-jocb': '№ зуба'}, inplace=True)

    # Set formula of № зуба
    Sheet_Name_Here['№ зуба'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                         SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск']), '(', ''), ')', ''), ' ', ''), None)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Наименование услуги по прейскуранту ': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена услуги ': 'Цена'}, inplace=True)

    # Renamed columns Пропуск2ъ
    Sheet_Name_Here.rename(columns={'Ст-ть услуг (руб.)': 'Пропуск2ъ'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'ФИО застрахован. ': 'ФИО'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код по МКБ-10': 'МКБ'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def samarskiy_dia_centr(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.drop(['номер строки'], axis=1, inplace=True)
    Sheet_Name_Here.insert(1, 'new-column-hqlz', 0)
    Sheet_Name_Here.rename(columns={'new-column-hqlz': 'ФИО'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF ( TYPE ( Sheet_Name_Here['Фамилия Имя Отчество']) != 'NaN' , SUBSTITUTE ( SUBSTITUTE ( SUBSTITUTE ( SUBSTITUTE ( SUBSTITUTE ( SUBSTITUTE ( PROPER ( Sheet_Name_Here['Фамилия Имя Отчество'] ), '  ', ' ') , '  ', ' '), '  ', ' ') , '  ', ' '), '  ', ' ') , '  ', ' ') , None)
    Sheet_Name_Here.drop(['Дата рожд'], axis=1, inplace=True)
    Sheet_Name_Here.insert(2, 'new-column-7mgk', 0)
    Sheet_Name_Here.rename(columns={'new-column-7mgk': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = IF(FIND(Sheet_Name_Here['Адрес'], 'полис:') > 0,
                                INT(LEFT(SUBSTITUTE(Sheet_Name_Here['Адрес'], 'полис: ', ''),
                                         FIND(SUBSTITUTE(Sheet_Name_Here['Адрес'], 'полис: ', ''), ' '))), None)
    Sheet_Name_Here.insert(4, 'new-column-5dz7', 0)
    Sheet_Name_Here.insert(4, 'new-column-txk2', 0)
    Sheet_Name_Here.rename(columns={'new-column-txk2': 'Врач'}, inplace=True)
    Sheet_Name_Here['Врач'] = IF(TYPE(Sheet_Name_Here['Адрес']) != 'NaN', PROPER(Sheet_Name_Here['Адрес']), None)
    Sheet_Name_Here.rename(columns={'new-column-5dz7': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Unnamed: 9'] != 'МКБ10']
    Sheet_Name_Here['Дата услуги'] = IF(TYPE(Sheet_Name_Here['Unnamed: 4']) != 'NaN', DATEVALUE(Sheet_Name_Here['Unnamed: 4']), None)
    Sheet_Name_Here.rename(columns={'Unnamed: 5': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.insert(9, 'new-column-ldd1', 0)
    Sheet_Name_Here.rename(columns={'new-column-ldd1': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here['Наименование услуги'] = IF(TYPE(Sheet_Name_Here['Unnamed: 6']) != 'NaN',
                                       SUBSTITUTE(Sheet_Name_Here['Unnamed: 6'], '- ', ''), None)
    Sheet_Name_Here.rename(columns={'Unnamed: 7': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.insert(12, 'new-column-6k3y', 0)
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-6k3y': 'Цена'}, inplace=True)
    Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['Unnamed: 8']) != 'NaN'),
                        INT(INT(Sheet_Name_Here['Unnamed: 8']) / INT(Sheet_Name_Here['Кол-во'])), None)
    Sheet_Name_Here.insert(15, 'new-column-c4r7', 0)
    Sheet_Name_Here.rename(columns={'new-column-c4r7': 'Скидка'}, inplace=True)
    Sheet_Name_Here['Скидка'] = IF(AND(TYPE(Sheet_Name_Here['Unnamed: 10']) != 'NaN', Sheet_Name_Here['Unnamed: 10'] > 0),
                          Sheet_Name_Here['Unnamed: 10'], None)
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here.rename(columns={'Адрес': 'Пропуск'}, inplace=True)
    Sheet_Name_Here['Наименование услуги'] = IF(TYPE(Sheet_Name_Here['Unnamed: 6']) != 'NaN',
                                       SUBSTITUTE(Sheet_Name_Here['Unnamed: 6'], '-', ''), None)
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Наименование услуги'].notnull()) & (Sheet_Name_Here['Наименование услуги'] != 'Итого:')]
    Sheet_Name_Here.rename(columns={'Фамилия Имя Отчество': 'Пропуск00'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nmictpm_mz_rf(file_name, file_path):
    wb = load_workbook(file_path, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_path, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgb_klinicheskaya_bonica_nomer_1(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=11)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(1, 'new-column-bu6d', 0)
    Sheet_Name_Here.insert(1, 'new-column-wkgv', 0)
    Sheet_Name_Here.rename(columns={'new-column-wkgv': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-bu6d': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 1': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 2': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 3': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 4': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 5': 'Цена'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 6': 'Стоимость'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(SUBSTITUTE(CLEAN(Sheet_Name_Here['Unnamed: 0']), ' ', '') == '',
                                Sheet_Name_Here['Unnamed: 0'], None)
    Sheet_Name_Here['Номер полиса'] = IF(SUBSTITUTE(CLEAN(Sheet_Name_Here['Unnamed: 0']), ' ', '') != '',
                                         Sheet_Name_Here['Unnamed: 0'], None)
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Unnamed: 0'].apply(lambda val: all(
        s not in str(val) for s in ['Пациент:', 'Номер карты:', 'Диагноз основной:', 'Количество процедур:']))]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here = Sheet_Name_Here[(~Sheet_Name_Here['Код услуги'].str.contains('Код услуги', na=False)) & (
                Sheet_Name_Here['Код услуги'] != '***************')]
    Sheet_Name_Here.insert(10, 'new-column-kuzp', 0)
    Sheet_Name_Here.insert(10, 'new-column-2c2v', 0)
    Sheet_Name_Here.rename(columns={'new-column-2c2v': 'Код врача'}, inplace=True)
    Sheet_Name_Here['Код врача'] = LEFT(SUBSTITUTE(Sheet_Name_Here['Unnamed: 7'], LEFT(Sheet_Name_Here['Unnamed: 7'],
                                                                                       FIND(Sheet_Name_Here[
                                                                                                'Unnamed: 7'],
                                                                                            '/') + 1), ''), FIND(
        SUBSTITUTE(Sheet_Name_Here['Unnamed: 7'],
                   LEFT(Sheet_Name_Here['Unnamed: 7'], FIND(Sheet_Name_Here['Unnamed: 7'], '/') + 1), ''), '/') - 2)
    Sheet_Name_Here['new-column-kuzp'] = SUBSTITUTE(Sheet_Name_Here['Unnamed: 7'], CONCAT(
        LEFT(Sheet_Name_Here['Unnamed: 7'],
             FIND(Sheet_Name_Here['Unnamed: 7'], '/') + 1 + FIND(Sheet_Name_Here['Unnamed: 7'], '/') + 1 + FIND(
                 Sheet_Name_Here['Unnamed: 7'], '/') + 4)), '')
    Sheet_Name_Here.rename(columns={'new-column-kuzp': 'МКБ'}, inplace=True)
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    Sheet_Name_Here['Цена'] = to_int_series(Sheet_Name_Here['Цена'])
    Sheet_Name_Here['Стоимость'] = to_int_series(Sheet_Name_Here['Стоимость'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def lim (file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(1, 'new-column-7enz', 0)
    Sheet_Name_Here.insert(1, 'new-column-3trw', 0)
    Sheet_Name_Here.insert(3, 'new-column-1uij', 0)
    # Added column new-column-p75x
    Sheet_Name_Here.insert(5, 'new-column-p75x', 0)

    # Renamed columns Пропуск0000
    Sheet_Name_Here.rename(columns={'Цена, руб.': 'Пропуск0000'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-p75x': 'Цена'}, inplace=True)

    Sheet_Name_Here.rename(columns={'new-column-7enz': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-1uij': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here['new-column-3trw'] = SUBSTITUTE(Sheet_Name_Here['Дата приёма'], RIGHT(Sheet_Name_Here['Дата приёма'], FIND(Sheet_Name_Here['Дата приёма'], ',')), '')
    Sheet_Name_Here['ФИО'] = IF(FIND(Sheet_Name_Here['new-column-3trw'], '.') > 0, None, SUBSTITUTE( Sheet_Name_Here['new-column-3trw'] , ',', '' ) )
    Sheet_Name_Here['Дата услуги'] = IF(FIND(Sheet_Name_Here['new-column-3trw'], '.') > 0, Sheet_Name_Here['new-column-3trw'], None)
    Sheet_Name_Here.insert(4, 'new-column-l0hc', 0)
    Sheet_Name_Here.insert(4, 'new-column-gczw', 0)
    Sheet_Name_Here.rename(columns={'new-column-gczw': 'Номер гп'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-l0hc': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here['Номер гп'] = IF(FIND(Sheet_Name_Here['Кол-во'], '/') > 0, Sheet_Name_Here['Кол-во'], None)
    Sheet_Name_Here['Номер полиса'] = IF( AND(FIND(Sheet_Name_Here['Кол-во'], '/') <= 0, INT(Sheet_Name_Here['Кол-во']) > 100), Sheet_Name_Here['Кол-во'], None)
    Sheet_Name_Here.insert(12, 'new-column-mfip', 0)
    Sheet_Name_Here.insert(12, 'new-column-nwqo', 0)
    Sheet_Name_Here.insert(12, 'new-column-gq00', 0)
    Sheet_Name_Here['new-column-gq00'] = IF(FIND(Sheet_Name_Here['Код МКБ'], 'На общую сумму:') > 0, FLOAT(SUBSTITUTE(SUBSTITUTE( SUBSTITUTE(Sheet_Name_Here['Кол-во'], ' руб.', ''), ',', '.'), ' ', '')), None)
    Sheet_Name_Here['new-column-nwqo'] = IF(FIND(Sheet_Name_Here['Код МКБ'], 'С учетом скидки:') > 0, FLOAT(SUBSTITUTE(SUBSTITUTE( SUBSTITUTE(Sheet_Name_Here['Кол-во'], ' руб.', ''), ',', '.'), ' ', '')), None)
    columns_to_fill_nan = ['new-column-gq00']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here['new-column-mfip'] = 100 - (Sheet_Name_Here['new-column-nwqo'] * 100 / Sheet_Name_Here['new-column-gq00'])
    columns_to_fill_nan = ['new-column-mfip']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='bfill')
    Sheet_Name_Here.rename(columns={'new-column-mfip': 'Скидка'}, inplace=True)
    Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Пропуск0000']) != 'NaN', Sheet_Name_Here['Пропуск0000'] - (
                FLOAT(Sheet_Name_Here['Пропуск0000']) / 100 * Sheet_Name_Here['Скидка']), None)
    columns_to_fill_nan = ['ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here['Номер гп'] = IF(FIND(Sheet_Name_Here['Кол-во'], '/') > 0, Sheet_Name_Here['Кол-во'], None)
    Sheet_Name_Here.insert(5, 'new-column-e7c2', 0)
    Sheet_Name_Here['new-column-e7c2'] = IF(Sheet_Name_Here['Код МКБ'] == 'Итого по пациенту:', 'STOP', None)
    Sheet_Name_Here.insert(6, 'new-column-mdq9', 0)
    Sheet_Name_Here['new-column-mdq9'] = CONCAT(Sheet_Name_Here['Номер гп'], Sheet_Name_Here['new-column-e7c2'])
    Sheet_Name_Here['new-column-e7c2'] = IF(Sheet_Name_Here['Код МКБ'] == 'Итого по пациенту:', 'STOP', ' ')
    Sheet_Name_Here['new-column-mdq9'] = CONCAT(Sheet_Name_Here['Номер гп'], Sheet_Name_Here['new-column-e7c2'])
    Sheet_Name_Here['Номер гп'] = IF(FIND(Sheet_Name_Here['Кол-во'], '/') > 0, Sheet_Name_Here['Кол-во'], ' ')
    Sheet_Name_Here['new-column-mdq9'] = CONCAT(Sheet_Name_Here['Номер гп'], Sheet_Name_Here['new-column-e7c2'])
    Sheet_Name_Here.insert(7, 'new-column-rucy', 0)
    Sheet_Name_Here['new-column-rucy'] = SUBSTITUTE(Sheet_Name_Here['new-column-mdq9'], ' ', '')
    Sheet_Name_Here['new-column-e7c2'] = IF(Sheet_Name_Here['Код МКБ'] == 'Итого по пациенту:', 'STOP', '')
    Sheet_Name_Here['new-column-mdq9'] = CONCAT(Sheet_Name_Here['Номер гп'], Sheet_Name_Here['new-column-e7c2'])
    Sheet_Name_Here['new-column-rucy'] = SUBSTITUTE(Sheet_Name_Here['new-column-mdq9'], ' ', '')
    Sheet_Name_Here['Номер гп'] = IF(FIND(Sheet_Name_Here['Кол-во'], '/') > 0, Sheet_Name_Here['Кол-во'], '')
    Sheet_Name_Here['new-column-mdq9'] = CONCAT(Sheet_Name_Here['Номер гп'], Sheet_Name_Here['new-column-e7c2'])
    Sheet_Name_Here['new-column-rucy'] = SUBSTITUTE(Sheet_Name_Here['new-column-mdq9'], ' ', '')
    Sheet_Name_Here['new-column-rucy'] = IF(Sheet_Name_Here['new-column-mdq9'] == '', None, Sheet_Name_Here['new-column-mdq9'])
    Sheet_Name_Here['Номер полиса'] = IF( AND(FIND(Sheet_Name_Here['Кол-во'], '/') <= 0, INT(Sheet_Name_Here['Кол-во']) > 100), Sheet_Name_Here['Кол-во'], '')
    Sheet_Name_Here.insert(9, 'new-column-cbdx', 0)
    Sheet_Name_Here.insert(9, 'new-column-prpp', 0)
    Sheet_Name_Here.insert(9, 'new-column-05dz', 0)
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'new-column-e7c2']
    Sheet_Name_Here_columns.insert(20, 'new-column-e7c2')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'new-column-mdq9']
    Sheet_Name_Here_columns.insert(20, 'new-column-mdq9')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here.rename(columns={'Номер гп': 'None_polyc'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-rucy': 'Номер гп'}, inplace=True)
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'Номер гп']
    Sheet_Name_Here_columns.insert(4, 'Номер гп')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'None_polyc']
    Sheet_Name_Here_columns.insert(20, 'None_polyc')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here['new-column-05dz'] = CONCAT(Sheet_Name_Here['Номер полиса'], Sheet_Name_Here['new-column-e7c2'])
    Sheet_Name_Here['new-column-prpp'] = IF(Sheet_Name_Here['new-column-05dz'] == '', None, Sheet_Name_Here['new-column-05dz'])
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'new-column-05dz']
    Sheet_Name_Here_columns.insert(20, 'new-column-05dz')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here.rename(columns={'Номер полиса': 'None_gp'}, inplace=True)
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'None_gp']
    Sheet_Name_Here_columns.insert(20, 'None_gp')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here.rename(columns={'new-column-prpp': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.drop(['new-column-cbdx'], axis=1, inplace=True)
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'new-column-gq00']
    Sheet_Name_Here_columns.insert(19, 'new-column-gq00')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'new-column-nwqo']
    Sheet_Name_Here_columns.insert(19, 'new-column-nwqo')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    columns_to_fill_nan = ['Номер гп', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here.insert(5, 'new-column-5sc5', 0)
    Sheet_Name_Here.insert(7, 'new-column-4lkx', 0)
    Sheet_Name_Here.rename(columns={'Номер гп': 'НЕ Номер гп'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-5sc5': 'Номер гп'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Номер полиса': ' НЕ Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-4lkx': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here['Номер гп'] = IF(Sheet_Name_Here['НЕ Номер гп'] == 'STOP', None, Sheet_Name_Here['НЕ Номер гп'])
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'НЕ Номер гп']
    Sheet_Name_Here_columns.insert(21, 'НЕ Номер гп')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here['Номер полиса'] = IF(Sheet_Name_Here[' НЕ Номер полиса'] == 'STOP', None, Sheet_Name_Here[' НЕ Номер полиса'])
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != ' НЕ Номер полиса']
    Sheet_Name_Here_columns.insert(21, ' НЕ Номер полиса')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]
    Sheet_Name_Here.rename(columns={'Номер гп': 'ГП'}, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Сумма, руб.'].notnull()]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    Sheet_Name_Here.rename(columns={'Дата приёма': 'Gjh'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Сумма, руб.': 'Gjhdd'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medecinskie_centri_2(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(1, 'new-column-03sx', 0)
    Sheet_Name_Here.rename(columns={'new-column-03sx': 'ФИО'}, inplace=True)
    Sheet_Name_Here.insert(2, 'new-column-t6mt', 0)
    Sheet_Name_Here.rename(columns={'new-column-t6mt': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Фамилия': 'Врач'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код': 'Номер Зуба'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-': 'Кол-во'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = IF(SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['№']), LEFT(CLEAN(Sheet_Name_Here['№']), INT(FIND(CLEAN(Sheet_Name_Here['№']), ' ') + 1)), ''), ' ', '') == '', None, SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['№']), LEFT(CLEAN(Sheet_Name_Here['№']), INT(FIND( CLEAN(Sheet_Name_Here['№']), ' ') + 1)), ''), ' ', ''))
    Sheet_Name_Here['ФИО'] = IF ( SUBSTITUTE ( Sheet_Name_Here['№'], LEFT ( Sheet_Name_Here['№'], INT(FIND( CLEAN( Sheet_Name_Here['№']), ' ') + 1) ) , '') == '', None,  PROPER ( SUBSTITUTE( SUBSTITUTE ( Sheet_Name_Here['№'], LEFT ( Sheet_Name_Here['№'], INT(FIND( CLEAN( Sheet_Name_Here['№']), ' ') + 1) ) , ''), SUBSTITUTE (SUBSTITUTE ( CLEAN( Sheet_Name_Here['№']), LEFT ( CLEAN( Sheet_Name_Here['№']), INT(FIND( CLEAN( Sheet_Name_Here['№']), ' ') + 1) ) , ''), ' ', ''), '', ) ) )
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата'].notnull()]
    Sheet_Name_Here.drop(['д'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Уда-'], axis=1, inplace=True)
    Sheet_Name_Here['Cумма к оплате'] = Sheet_Name_Here['Cумма к оплате'].fillna(0).astype('int')
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    Sheet_Name_Here.insert(3, 'new-column-aswh', 0)
    Sheet_Name_Here['new-column-aswh'] = INT(Sheet_Name_Here['Номер полиса'])
    Sheet_Name_Here.rename(columns={'Номер полиса': 'temp_value'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-aswh': 'Номер полиса'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def klinicheskaya_bolnica_1(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=11)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_obp(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)
    Sheet_Name_Here.rename(columns={'Описание': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = Sheet_Name_Here['Цена'].fillna(0).astype('int')
    Sheet_Name_Here.rename(columns={'Дата исп.': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.insert(1, 'new-column-g6wi', 0)
    Sheet_Name_Here.insert(1, 'new-column-h89p', 0)
    Sheet_Name_Here.rename(columns={'new-column-h89p': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-g6wi': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(FIND(Sheet_Name_Here['Дата услуги'], 'Полис') > 0,
                        LEFT(Sheet_Name_Here['Дата услуги'], INT(FIND(Sheet_Name_Here['Дата услуги'], ',') - 1)), None)
    Sheet_Name_Here['Номер полиса'] = IF(FIND(Sheet_Name_Here['Дата услуги'], 'Полис') > 0, SUBSTITUTE(LEFT(CLEAN(
        SUBSTITUTE(Sheet_Name_Here['Дата услуги'], LEFT(Sheet_Name_Here['Дата услуги'], INT(FIND(Sheet_Name_Here['Дата услуги'], ',') + 1)),
                   '')), INT(FIND(CLEAN(
        SUBSTITUTE(Sheet_Name_Here['Дата услуги'], LEFT(Sheet_Name_Here['Дата услуги'], INT(FIND(Sheet_Name_Here['Дата услуги'], ',') + 1)),
                   '')), ':') - 1)), ' ', ''), None)
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here.drop(['Вид цены'], axis=1, inplace=True)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ldc_kazanskaya_klinka(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)
    Sheet_Name_Here.insert(2, 'new-column-oerd', 0)
    Sheet_Name_Here.rename(columns={'new-column-oerd': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'№ полиса': 'Пропуск'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск'], ',', ''), None)
    Sheet_Name_Here.rename(columns={'Наименование мед. услуги': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here['Стоимость'] = Sheet_Name_Here['Стоимость'].fillna(0).astype('int')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    columns_to_fill_nan = ['Фамилия Имя Отчетсво', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Фамилия Имя Отчетсво': 'ФИО'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gkb_nomer_1_n_i_pirogova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[2]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5, converters={'Код услуги по Прейскуранту': str})
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(2, 'new-column-kowb', 0)
    Sheet_Name_Here.rename(columns={'new-column-kowb': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Данные полиса': 'Пропуск'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск'], LEFT(Sheet_Name_Here['Пропуск'],
                                                                                                     INT(FIND(Sheet_Name_Here[
                                                                                                                  'Пропуск'],
                                                                                                              ':'))),
                                                                             ''), None)
    Sheet_Name_Here.insert(7, 'new-column-m2i5', 0)
    Sheet_Name_Here.rename(columns={'ФИО': 'Пропуск2'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-m2i5': 'ФИО'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск2'], SUBSTITUTE(Sheet_Name_Here['Пропуск2'],
                                                                                                    LEFT(Sheet_Name_Here[
                                                                                                             'Пропуск2'],
                                                                                                         INT(FIND(
                                                                                                             Sheet_Name_Here[
                                                                                                                 'Пропуск2'],
                                                                                                             ',') - 1)),
                                                                                                    ''), ''), None)
    Sheet_Name_Here.insert(10, 'new-column-giyl', 0)
    Sheet_Name_Here.insert(10, 'new-column-259q', 0)
    Sheet_Name_Here.rename(columns={'Дата оказания услуг': 'Пропуск3'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-259q': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-giyl': 'Дата окончания услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Диагноза по МКБ-Х': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код услуги по Прейскуранту': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Наименование медицинской услуги по Прейскуранту': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Количество услуг': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Стоимость'}, inplace=True)
    Sheet_Name_Here.drop(['Общая Сумма'], axis=1, inplace=True)
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Дата услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск3']) != 'NaN', LEFT(Sheet_Name_Here['Пропуск3'], INT(FIND(Sheet_Name_Here['Пропуск3'], '-') - 1)), None)
    Sheet_Name_Here['Дата окончания услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск3']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск3'], LEFT(Sheet_Name_Here['Пропуск3'], INT(FIND(Sheet_Name_Here['Пропуск3'], '-'))), ''), None)
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Стоимость'].notnull()]
    Sheet_Name_Here['Стоимость'] = Sheet_Name_Here['Стоимость'].fillna(0).astype('int')
    columns_to_fill_nan = ['Номер полиса', 'ФИО', 'Дата услуги', 'Дата окончания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def blesk_na_geodezicheskoi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=16)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'№\nзуба': 'ППропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 5': 'Номер зуба'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 2': 'Врач'}, inplace=True)
    Sheet_Name_Here.insert(3, 'new-column-kq5i', 0)
    Sheet_Name_Here.insert(3, 'new-column-ot1w', 0)
    Sheet_Name_Here.rename(columns={'new-column-ot1w': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-kq5i': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.insert(6, 'new-column-0mit', 0)
    Sheet_Name_Here.insert(6, 'new-column-ais5', 0)
    Sheet_Name_Here.rename(columns={'new-column-ais5': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Диагноз': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-0mit': 'Диагноз'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Unnamed: 4': 'Пропуск2'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Название работы': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'ФИО': 'Врач '}, inplace=True)
    Sheet_Name_Here.rename(columns={'Врач': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Врач ': 'Врач'}, inplace=True)
    Sheet_Name_Here.insert(2, 'new-column-60zy', 0)
    Sheet_Name_Here.rename(columns={'new-column-60zy': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['Unnamed: 27']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Unnamed: 27'], '№ СП: ', ''),
                               None)
    Sheet_Name_Here['МКБ'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], ' ') - 1)),
                      None)
    Sheet_Name_Here['Диагноз'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                          SUBSTITUTE(Sheet_Name_Here['Пропуск'], LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], ' '))), ''),
                          None)
    Sheet_Name_Here.rename(columns={'Код\nработы': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.insert(13, 'new-column-fg6h', 0)
    Sheet_Name_Here.rename(columns={'new-column-fg6h': 'Цена'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Сумма\nруб.': 'Пропуск4'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-\nво': 'Кол-во'}, inplace=True)
    Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', INT(Sheet_Name_Here['Пропуск4']) / INT(Sheet_Name_Here['Кол-во']), None)
    Sheet_Name_Here['Врач'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN',
                       SUBSTITUTE(Sheet_Name_Here['Пропуск2'], SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск2']), ' ', ''), ''), None)
    Sheet_Name_Here['Дата услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск2'], Sheet_Name_Here['Врач'], ''), None)
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Врач', 'Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', INT(Sheet_Name_Here['Пропуск4']) / INT(Sheet_Name_Here['Кол-во']), None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nash_mc_paracels(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(4, 'new-column-csyg', 0)
    Sheet_Name_Here.rename(columns={'new-column-csyg': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Полис застрахованного': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'ФИО застрахованного': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.insert(8, 'new-column-3hrq', 0)
    Sheet_Name_Here.rename(columns={'new-column-3hrq': 'Цена'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-во оказанных услуг': 'Кол-во'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['Код услуги']) != 'NaN', INT(FIND(Sheet_Name_Here['Код услуги'], '.') <= 0)),
                        Sheet_Name_Here['Код услуги'], None)
    columns_to_fill_nan = ['ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Стоимость одной услуги, руб.'].notnull()]
    Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Стоимость одной услуги, руб.']) != 'NaN',
                         INT(INT(Sheet_Name_Here['Стоимость одной услуги, руб.']) / INT(Sheet_Name_Here['Кол-во'])), None)
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Стоимость одной услуги, руб.']) != 'NaN',
                         INT(INT(Sheet_Name_Here['Стоимость одной услуги, руб.']) / INT(Sheet_Name_Here['Кол-во'])), None)
    Sheet_Name_Here['Номер полиса'] = Sheet_Name_Here['Номер полиса'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def cs_prior_m(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4, converters={'Код услуги (клиника)':str, '№ зуба':str})
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Ф.И.О. пациента': '4'}, inplace=True)
    Sheet_Name_Here.insert(9, 'new-column-6vah', 0)
    Sheet_Name_Here.rename(columns={'Диагноз': 'Пропуск2'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-6vah': 'Диaгноз'}, inplace=True)
    Sheet_Name_Here['Диaгноз'] = IF(TYPE(Sheet_Name_Here['Пропуск2']) != 'NaN',
                                      SUBSTITUTE(Sheet_Name_Here['Пропуск2'], LEFT(Sheet_Name_Here['Пропуск2'],
                                                                                      INT(FIND(Sheet_Name_Here[
                                                                                                   'Пропуск2'], ','))),
                                                 ''), None)
    Sheet_Name_Here.insert(8, 'new-column-q4ui', 0)
    Sheet_Name_Here.rename(columns={'new-column-q4ui': 'Номер зуба'}, inplace=True)
    Sheet_Name_Here['Номер зуба'] = IF(TYPE(Sheet_Name_Here['№ зуба']) != 'NaN',
                                          SUBSTITUTE(CLEAN(Sheet_Name_Here['№ зуба']), '-', ''), None)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Пропуск3'}, inplace=True)
    Sheet_Name_Here.insert(1, 'new-column-4izj', 0)
    Sheet_Name_Here.rename(columns={'new-column-4izj': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'4': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Адрес регистрации'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Место работы'], axis=1, inplace=True)
    Sheet_Name_Here.insert(2, 'new-column-e627', 0)
    Sheet_Name_Here.rename(columns={'new-column-e627': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.insert(3, 'new-column-csba', 0)
    Sheet_Name_Here.rename(columns={'new-column-csba': 'Пропуск11'}, inplace=True)
    Sheet_Name_Here.insert(4, 'new-column-z95n', 0)
    Sheet_Name_Here.rename(columns={'new-column-z95n': 'Пропуск12'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Страховой полис': 'Пропуск22'}, inplace=True)
    Sheet_Name_Here['Пропуск12'] = IF(TYPE(Sheet_Name_Here['Пропуск22']) != 'NaN', SUBSTITUTE(
        SUBSTITUTE(Sheet_Name_Here['Пропуск22'],
                   LEFT(Sheet_Name_Here['Пропуск22'], INT(FIND(Sheet_Name_Here['Пропуск22'], '-'))), ''), '\n',
        ' '), None)
    Sheet_Name_Here['Пропуск11'] = IF(
        AND(FIND(Sheet_Name_Here['Пропуск12'], ' ') > 0, TYPE(Sheet_Name_Here['Пропуск12']) != 'NaN'), LEFT(
            SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['Пропуск22'], LEFT(Sheet_Name_Here['Пропуск22'],
                                                                        INT(FIND(Sheet_Name_Here['Пропуск22'],
                                                                                 '-'))), ''), '\n', ' '), INT(FIND(
                SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['Пропуск22'], LEFT(Sheet_Name_Here['Пропуск22'],
                                                                            INT(FIND(Sheet_Name_Here['Пропуск22'],
                                                                                     '-'))), ''), '\n', ' '), ' '))),
        Sheet_Name_Here['Пропуск12'])
    Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['Пропуск11']) != 'NaN',
                                            SUBSTITUTE(Sheet_Name_Here['Пропуск11'],
                                                       LEFT(Sheet_Name_Here['Пропуск11'],
                                                            INT(FIND(Sheet_Name_Here['Пропуск11'], '_0'))), ''),
                                            None)
    Sheet_Name_Here['Номер полиса'] = [x if x != 'nan' else None for x in Sheet_Name_Here['Номер полиса']  ]
    Sheet_Name_Here.rename(columns={'Дата лечения': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', PROPER(
        SUBSTITUTE(Sheet_Name_Here['Пропуск'],
                   LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], ','))), '')), None)
    Sheet_Name_Here.rename(columns={'Манипуляции, материалы': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код услуги (клиника)': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.drop(['Unnamed: 11'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['К оплате страховой компанией'], axis=1, inplace=True)
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата услуги', 'Врач', 'Номер зуба', 'Диaгноз']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def rzd_spb(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(2, 'new-column-aadz', 0)
    Sheet_Name_Here.insert(2, 'new-column-958k', 0)
    Sheet_Name_Here.rename(columns={'Фамилия, И.О. застрахованного': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-958k': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-aadz': 'Номер мед.карты'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код мед. услуги по прейскуранту ЛПУ': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Наименование мед. услуги': 'Наименование услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код МКБ': 'МКБ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Кол-во мед. услуг': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Стоимость медицинской услуги': 'Цена'}, inplace=True)
    Sheet_Name_Here.drop(['Итого'], axis=1, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                LEFT(Sheet_Name_Here['Пропуск'], INT(FIND(Sheet_Name_Here['Пропуск'], '.') - 4)), None)
    Sheet_Name_Here['Номер мед.карты'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                            SUBSTITUTE(Sheet_Name_Here['Пропуск'], LEFT(Sheet_Name_Here['Пропуск'],
                                                                                        INT(FIND(
                                                                                            Sheet_Name_Here['Пропуск'],
                                                                                            '№') + 1)), ''), None)
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull()) & (
        Sheet_Name_Here['Кол-во'].apply(lambda val: all(s not in str(val) for s in ['ИТОГО', 'бухгалтер']))) \
                                      & (Sheet_Name_Here['Кол-во'].apply(lambda val: all(
        val != s for s in ['Поликлиника      ', 'Итого:', 'Итого по подразделению  Поликлиника      :'])))]
    columns_to_fill_nan = ['ФИО', 'Номер мед.карты']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])
    Sheet_Name_Here['№ полиса'] = to_int_series(Sheet_Name_Here['№ полиса'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def dinastia(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.insert(1, 'new-column-gq9a', 0)
        Sheet_Name_Here.rename(columns={'new-column-gq9a': 'Филиал'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Полис. Номер': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Пациент': 'ФИО'}, inplace=True)
        Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)
        Sheet_Name_Here.rename(columns={'Дата оказания': 'Дата услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Номенклатура. Код': 'Код услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Наименование медицинских услуг': 'Наименование услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код МКБ': 'МКБ'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Итого': 'Кол-во'}, inplace=True)
        Sheet_Name_Here.insert(9, 'new-column-1ljk', 0)
        Sheet_Name_Here.rename(columns={'new-column-1ljk': 'Цена'}, inplace=True)
        Sheet_Name_Here['Филиал'] = IF(AND(TYPE(Sheet_Name_Here['№ п/п']) != 'NaN', CLEAN(Sheet_Name_Here['№ п/п']) == ''), Sheet_Name_Here['№ п/п'], None)
        Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['Unnamed: 9']) != 'NaN'),
                             FLOAT(Sheet_Name_Here['Unnamed: 9']) / INT(Sheet_Name_Here['Кол-во']), None)
        columns_to_fill_nan = ['ФИО', 'Филиал', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here.insert(4, 'new-column-4s5r', 0)
        Sheet_Name_Here.rename(columns={'Номер полиса': 'Пропуск'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-4s5r': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(Sheet_Name_Here['Пропуск'], ' ', ''), None)
        Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Дата услуги'].str.contains('Итого', na=False)]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
        Sheet_Name_Here.insert(5, 'new-column-qzze', 0)
        Sheet_Name_Here.rename(columns={'Номер полиса': 'Пропуск2'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-qzze': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here['Номер полиса'] = IF(FIND(Sheet_Name_Here['Пропуск2'], '-') > 0, Sheet_Name_Here['Пропуск2'], INT(Sheet_Name_Here['Пропуск2']))
        Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['Unnamed: 9']) != 'NaN'),
                             FLOAT(FLOAT(Sheet_Name_Here['Unnamed: 9']) / INT(Sheet_Name_Here['Кол-во'])), None)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        Sheet_Name_Here.insert(1, 'new-column-gq9a', 0)
        Sheet_Name_Here.rename(columns={'new-column-gq9a': 'Филиал'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Полис. Номер': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Пациент': 'ФИО'}, inplace=True)
        Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)
        Sheet_Name_Here.rename(columns={'Дата оказания': 'Дата услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Номенклатура. Код': 'Код услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Наименование медицинских услуг': 'Наименование услуги'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Код МКБ': 'МКБ'}, inplace=True)
        Sheet_Name_Here.rename(columns={'Итого': 'Кол-во'}, inplace=True)
        Sheet_Name_Here.insert(9, 'new-column-1ljk', 0)
        Sheet_Name_Here.rename(columns={'new-column-1ljk': 'Цена'}, inplace=True)
        Sheet_Name_Here['Филиал'] = IF(
            AND(TYPE(Sheet_Name_Here['№ п/п']) != 'NaN', CLEAN(Sheet_Name_Here['№ п/п']) == ''),
            Sheet_Name_Here['№ п/п'], None)
        Sheet_Name_Here['Цена'] = IF(
            AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['Unnamed: 9']) != 'NaN'),
            FLOAT(Sheet_Name_Here['Unnamed: 9']) / INT(Sheet_Name_Here['Кол-во']), None)
        columns_to_fill_nan = ['ФИО', 'Филиал', 'Номер полиса']
        Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
        Sheet_Name_Here.insert(4, 'new-column-4s5r', 0)
        Sheet_Name_Here.rename(columns={'Номер полиса': 'Пропуск'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-4s5r': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                             SUBSTITUTE(Sheet_Name_Here['Пропуск'], ' ', ''), None)
        Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Дата услуги'].str.contains('Итого', na=False)]
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
        Sheet_Name_Here.insert(5, 'new-column-qzze', 0)
        Sheet_Name_Here.rename(columns={'Номер полиса': 'Пропуск2'}, inplace=True)
        Sheet_Name_Here.rename(columns={'new-column-qzze': 'Номер полиса'}, inplace=True)
        Sheet_Name_Here['Номер полиса'] = IF(FIND(Sheet_Name_Here['Пропуск2'], '-') > 0, Sheet_Name_Here['Пропуск2'],
                                             INT(Sheet_Name_Here['Пропуск2']))
        Sheet_Name_Here['Цена'] = IF(
            AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here['Unnamed: 9']) != 'NaN'),
            FLOAT(FLOAT(Sheet_Name_Here['Unnamed: 9']) / INT(Sheet_Name_Here['Кол-во'])), None)
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def laus_deo(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Серия и номер полиса': 'Номер полиса'}, inplace=True)
    Sheet_Name_Here.insert(1, 'new-column-qmxr', 0)
    Sheet_Name_Here.rename(columns={'new-column-qmxr': 'ФИО'}, inplace=True)
    Sheet_Name_Here.rename(columns={'ФИО  пациента': 'ПропускФ'}, inplace=True)
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['ПропускФ']) != 'NaN',
                                SUBSTITUTE(Sheet_Name_Here['ПропускФ'], '\n', ' '), None)
    Sheet_Name_Here.rename(columns={'Диагноз по МКБ код': 'Диагноз'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Цена услуги': 'Цена'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Количество услуг': 'Кол-во'}, inplace=True)
    Sheet_Name_Here.rename(columns={'ФИО доктора полностью': 'Врач'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Итоговая цена': 'Пропуск'}, inplace=True)
    Sheet_Name_Here.insert(11, 'new-column-s14t', 0)
    Sheet_Name_Here.insert(11, 'new-column-85kh', 0)
    Sheet_Name_Here.rename(columns={'№ ГП': 'ПропускГ'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-85kh': 'ГП'}, inplace=True)
    Sheet_Name_Here.rename(columns={'new-column-s14t': 'Дата ГП'}, inplace=True)
    Sheet_Name_Here['ГП'] = IF(TYPE(Sheet_Name_Here['ПропускГ']) != 'NaN',
                               LEFT(Sheet_Name_Here['ПропускГ'], INT(FIND(Sheet_Name_Here['ПропускГ'], ' '))), None)
    Sheet_Name_Here['Дата ГП'] = IF(TYPE(Sheet_Name_Here['ПропускГ']) != 'NaN', SUBSTITUTE(
        CLEAN(SUBSTITUTE(Sheet_Name_Here['ПропускГ'], Sheet_Name_Here['ГП'], '')), ' ', ''), None)
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def centr_luchevoi_diagnostiki_tonus_premium(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Ф.И.О. Застрахованного': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис': 'Номер полиса'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код \nмед-й \nуслуги': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код \nмед-ой \nуслуги': 'Код услуги'}, inplace=True)
    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Стоимость \nодной \nуслуги': 'Цена'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-во\n оказ-х \nуслуг': 'Кол-во'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def tonus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Ф.И.О. Застрахованного': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис': 'Номер полиса'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код \nмед-й \nуслуги': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код \nмед-ой \nуслуги': 'Код услуги'}, inplace=True)
    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Стоимость \nодной \nуслуги': 'Цена'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-во\n оказ-х \nуслуг': 'Кол-во'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def tonus_plus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Ф.И.О. Застрахованного': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис': 'Номер полиса'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код \nмед-й \nуслуги': 'Код услуги'}, inplace=True)
    Sheet_Name_Here.rename(columns={'Код \nмед-ой \nуслуги': 'Код услуги'}, inplace=True)
    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Стоимость \nодной \nуслуги': 'Цена'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-во\n оказ-х \nуслуг': 'Кол-во'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Номер полиса', 'Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def lotos(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО'}, inplace=True)

    # Added column new-column-1lu8
    Sheet_Name_Here.insert(5, 'new-column-1lu8', 0)

    # Renamed columns ПропускГ
    Sheet_Name_Here.rename(columns={'№ и дата направления': 'Пропуск1'}, inplace=True)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-1lu8': 'ГП'}, inplace=True)

    # Renamed columns Дата окончания услуги
    Sheet_Name_Here.rename(columns={'Дата реализации': 'Дата окончания услуги'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата заказа': 'Дата услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код по прайсу': 'Код услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Все МКБ10': 'МКБ'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Количество': 'Кол-во'}, inplace=True)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = IF(TYPE(Sheet_Name_Here['Пропуск1']) != 'NaN',
                               LEFT(Sheet_Name_Here['Пропуск1'], INT(FIND(Sheet_Name_Here['Пропуск1'], ' ') - 1)), None)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Filtered Цена
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def odinmed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-x4yf
    Sheet_Name_Here.insert(1, 'new-column-x4yf', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-x4yf': 'ФИО'}, inplace=True)

    # Added column new-column-3yem
    Sheet_Name_Here.insert(2, 'new-column-3yem', 0)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'new-column-3yem': 'Номер полиса'}, inplace=True)

    # Added column new-column-tw4g
    Sheet_Name_Here.insert(1, 'new-column-tw4g', 0)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'new-column-tw4g': 'Дата услуги'}, inplace=True)

    # Renamed columns Проск
    Sheet_Name_Here.rename(columns={'Дата оказания услуг': 'Проск'}, inplace=True)

    # Added column new-column-iu8m
    Sheet_Name_Here.insert(5, 'new-column-iu8m', 0)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'new-column-iu8m': 'Врач'}, inplace=True)

    # Added column new-column-guw0
    Sheet_Name_Here.insert(3, 'new-column-guw0', 0)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-guw0': 'ГП'}, inplace=True)

    # Renamed columns Ппр
    Sheet_Name_Here.rename(columns={'ФИО доктора': 'Ппр'}, inplace=True)

    # Set formula of Врач
    Sheet_Name_Here['Врач'] = IF(AND(TYPE(Sheet_Name_Here['Ппр']) != 'NaN', FIND(Sheet_Name_Here['Ппр'], '.')),
                                 Sheet_Name_Here['Ппр'], None)

    # Set formula of Дата услуги
    Sheet_Name_Here['Дата услуги'] = IF(
        AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN',
            OR(FIND(Sheet_Name_Here['Проск'], '.'), FIND(Sheet_Name_Here['Проск'], '-'))),
        Sheet_Name_Here['Проск'], None)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = IF(AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], ':'),
                                   FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту:') <= 0,
                                   FIND(Sheet_Name_Here['Проск'], 'ГП')), SUBSTITUTE(
        SUBSTITUTE(CLEAN(SUBSTITUTE(Sheet_Name_Here['Проск'],
                                    LEFT(Sheet_Name_Here['Проск'], INT(FIND(Sheet_Name_Here['Проск'], '('))), '')), ' ',
                   ''), ')', ''), None)

    # Added column new-column-euq8
    Sheet_Name_Here.insert(5, 'new-column-euq8', 0)

    # Renamed columns ПереходныйПолис
    Sheet_Name_Here.rename(columns={'new-column-euq8': 'ПереходныйПолис'}, inplace=True)

    # Added column new-column-jpgx
    Sheet_Name_Here.insert(6, 'new-column-jpgx', 0)

    # Renamed columns ПереходныйПОЛИС2
    Sheet_Name_Here.rename(columns={'new-column-jpgx': 'ПереходныйПОЛИС2'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], ':'),
                                    FIND(Sheet_Name_Here['Проск'], '-') <= 0,
                                    FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту:') <= 0),
                                SUBSTITUTE(Sheet_Name_Here['Проск'],
                                           SUBSTITUTE(Sheet_Name_Here['Проск'],
                                                      LEFT(
                                                          Sheet_Name_Here['Проск'],
                                                          INT(FIND(Sheet_Name_Here[
                                                                       'Проск'],
                                                                   '№') - 1)),
                                                      ''), ''), None)

    # Set formula of ПереходныйПолис
    Sheet_Name_Here['ПереходныйПолис'] = IF(
        AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], ':'),
            FIND(Sheet_Name_Here['Проск'], '-') <= 0,
            FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту:') <= 0),
        SUBSTITUTE(Sheet_Name_Here['Проск'], SUBSTITUTE(Sheet_Name_Here['Проск'],
                                                        SUBSTITUTE(Sheet_Name_Here[
                                                                       'Проск'],
                                                                   LEFT(
                                                                       Sheet_Name_Here[
                                                                           'Проск'],
                                                                       INT(FIND(
                                                                           Sheet_Name_Here[
                                                                               'Проск'],
                                                                           '№') - 1)),
                                                                   ''),
                                                        ''), ''), None)

    # Set formula of ПереходныйПОЛИС2
    Sheet_Name_Here['ПереходныйПОЛИС2'] = IF(
        AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], ':'),
            FIND(Sheet_Name_Here['Проск'], '-') <= 0,
            FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту:') <= 0, FIND(Sheet_Name_Here['Проск'], 'ГП')),
        SUBSTITUTE(Sheet_Name_Here['Проск'],
                   SUBSTITUTE(
                       Sheet_Name_Here[
                           'Проск'],
                       SUBSTITUTE(
                           Sheet_Name_Here[
                               'Проск'],
                           LEFT(
                               Sheet_Name_Here[
                                   'Проск'],
                               INT(FIND(
                                   Sheet_Name_Here[
                                       'Проск'],
                                   '№') - 1)),
                           ''), ''),
                   ''), None)

    # Set formula of ПереходныйПолис
    Sheet_Name_Here['ПереходныйПолис'] = IF(
        AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], ':'),
            FIND(Sheet_Name_Here['Проск'], '-') <= 0,
            FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту:') <= 0, FIND(Sheet_Name_Here['Проск'], 'ГП') <= 0),
        SUBSTITUTE(Sheet_Name_Here['Проск'], SUBSTITUTE(Sheet_Name_Here['Проск'], SUBSTITUTE(Sheet_Name_Here['Проск'],
                                                                                             LEFT(Sheet_Name_Here[
                                                                                                      'Проск'],
                                                                                                  INT(FIND(
                                                                                                      Sheet_Name_Here[
                                                                                                          'Проск'],
                                                                                                      '№') - 1)), ''),
                                                        ''), ''), None)

    # Added column new-column-4y17
    Sheet_Name_Here.insert(5, 'new-column-4y17', 0)

    # Added column new-column-9nva
    Sheet_Name_Here.insert(5, 'new-column-9nva', 0)

    # Renamed columns ПОЛИСАБЕЗГП
    Sheet_Name_Here.rename(columns={'new-column-9nva': 'ПОЛИСАБЕЗГП'}, inplace=True)

    # Renamed columns ПОЛИСАСГП
    Sheet_Name_Here.rename(columns={'new-column-4y17': 'ПОЛИСАСГП'}, inplace=True)

    # Set formula of ПОЛИСАБЕЗГП
    Sheet_Name_Here['ПОЛИСАБЕЗГП'] = IF(
        AND(TYPE(Sheet_Name_Here['ПереходныйПолис']) != 'NaN', TYPE(Sheet_Name_Here['ПереходныйПолис']) != 'object'),
        SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['ПереходныйПолис']), ':', ''), ' ', ''), None)

    # Set formula of ПОЛИСАСГП
    Sheet_Name_Here['ПОЛИСАСГП'] = IF(
        AND(TYPE(Sheet_Name_Here['ПереходныйПОЛИС2']) != 'NaN', TYPE(Sheet_Name_Here['ПереходныйПОЛИС2']) != 'object'),
        FLOAT(SUBSTITUTE(
            SUBSTITUTE(
                SUBSTITUTE(
                    SUBSTITUTE(CLEAN(SUBSTITUTE(Sheet_Name_Here['ПереходныйПОЛИС2'], Sheet_Name_Here['ГП'], '')), ':',
                               ''), '(',
                    ''), ')', ''), ' ', '')), None)

    # Set formula of ПереходныйПолис
    Sheet_Name_Here['ПереходныйПолис'] = IF(
        AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], ':'),
            FIND(Sheet_Name_Here['Проск'], '-') <= 0,
            FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту') <= 0, FIND(Sheet_Name_Here['Проск'], 'ГП') <= 0),
        SUBSTITUTE(Sheet_Name_Here['Проск'], SUBSTITUTE(Sheet_Name_Here['Проск'], SUBSTITUTE(Sheet_Name_Here['Проск'],
                                                                                             LEFT(Sheet_Name_Here[
                                                                                                      'Проск'],
                                                                                                  INT(FIND(
                                                                                                      Sheet_Name_Here[
                                                                                                          'Проск'],
                                                                                                      '№') - 1)), ''),
                                                        ''), ''), None)
    Sheet_Name_Here['ПОЛИСАБЕЗГП'] = IF(
        AND(TYPE(Sheet_Name_Here['ПереходныйПолис']) != 'NaN', TYPE(Sheet_Name_Here['ПереходныйПолис']) != 'object'),
        SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['ПереходныйПолис']), ':', ''), ' ', ''), None)

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = FILLNAN(Sheet_Name_Here['ПОЛИСАБЕЗГП'], Sheet_Name_Here['ПОЛИСАСГП'])

    # Added column new-column-8c81
    Sheet_Name_Here.insert(4, 'new-column-8c81', 0)

    # Renamed columns ПереходноеГП
    Sheet_Name_Here.rename(columns={'new-column-8c81': 'ПереходноеГП'}, inplace=True)

    # Added column new-column-f5ea
    Sheet_Name_Here.insert(5, 'new-column-f5ea', 0)

    # Renamed columns Переход2
    Sheet_Name_Here.rename(columns={'ГП': 'Переход2'}, inplace=True)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-f5ea': 'ГП'}, inplace=True)

    # Set formula of ПереходноеГП
    Sheet_Name_Here['ПереходноеГП'] = IF(
        AND(TYPE(Sheet_Name_Here['Проск']) != 'NaN', FIND(Sheet_Name_Here['Проск'], 'Итого по пациенту:')), '',
        None)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = FILLNAN(Sheet_Name_Here['Переход2'], Sheet_Name_Here['ПереходноеГП'])

    # Filled NaN values in 5 columns in Sheet1
    columns_to_fill_nan = ['Дата услуги', 'ФИО', 'ГП', 'Номер полиса', 'Врач']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во услуг
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во услуг'].notnull()]

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    # Changed Кол-во услуг to dtype int
    Sheet_Name_Here['Кол-во услуг'] = Sheet_Name_Here['Кол-во услуг'].fillna(0).astype('int')

    # Renamed columns sudhadh
    Sheet_Name_Here.rename(columns={'Сумма по услугам': 'sudhadh'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ': 'МКБ'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена услуги': 'Цена'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_bo_mgu_im_i_m_sechenova_mizdrava_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.insert(1, 'new-column-rfgp', 0)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-rfgp': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(
        AND(TYPE(Sheet_Name_Here['Дата услуги']) != 'NaN', FIND(Sheet_Name_Here['Дата услуги'], 'д/р') > 0),
        PROPER(LEFT(Sheet_Name_Here['Дата услуги'], INT(FIND(Sheet_Name_Here['Дата услуги'], 'д/р') - 1))), None)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'п/п': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Вид услуг': 'Наименование услуги'}, inplace=True)

    # Added column new-column-ufh9
    Sheet_Name_Here.insert(15, 'new-column-ufh9', 0)

    # Renamed columns ГП
    Sheet_Name_Here.rename(columns={'new-column-ufh9': 'ГП'}, inplace=True)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = IF(TYPE(Sheet_Name_Here['Unnamed: 13']) != 'NaN', SUBSTITUTE(
        SUBSTITUTE(CLEAN(LEFT(Sheet_Name_Here['Unnamed: 13'], INT(FIND(Sheet_Name_Here['Unnamed: 13'], 'от')))), ' ',
                   ''), '№', ''),
                               None)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'МКБ10': 'МКБ'}, inplace=True)

    # Added column new-column-2dwe
    Sheet_Name_Here.insert(27, 'new-column-2dwe', 0)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'new-column-2dwe': 'Номер полиса'}, inplace=True)

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = IF(
        AND(TYPE(Sheet_Name_Here['Цена']) != 'NaN', FIND(Sheet_Name_Here['Цена'], 'полис') > 0),
        FLOAT(SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Цена']), ' ', ''), '№', '')), None)

    # Added column new-column-v36r
    Sheet_Name_Here.insert(27, 'new-column-v36r', 0)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Цена': 'Пропуск'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-v36r': 'Цена'}, inplace=True)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', FIND(Sheet_Name_Here['Пропуск'], 'полис') <= 0),
        FLOAT(Sheet_Name_Here['Пропуск']), None)

    # Filled NaN values in 3 columns in TDSheet
    columns_to_fill_nan = ['ГП', 'Номер полиса', 'ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'К-во': 'Кол-во'}, inplace=True)

    # Renamed columns Пропуск2
    Sheet_Name_Here.rename(columns={'Сумма': 'Пропуск2'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Added column new-column-skza
    Sheet_Name_Here.insert(5, 'new-column-skza', 0)

    # Renamed columns Пропуск22
    Sheet_Name_Here.rename(columns={'Код услуги': 'Пропуск22'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'new-column-skza': 'Код услуги'}, inplace=True)

    # Set formula of Код услуги
    Sheet_Name_Here['Код услуги'] = IF(TYPE(Sheet_Name_Here['Пропуск22']) != 'NaN',
                                       SUBSTITUTE(Sheet_Name_Here['Пропуск22'], "'", ''), None)

    # Renamed columns Пропуск999
    Sheet_Name_Here.rename(columns={'Коэф': 'Пропуск999'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def beka_invest(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here[0].str.contains('ООО "Бестдоктор"', na=False)]

    # Added column
    Sheet_Name_Here.insert(7, 'new-column-5txp', 0)
    Sheet_Name_Here.insert(1, 'new-column-yjlw', 0)
    Sheet_Name_Here.insert(1, 'new-column-uaqj', 0)
    Sheet_Name_Here.insert(3, 'new-column-qdyr', 0)
    Sheet_Name_Here.insert(4, 'new-column-x9gq', 0)
    Sheet_Name_Here.insert(4, 'new-column-lpos', 0)

    # Renamed columns
    Sheet_Name_Here.rename(columns={1: 'Дата услуги',
                                    2: 'Код услуги',
                                    3: 'Наименование услуги',
                                    4: 'МКБ',
                                    5: 'Кол-во',
                                    'new-column-5txp': 'Цена',
                                    'new-column-uaqj': 'Номер полиса',
                                    'new-column-yjlw': 'ФИО',
                                    'new-column-qdyr': 'ГП',
                                    0: 'Ghjcr'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Наименование услуги']) == 'NaN',
                                SUBSTITUTE(Sheet_Name_Here['Ghjcr'], RIGHT(Sheet_Name_Here['Ghjcr'], 12), ''), None)

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = IF(
        AND(TYPE(Sheet_Name_Here['Наименование услуги']) != 'NaN', FIND(Sheet_Name_Here['Ghjcr'], '/') <= 0),
        Sheet_Name_Here['Ghjcr'], None)

    # Set formula of new-column-ured
    Sheet_Name_Here['new-column-x9gq'] = IF(
        AND(TYPE(Sheet_Name_Here['Наименование услуги']) != 'NaN', FIND(Sheet_Name_Here['Ghjcr'], '/') > 0),
        Sheet_Name_Here['Ghjcr'], None)

    # Set formula of new-column-2fnf
    Sheet_Name_Here['new-column-lpos'] = IF(
        AND(TYPE(Sheet_Name_Here['Номер полиса']) == 'object', TYPE(Sheet_Name_Here['new-column-x9gq']) == 'object'),
        '', None)

    # Set formula of ГП
    Sheet_Name_Here['ГП'] = FILLNAN(Sheet_Name_Here['new-column-x9gq'], Sheet_Name_Here['new-column-lpos'])

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'ГП', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    # Set formula of Кол-во
    Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here['Кол-во']) != 'NaN', TYPE(Sheet_Name_Here[6]) != 'NaN'),
                                 Sheet_Name_Here[6] / Sheet_Name_Here['Кол-во'], None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def klinika_sanitas_v_medparke(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Added column new-column-cs57
    Sheet_Name_Here.insert(1, 'new-column-cs57', 0)

    # Renamed columns ПропускФ
    Sheet_Name_Here.rename(columns={'Ф.И.О. застрахованного': 'ПропускФ'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-cs57': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['ПропускФ']) != 'NaN', Sheet_Name_Here['ПропускФ'] != ' '),
                                PROPER(Sheet_Name_Here['ПропускФ']), None)

    # Deleted columns рождения
    Sheet_Name_Here.drop(['рождения'], axis=1, inplace=True)

    # Added column new-column-knnz
    Sheet_Name_Here.insert(3, 'new-column-knnz', 0)

    # Renamed columns ПропускП
    Sheet_Name_Here.rename(columns={'полиса': 'ПропускП'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'new-column-knnz': 'Номер полиса'}, inplace=True)

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = IF(
        AND(TYPE(Sheet_Name_Here['ПропускП']) != 'NaN', Sheet_Name_Here['ПропускП'] != ' '),
        Sheet_Name_Here['ПропускП'], None)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'мед.': 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={' услуги': 'Цена'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'оказания': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'МКБ10': 'МКБ'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'услуги': 'Код услуги'}, inplace=True)

    # Filled NaN values in 2 columns in Отчет_по_ДМС_v2
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Filtered Цена
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    # Added column new-column-8cts
    Sheet_Name_Here.insert(11, 'new-column-8cts', 0)

    # Renamed columns ПропускЗ
    Sheet_Name_Here.rename(columns={'№ зуба': 'ПропускЗ'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={'new-column-8cts': '№ зуба'}, inplace=True)

    # Set formula of № зуба
    Sheet_Name_Here['№ зуба'] = IF(AND(TYPE(Sheet_Name_Here['ПропускЗ']) != 'NaN', Sheet_Name_Here['ПропускЗ'] != ' '),
                                   Sheet_Name_Here['ПропускЗ'], None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def poliklinaka_konsultativno_diagnosticheskaya_im_e_m_nigibskogo(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0,
                                       converters={'Код услуги': str})
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ПропускД
    Sheet_Name_Here.rename(columns={'Дата': 'ПропускД'}, inplace=True)

    # Added column new-column-x1to
    Sheet_Name_Here.insert(1, 'new-column-x1to', 0)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'new-column-x1to': 'Дата услуги'}, inplace=True)

    # Added column new-column-fu44
    Sheet_Name_Here.insert(5, 'new-column-fu44', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-fu44': 'ФИО'}, inplace=True)

    # Renamed columns ПропускФ
    Sheet_Name_Here.rename(columns={'Фамилия': 'ПропускФ'}, inplace=True)

    # Renamed columns ПропускИ
    Sheet_Name_Here.rename(columns={'Имя': 'ПропускИ'}, inplace=True)

    # Renamed columns ПропускО
    Sheet_Name_Here.rename(columns={'Отчество': 'ПропускО'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'№ МКБ, код основного заболевания': 'МКБ'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'№ МКБ, наименование основного заболевания': 'Диагноз'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'Количество': 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена услуги': 'Цена'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Сумма к оплате': 'Пропуск'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(
        AND(TYPE(Sheet_Name_Here['ПропускФ']) != 'NaN', TYPE(Sheet_Name_Here['ПропускИ']) != 'NaN',
            TYPE(Sheet_Name_Here['ПропускО']) != 'NaN'),
        PROPER(CONCAT(Sheet_Name_Here['ПропускФ'], ' ', Sheet_Name_Here['ПропускИ'], ' ', Sheet_Name_Here['ПропускО'])),
        None)

    # Set formula of Дата услуги
    Sheet_Name_Here['Дата услуги'] = IF(TYPE(Sheet_Name_Here['ПропускД']) != 'NaN',
                                        LEFT(Sheet_Name_Here['ПропускД'], INT(FIND(Sheet_Name_Here['ПропускД'], ' '))),
                                        None)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def daliz(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-yb9z
    Sheet_Name_Here.insert(3, 'new-column-yb9z', 0)

    # Set formula of new-column-yb9z
    Sheet_Name_Here['new-column-yb9z'] = CONCAT(Sheet_Name_Here['ФАMИЛИЯ'], ' ', Sheet_Name_Here['ИMя'], ' ',
                                                Sheet_Name_Here['Отчество'])

    # Filtered Kод Услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Kод Услуги'].notnull()]

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['new-column-yb9z']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['№Полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Deleted columns ФАMИЛИЯ
    Sheet_Name_Here.drop(['ФАMИЛИЯ'], axis=1, inplace=True)

    # Deleted columns ИMя
    Sheet_Name_Here.drop(['ИMя'], axis=1, inplace=True)

    # Deleted columns Отчество
    Sheet_Name_Here.drop(['Отчество'], axis=1, inplace=True)

    # Added column new-column-p37g
    Sheet_Name_Here.insert(6, 'new-column-p37g', 0)

    # Set formula of new-column-p37g
    Sheet_Name_Here['new-column-p37g'] = Sheet_Name_Here['СуMMа'] / Sheet_Name_Here['Kоличество']

    # Deleted columns СуMMа
    Sheet_Name_Here.drop(['СуMMа'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-yb9z': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№Полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата\nначала услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'new-column-p37g': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Kод MKБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Kод\nвида обращения': 'Диагноз'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def prizvanie(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-827x
    Sheet_Name_Here.insert(1, 'new-column-827x', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-827x': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here['П/п']) != 'NaN', FIND(Sheet_Name_Here['П/п'], ' ') > 0),
                                PROPER(Sheet_Name_Here['П/п']), None)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Filled NaN values in 1 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Added column new-column-vvg9
    Sheet_Name_Here.insert(13, 'new-column-vvg9', 0)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Цена': 'Пропуск'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'new-column-vvg9': 'Цена'}, inplace=True)

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                 FLOAT(SUBSTITUTE(Sheet_Name_Here['Пропуск'], ' ', '')), None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_nmic_onkologii_im_n_n_blohina_min_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=12)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-x4o2
    Sheet_Name_Here.insert(4, 'new-column-x4o2', 0)

    # Renamed columns Пропуск Ф
    Sheet_Name_Here.rename(columns={'Ф.И.О. пациента': 'Пропуск Ф'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-x4o2': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск Ф']) != 'NaN', FIND(Sheet_Name_Here['Пропуск Ф'], '/') <= 0),
        PROPER(Sheet_Name_Here['Пропуск Ф']), None)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Сроки оказания услуги': 'Дата услуги'}, inplace=True)

    # Renamed columns Дата окончания услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 17': 'Дата окончания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код ПМУ': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование ПМУ': 'Наименование услуги'}, inplace=True)

    # Renamed columns Пропуск99
    Sheet_Name_Here.rename(columns={'Сумма': 'Пропуск99'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Дата услуги', 'Дата окончания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_policlinika_nomer_1_uod_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис': 'Номер полиса'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'ФАМИЛИЯ ИО': 'ФИО'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Коды диагнозов': 'МКБ'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуга': 'Наименование услуги'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Сумма': 'Пропуск'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Renamed columns Ghjgecr
    Sheet_Name_Here.rename(columns={'№ ИБ': 'Ghjgecr'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgbu_nmic_cniisichlh_min_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    Sheet_Name_Here.rename(columns={'Услуга': 'Наименование услуги'}, inplace=True)

    # Renamed columns пропуск
    Sheet_Name_Here.rename(columns={'Коэф-т': 'пропуск'}, inplace=True)

    # Renamed columns Пропуск2
    Sheet_Name_Here.rename(columns={'Стоимость': 'Пропуск2'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код Минздрава': 'Код услуги'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'Полис': 'Номер полиса'}, inplace=True)

    # Renamed columns Номер мед.карты
    Sheet_Name_Here.rename(columns={'Карта': 'Номер мед.карты'}, inplace=True)

    # Renamed columns Пропуск Ф
    Sheet_Name_Here.rename(columns={'Фамилия И.О.': 'Пропуск Ф'}, inplace=True)

    # Added column new-column-atah
    Sheet_Name_Here.insert(6, 'new-column-atah', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-atah': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск Ф']) != 'NaN', PROPER(Sheet_Name_Here['Пропуск Ф']), None)

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Наименование услуги'].str.contains('7', na=False)]

    # Added column new-column-81v4
    Sheet_Name_Here.insert(25, 'new-column-81v4', 0)

    # Renamed columns Ghjge
    Sheet_Name_Here.rename(columns={'№ зуба': 'Ghjge'}, inplace=True)

    # Added column new-column-r1hx
    Sheet_Name_Here.insert(26, 'new-column-r1hx', 0)

    # Set formula of new-column-r1hx
    Sheet_Name_Here['new-column-r1hx'] = IF(TYPE(Sheet_Name_Here['Ghjge']) != 'NaN',
                                            SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Ghjge']), '- ', '-'), ' ',
                                                       ', '), None)

    # Renamed columns ПереходныйЗУБ
    Sheet_Name_Here.rename(columns={'new-column-r1hx': 'ПереходныйЗУБ'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={'new-column-81v4': '№ зуба'}, inplace=True)

    # Added column new-column-vtwh
    Sheet_Name_Here.insert(26, 'new-column-vtwh', 0)

    # Renamed columns Переходный зуб2
    Sheet_Name_Here.rename(columns={'new-column-vtwh': 'Переходный зуб2'}, inplace=True)

    # Set formula of Переходный зуб2
    Sheet_Name_Here['Переходный зуб2'] = IF(
        AND(TYPE(Sheet_Name_Here['ПереходныйЗУБ']) != 'NaN', TYPE(Sheet_Name_Here['ПереходныйЗУБ']) != 'object',
            RIGHT(Sheet_Name_Here['ПереходныйЗУБ'], 2) != ', '), Sheet_Name_Here['ПереходныйЗУБ'], None)

    # Added column new-column-veeq
    Sheet_Name_Here.insert(27, 'new-column-veeq', 0)

    # Renamed columns Пропуск999
    Sheet_Name_Here.rename(columns={'new-column-veeq': 'Пропуск999'}, inplace=True)

    # Set formula of Пропуск999
    Sheet_Name_Here['Пропуск999'] = IF(
        AND(TYPE(Sheet_Name_Here['ПереходныйЗУБ']) != 'NaN', RIGHT(Sheet_Name_Here['ПереходныйЗУБ'], 2) == ', '),
        SUBSTITUTE(Sheet_Name_Here['ПереходныйЗУБ'], ', ', ''), None)

    # Set formula of № зуба
    Sheet_Name_Here['№ зуба'] = FILLNAN(Sheet_Name_Here['Переходный зуб2'], Sheet_Name_Here['Пропуск999'])

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def era_1(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код по МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата услуги'}, inplace=True)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'ФИО застрахован. полис': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)

    # Added column new-column-3v5x
    Sheet_Name_Here.insert(7, 'new-column-3v5x', 0)

    # Added column new-column-v82n
    Sheet_Name_Here.insert(7, 'new-column-v82n', 0)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'new-column-v82n': 'Диагноз'}, inplace=True)

    # Renamed columns № зуба
    Sheet_Name_Here.rename(columns={'new-column-3v5x': '№ зуба'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена услуги ': 'Цена'}, inplace=True)

    # Renamed columns Пропуски
    Sheet_Name_Here.rename(columns={'Ст-ть услуг (руб.)': 'Пропуски'}, inplace=True)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Диагноз, номер зуба': 'Пропуск'}, inplace=True)

    # Set formula of № зуба
    Sheet_Name_Here['№ зуба'] = IF(
        AND(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск']), ' ', '') != ''),
        SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(CLEAN(Sheet_Name_Here['Пропуск']), ' ', ''), ')', ''), '(', ''), None)

    # Set formula of Диагноз
    Sheet_Name_Here['Диагноз'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN',
                                    SUBSTITUTE(Sheet_Name_Here['Пропуск'], Sheet_Name_Here['№ зуба'], ''), None)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Наименование услуги по прейскуранту ': 'Код услуги'}, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def cs_32_praktika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    try:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7,
                                           header=None)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        # Renamed columns ФИО
        Sheet_Name_Here.rename(columns={1: 'ФИО'}, inplace=True)

        # Renamed columns Номер полиса
        Sheet_Name_Here.rename(columns={3: 'Номер полиса'}, inplace=True)

        # Renamed columns Дата услуги
        Sheet_Name_Here.rename(columns={4: 'Дата услуги'}, inplace=True)

        # Renamed columns Наименование услуги
        Sheet_Name_Here.rename(columns={5: 'Наименование услуги'}, inplace=True)

        # Renamed columns Код услуги
        Sheet_Name_Here.rename(columns={6: 'Код услуги'}, inplace=True)

        # Added column new-column-kmv6
        Sheet_Name_Here.insert(8, 'new-column-kmv6', 0)

        # Added column new-column-uzq8
        Sheet_Name_Here.insert(8, 'new-column-uzq8', 0)

        # Renamed columns Диагноз
        Sheet_Name_Here.rename(columns={'new-column-uzq8': 'Диагноз'}, inplace=True)

        # Renamed columns № зуба
        Sheet_Name_Here.rename(columns={'new-column-kmv6': '№ зуба'}, inplace=True)

        # Renamed columns Цена
        Sheet_Name_Here.rename(columns={8: 'Цена'}, inplace=True)

        # Renamed columns Кол-во
        Sheet_Name_Here.rename(columns={9: 'Кол-во'}, inplace=True)

        # Renamed columns Скидка
        Sheet_Name_Here.rename(columns={11: 'Скидка'}, inplace=True)

        # Renamed columns Наименование клиники
        Sheet_Name_Here.rename(columns={13: 'Наименование филиала клиники'}, inplace=True)

        # Renamed columns ГП
        Sheet_Name_Here.rename(columns={14: 'ГП'}, inplace=True)

        # Set formula of № зуба
        Sheet_Name_Here['№ зуба'] = IF(AND(TYPE(Sheet_Name_Here[7]) != 'NaN', FIND(Sheet_Name_Here[7], ',')),
                                       SUBSTITUTE(Sheet_Name_Here[7],
                                                  LEFT(Sheet_Name_Here[7], INT(FIND(Sheet_Name_Here[7], ','))), ''),
                                       None)

        # Filtered Кол-во
        Sheet_Name_Here = Sheet_Name_Here[
            (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Кол-во', na=False))]

        # Filtered 2
        Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here[2].notnull()]

        # Changed Кол-во to dtype int
        Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

        # Changed Цена to dtype float
        Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

        # Set formula of Диагноз
        Sheet_Name_Here['Диагноз'] = IF(TYPE(Sheet_Name_Here[7]) != 'NaN',
                                        SUBSTITUTE(Sheet_Name_Here[7], ',' + Sheet_Name_Here['№ зуба'], ''), None)

        # Changed Номер полиса to dtype float
        Sheet_Name_Here['Номер полиса'] = to_float_series(Sheet_Name_Here['Номер полиса'])
    except:
        sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
        Sheet_Name_Here = sheet_df_dictonary[sheet_name]
        # Deleted columns Unnamed: 0
        Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

        # Deleted columns Дата рождения
        Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

        # Deleted columns Стоимость
        Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

        # Deleted columns Скидка
        Sheet_Name_Here.drop(['Скидка'], axis=1, inplace=True)

        # Deleted columns Сумма
        Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

        # Changed Дата оказания услуги to dtype datetime
        Sheet_Name_Here['Дата оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата оказания услуги'],
                                                                 infer_datetime_format=True,
                                                                 errors='coerce')

        # Renamed columns ФИО пациента
        Sheet_Name_Here.rename(columns={'\nФИО': 'ФИО пациента'}, inplace=True)

        # Renamed columns Страховой полис
        Sheet_Name_Here.rename(columns={'ID карта пациента': 'Страховой полис'}, inplace=True)

        # Renamed columns Дата начала оказания услуги
        Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

        # Renamed columns Диагноз
        Sheet_Name_Here.rename(columns={'Диагноз (код по МКБ)': 'Диагноз'}, inplace=True)

        # Renamed columns Цена услуги
        Sheet_Name_Here.rename(columns={'Цена 1-й услуги, руб.': 'Цена услуги'}, inplace=True)

        # Changed Цена услуги to dtype float
        Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

        # Renamed columns Наименование филиала клиники (точки)
        Sheet_Name_Here.rename(columns={'Название клиники': 'Наименование филиала клиники (точки)'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def eramed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3,
                                       converters={'Код услуги': str})
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'ФИО застрахован. полис': 'ФИО'}, inplace=True)

    # Renamed columns Номер полиса
    Sheet_Name_Here.rename(columns={'№ полиса': 'Номер полиса'}, inplace=True)

    # Renamed columns Врач
    Sheet_Name_Here.rename(columns={'Специальность врача': 'Врач'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Код по МКБ-10': 'МКБ'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Наименование услуги по прейскуранту ': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'Цена услуги ': 'Цена'}, inplace=True)

    # Renamed columns Пропустить
    Sheet_Name_Here.rename(columns={'Ст-ть услуг (руб.)': 'Пропустить'}, inplace=True)

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def szgmu_mechnikova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def paracels_krasnodar(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер ID ': 'Страховой полис'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Added column new-column-wag5
    Sheet_Name_Here.insert(3, 'new-column-wag5', 0)

    # Deleted columns new-column-wag5
    Sheet_Name_Here.drop(['new-column-wag5'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nash_doctor_p_mehzavodtest(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис ID': 'Страховой полис'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Стоимость, Скидка, %, Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость', 'Скидка, %', 'Стоимость со скидкой'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nash_doctor_p_mehzavod(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис ID': 'Страховой полис'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Стоимость, Скидка, %, Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость', 'Скидка, %', 'Стоимость со скидкой'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def alfa_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО клиента': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз с кодом по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость по прайсу, руб.': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Deleted columns Сумма, руб.
    Sheet_Name_Here.drop(['Сумма, руб.'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_so_ssp_3(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(4, 'new-column-kcup', 0)
    Sheet_Name_Here.insert(4, 'new-column-dzaj', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={'new-column-kcup': 'ФИО пациента',
                                    'new-column-dzaj': 'Номер полиса',
                                    5: 'Дата услуги',
                                    6: 'Код услуги',
                                    7: 'Наименование услуги',
                                    8: 'Цена услуги',
                                    9: 'Количество',
                                    11: 'Код МКБ-10',
                                    12: 'Номер зуба (для стоматологических услуг)',
                                    13: '№ ГП',
                                    15: 'Врач (ФИО)'
                                    }, inplace=True)
    # Set formula
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here[1], ' ', Sheet_Name_Here[2], ' ', Sheet_Name_Here[3])
    Sheet_Name_Here['Номер полиса'] = IF(TYPE(Sheet_Name_Here[4]) != 'NaN', SUBSTITUTE(Sheet_Name_Here[4], '.', ''),
                                         None)
    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Количество'].notnull()) & (~Sheet_Name_Here['Количество'].str.contains('Кол-во', na=False))]
    # Deleted columns
    Sheet_Name_Here.drop([0, 1, 2, 3, 10, 14], axis=1, inplace=True)
    # Changed dtype
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    Sheet_Name_Here['Количество'] = to_int_series(Sheet_Name_Here['Количество'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def dent_real(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(13, 'new-column-xymf', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={1: 'ФИО пациента',
                                    2: 'Страховой полис',
                                    4: 'Дата начала оказания услуги',
                                    5: 'Код услуги',
                                    6: 'Наименование услуги',
                                    7: 'Цена услуги',
                                    8: 'Кол-во',
                                    10: 'МКБ',
                                    11: 'Номер зуба (для стоматологических услуг)',
                                    'new-column-xymf': '№ ГП',
                                    13: 'Врач (ФИО)'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop([0, 3, 9], axis=1, inplace=True)

    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги', 'МКБ',
                           'Номер зуба (для стоматологических услуг)', '№ ГП', 'Врач (ФИО)']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Set formula
    Sheet_Name_Here['№ ГП'] = IF(TYPE(Sheet_Name_Here[12]) != 'NaN',
                                 LEFT(Sheet_Name_Here[12], INT(FIND(Sheet_Name_Here[12], 'о') - 1)), None)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Количество', na=False))]
    # Changed dtype
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def dent_real_plus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(13, 'new-column-xymf', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={1: 'ФИО пациента',
                                    2: 'Страховой полис',
                                    4: 'Дата начала оказания услуги',
                                    5: 'Код услуги',
                                    6: 'Наименование услуги',
                                    7: 'Цена услуги',
                                    8: 'Кол-во',
                                    10: 'МКБ',
                                    11: 'Номер зуба (для стоматологических услуг)',
                                    'new-column-xymf': '№ ГП',
                                    13: 'Врач (ФИО)'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop([0, 3, 9], axis=1, inplace=True)

    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги', 'МКБ',
                           'Номер зуба (для стоматологических услуг)', '№ ГП', 'Врач (ФИО)']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Set formula
    Sheet_Name_Here['№ ГП'] = IF(TYPE(Sheet_Name_Here[12]) != 'NaN',
                                 LEFT(Sheet_Name_Here[12], INT(FIND(Sheet_Name_Here[12], 'о') - 1)), None)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Количество', na=False))]
    # Changed dtype
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medicina_buduschego(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=11)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Filtered Услуга
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Услуга'].notnull()]

    # Filtered Дата услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата услуги'].notnull()]

    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО Застрахованного': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуга': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена, руб.': 'Цена услуги'}, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns К оплате Заказчиком, руб.
    Sheet_Name_Here.drop(['К оплате Заказчиком, руб.'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_ssp_2(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О. застрахованного лица': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'серия и номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'},
                           inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(
        columns={'Номер пролеч. зубов': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диаг. МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Доктор': 'Врач (ФИО)'}, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[
        Sheet_Name_Here['Код услуги'].notnull()]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns № истории
    Sheet_Name_Here.drop(['№ истории'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Реестр_оказанных_медицинских_ус_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[
        columns_to_fill_nan].fillna(method='ffill')

    # Filtered ФИО пациента
    Sheet_Name_Here = Sheet_Name_Here[
        ~Sheet_Name_Here['ФИО пациента'].str.contains('3', na=False, regex=False)]

    # Changed Номер зуба (для стоматологических услуг) to dtype float
    Sheet_Name_Here['Номер зуба (для стоматологических услуг)'] = to_float_series(
        Sheet_Name_Here['Номер зуба (для стоматологических услуг)'])

    # Added column new-column-eyj8
    Sheet_Name_Here.insert(6, 'new-column-eyj8', 0)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'new-column-eyj8': 'Цена услуги'}, inplace=True)

    # Set formula of Цена услуги
    Sheet_Name_Here['Цена услуги'] = Sheet_Name_Here['Сумма'] / \
                                     Sheet_Name_Here['Количество']

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def spb_gup_peterburskiy_metropoliten(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns ID_номер
    Sheet_Name_Here.drop(['ID_номер'], axis=1, inplace=True)

    # Deleted columns Пол
    Sheet_Name_Here.drop(['Пол'], axis=1, inplace=True)

    # Deleted columns ДР
    Sheet_Name_Here.drop(['ДР'], axis=1, inplace=True)

    # Deleted columns dms_nameKrat
    Sheet_Name_Here.drop(['dms_nameKrat'], axis=1, inplace=True)

    # Deleted columns dms_namePoln
    Sheet_Name_Here.drop(['dms_namePoln'], axis=1, inplace=True)

    # Added column new-column-0u7s
    Sheet_Name_Here.insert(4, 'new-column-0u7s', 0)

    # Set formula of new-column-0u7s
    Sheet_Name_Here['new-column-0u7s'] = CONCAT(Sheet_Name_Here['Фамилия'], ' ', Sheet_Name_Here['Имя'], ' ',
                                                Sheet_Name_Here['Отчество'])

    # Deleted columns Фамилия
    Sheet_Name_Here.drop(['Фамилия'], axis=1, inplace=True)

    # Deleted columns Имя
    Sheet_Name_Here.drop(['Имя'], axis=1, inplace=True)

    # Deleted columns Отчество
    Sheet_Name_Here.drop(['Отчество'], axis=1, inplace=True)

    # Changed Полис to dtype str
    Sheet_Name_Here['Полис'] = Sheet_Name_Here['Полис'].astype('str')

    # Deleted columns Сумма_ск
    Sheet_Name_Here.drop(['Сумма_ск'], axis=1, inplace=True)

    # Changed Цена_ск to dtype float
    Sheet_Name_Here['Цена_ск'] = Sheet_Name_Here['Цена_ск'].astype('float')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-0u7s': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Услуга (код)': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуга назв': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена_ск': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кратность': 'Количество'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Зуб': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ (код)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'s_speshislPol': 'Наименование филиала клиники (точки)'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mfc_garmoniya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер направления': '№ ГП'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата начала. услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'КодУслуги': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Kод': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата  окончания  услуги': 'Дата окончания оказания услуги'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Итого по пациенту
    Sheet_Name_Here.drop(['Итого по пациенту'], axis=1, inplace=True)

    # Deleted columns Код вида обращения
    Sheet_Name_Here.drop(['Код вида обращения'], axis=1, inplace=True)

    # Added column new-column-fh19
    Sheet_Name_Here.insert(4, 'new-column-fh19', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-fh19': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['ФАMИЛИЯ'], ' ', Sheet_Name_Here['Имя'], ' ',
                                             Sheet_Name_Here['Отчество'])

    # Deleted columns Отчество, Имя, ФАMИЛИЯ
    Sheet_Name_Here.drop(['Отчество', 'Имя', 'ФАMИЛИЯ'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Дата окончания оказания услуги to dtype datetime
    Sheet_Name_Here['Дата окончания оказания услуги'] = pd.to_datetime(
        Sheet_Name_Here['Дата окончания оказания услуги'],
        infer_datetime_format=True, errors='coerce')

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = Sheet_Name_Here['Цена услуги'].astype('float')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medecina_alfastrahovaniya_murmansk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns Дата рожд.
    Sheet_Name_Here.drop(['Дата рожд.'], axis=1, inplace=True)

    # Changed Дата услуги to dtype datetime
    Sheet_Name_Here['Дата услуги'] = pd.to_datetime(Sheet_Name_Here['Дата услуги'],
                                                    infer_datetime_format=True, errors='coerce')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Deleted columns Стоимость (руб)
    Sheet_Name_Here.drop(['Стоимость (руб)'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medecina_alfastrahovaniya_tumen(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Цена (руб) to dtype float
    Sheet_Name_Here['Цена (руб)'] = to_float_series(Sheet_Name_Here['Цена (руб)'])

    # Deleted columns Стоимость (руб.)
    Sheet_Name_Here.drop(['Стоимость (руб.)'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vitanika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Номенклатурный код
    Sheet_Name_Here.drop(['Номенклатурный код'], axis=1, inplace=True)

    # Deleted columns Unnamed: 13
    Sheet_Name_Here.drop(['Unnamed: 13'], axis=1, inplace=True)

    # Filled NaN values in 1 columns in Лист_1_1
    columns_to_fill_nan = ['ФИО']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 3 columns in Лист_1_1
    columns_to_fill_nan = ['(ID) Пациента', 'Дата', 'Диагноз']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Deleted columns Итого
    Sheet_Name_Here.drop(['Итого'], axis=1, inplace=True)

    # Filled NaN values in 1 columns in Лист_1_1
    columns_to_fill_nan = ['Врач']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'(ID) Пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость 1 усл.': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stavropolskiy_kraevoy_klinicheskiy_kons_diagnos_center(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость одной услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Количество оказанных услуг': 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Код пациента в СКККДЦ
    Sheet_Name_Here.drop(['Код пациента в СКККДЦ'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Пол
    Sheet_Name_Here.drop(['Пол'], axis=1, inplace=True)

    # Deleted columns Адрес пациента
    Sheet_Name_Here.drop(['Адрес пациента'], axis=1, inplace=True)

    # Deleted columns Сумма по услугам
    Sheet_Name_Here.drop(['Сумма по услугам'], axis=1, inplace=True)

    # Deleted columns Дополнительная информация
    Sheet_Name_Here.drop(['Дополнительная информация'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Filled NaN values in 2 columns in Mail
    columns_to_fill_nan = ['ФИО пациента', 'Полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gobuz_murmanskiy_olastnoy_klinicheskiy_mnogoprof_center(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО                              застрахованного': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса, договора': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз, код МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Filled NaN values in 2 columns in Результат
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Цена, руб.
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена, руб.'].notnull()]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Стоимость, руб.
    Sheet_Name_Here.drop(['Стоимость, руб.'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_moskvi_mknpc_im_a_s_lohinova_dzgm(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(4, 'new-column-04ol', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={2: 'Страховой полис',
                                    'new-column-04ol': '№ ГП',
                                    5: 'ФИО пациента',
                                    7: 'Дата начала оказания услуги',
                                    9: 'Код услуги',
                                    10: 'Наименование услуги',
                                    12: 'Количество',
                                    14: 'Цена услуги'}, inplace=True)
    # Set formula
    Sheet_Name_Here['№ ГП'] = IF(TYPE(Sheet_Name_Here[3]) != 'NaN',
                                 LEFT(Sheet_Name_Here[3], INT(FIND(Sheet_Name_Here[3], 'о') - 1)), None)
    # Filled NaN values
    columns_to_fill_nan = ['Страховой полис', '№ ГП', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Количество'].notnull()) & (~Sheet_Name_Here['Количество'].str.contains('Кол-во', na=False))]
    # Changed dtype
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    # Deleted columns
    Sheet_Name_Here.drop([0, 1, 4, 6, 8, 11, 13, 15, 16], axis=1, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fgau_nmic_mntk_mg_im_akad_s_n_fedorova_min_rf_krasnodar(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса ДМС': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата начала услуги (ДД.ММ.ГГГГ)': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата окончания услуги (ДД.ММ.ГГГГ)': 'Дата окончания оказания услуги'},
                           inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название услуги по прейскуранту': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена услуги по прейскуранту': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по  МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'Номер карты': '№ истории болезни'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач     (Фамилия И.О.)': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер ГП': '№ ГП'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns итого
    Sheet_Name_Here.drop(['итого'], axis=1, inplace=True)

    # Deleted columns Код вида обращения (амб, сто, стц)
    Sheet_Name_Here.drop(['Код вида обращения (амб, сто, стц)'], axis=1, inplace=True)

    # Deleted columns Дата начала ГП
    Sheet_Name_Here.drop(['Дата начала ГП'], axis=1, inplace=True)

    # Filled NaN values in 1 columns in DOGFAMITOG2
    Sheet_Name_Here.fillna({'Количество': 1}, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def farmaciya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    # Deleted columns Стоимость услуги
    Sheet_Name_Here.drop(['Стоимость услуги'], axis=1, inplace=True)

    # Filled NaN values in 1 columns in Sheet1_1
    columns_to_fill_nan = ['ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Sheet1_1
    columns_to_fill_nan = ['Номер инливидуальной карты (ID) Пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Added column new-column-2rc6
    Sheet_Name_Here.insert(2, 'new-column-2rc6', 0)

    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'ФИО пациента': 'Пропуск'}, inplace=True)

    # Set formula of new-column-2rc6
    Sheet_Name_Here['new-column-2rc6'] = IF(Sheet_Name_Here['Пропуск'] != '  ', Sheet_Name_Here['Пропуск'], None)

    # Deleted columns Пропуск
    Sheet_Name_Here.drop(['Пропуск'], axis=1, inplace=True)

    # Changed Номер инливидуальной карты (ID) Пациента to dtype str
    Sheet_Name_Here['Номер инливидуальной карты (ID) Пациента'] = Sheet_Name_Here[
        'Номер инливидуальной карты (ID) Пациента'].astype(
        'str')

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер инливидуальной карты (ID) Пациента': 'Страховой полис'}, inplace=True)

    # Filled NaN values in 1 columns in Sheet1_1
    columns_to_fill_nan = ['new-column-2rc6']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-2rc6': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Filtered Код  услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код  услуги'].notnull()]

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fbuz_pomc_fmba_rossii(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата посещения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуга': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз (МКБ-10)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Расшифровка диагноза': 'Диагноз'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'№ карты': '№ истории болезни'}, inplace=True)

    # Deleted columns № талона
    Sheet_Name_Here.drop(['№ талона'], axis=1, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Результат
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Added column new-column-iw3k
    Sheet_Name_Here.insert(3, 'new-column-iw3k', 0)

    # Set formula of new-column-iw3k
    Sheet_Name_Here['new-column-iw3k'] = SUBSTITUTE(Sheet_Name_Here['Страховой полис'], 'ДМС N ', '')

    # Deleted columns Страховой полис
    Sheet_Name_Here.drop(['Страховой полис'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-iw3k': 'Страховой полис'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_so_samarskaya_gp_6_prom_r(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Серия и № полиса to dtype str
    Sheet_Name_Here['Серия и № полиса'] = Sheet_Name_Here['Серия и № полиса'].astype('str')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф. И. О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Серия и № полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания мед. услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ 10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Общая стоимость мед. услуг. (руб)': 'Цена услуги'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ao_k31_city(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns надбавка (коэфф. Повышения за спецсектор или отделение семейной медицины)
    Sheet_Name_Here.drop(['надбавка (коэфф. Повышения за спецсектор или отделение семейной медицины)'], axis=1,
                         inplace=True)

    # Deleted columns Франшиза %
    Sheet_Name_Here.drop(['Франшиза %'], axis=1, inplace=True)

    # Deleted columns Начислено
    Sheet_Name_Here.drop(['Начислено'], axis=1, inplace=True)

    # Deleted columns Сумма скидки
    Sheet_Name_Here.drop(['Сумма скидки'], axis=1, inplace=True)

    # Deleted columns Начислено к оплате
    Sheet_Name_Here.drop(['Начислено к оплате'], axis=1, inplace=True)

    # Deleted columns Номер ИБ
    Sheet_Name_Here.drop(['Номер ИБ'], axis=1, inplace=True)

    # Deleted columns Клиника-Исполнитель (при Сети Клиник)
    Sheet_Name_Here.drop(['Клиника-Исполнитель (при Сети Клиник)'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Changed № Полиса ДМС to dtype str
    Sheet_Name_Here['№ Полиса ДМС'] = Sheet_Name_Here['№ Полиса ДМС'].astype('str')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ Полиса ДМС': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена услуги по прейскуранту': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Диагноз клинический': 'Диагноз'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Зуб - номер': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Отделение / Специальность': 'Наименование филиала клиники (точки)'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач (ФИО)'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_moskvi_gkb_15_im_o_m_filatova_dzm(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=17)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Пол
    Sheet_Name_Here.drop(['Пол'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Номер истории болезни
    Sheet_Name_Here.drop(['Номер истории болезни'], axis=1, inplace=True)

    # Deleted columns Номер договора
    Sheet_Name_Here.drop(['Номер договора'], axis=1, inplace=True)

    # Deleted columns Работодатель
    Sheet_Name_Here.drop(['Работодатель'], axis=1, inplace=True)

    # Deleted columns Дата гарантийного письма
    Sheet_Name_Here.drop(['Дата гарантийного письма'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Дата конца оказания услуги to dtype datetime
    Sheet_Name_Here['Дата конца оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата конца оказания услуги'],
                                                                   infer_datetime_format=True, errors='coerce')

    # Deleted columns Дата поступления (госпитализация)
    Sheet_Name_Here.drop(['Дата поступления (госпитализация)'], axis=1, inplace=True)

    # Deleted columns Дата выбытия (госпитализация)
    Sheet_Name_Here.drop(['Дата выбытия (госпитализация)'], axis=1, inplace=True)

    # Deleted columns Категория услуги
    Sheet_Name_Here.drop(['Категория услуги'], axis=1, inplace=True)

    # Deleted columns Поверхность зуба
    Sheet_Name_Here.drop(['Поверхность зуба'], axis=1, inplace=True)

    # Deleted columns Комментарий
    Sheet_Name_Here.drop(['Комментарий'], axis=1, inplace=True)

    # Deleted columns Скидка или надбавка
    Sheet_Name_Here.drop(['Скидка или надбавка'], axis=1, inplace=True)

    # Deleted columns Сумма без скидки/надбавки
    Sheet_Name_Here.drop(['Сумма без скидки/надбавки'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер страхового полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Наименование филиала': 'Наименование филиала клиники (точки)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Специальность врача
    Sheet_Name_Here.rename(columns={'Специализация врача': 'Специальность врача'}, inplace=True)

    # Deleted columns Код услуги, СК
    Sheet_Name_Here.drop(['Код услуги, СК'], axis=1, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги, ЛПУ': 'Код услуги'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер гарантийного письма': '№ ГП'}, inplace=True)

    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата конца оказания услуги': 'Дата окончания оказания услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Added column new-column-1vzf
    Sheet_Name_Here.insert(3, 'new-column-1vzf', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-1vzf': 'ФИО пациента'}, inplace=True)

    # Filtered Имя
    Sheet_Name_Here = Sheet_Name_Here[~Sheet_Name_Here['Имя'].str.contains('2', na=False, regex=False)]

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['Фамилия'], ' ', Sheet_Name_Here['Имя'], ' ',
                                             Sheet_Name_Here['Отчество'])

    # Deleted columns Имя, Отчество, Фамилия
    Sheet_Name_Here.drop(['Имя', 'Отчество', 'Фамилия'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def emal(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Deleted columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.drop(['Номер зуба (для стоматологических услуг)'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Changed Страховой полис to dtype str
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Unnamed: 0']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Страховой полис to dtype float
    Sheet_Name_Here['Страховой полис'] = to_float_series(Sheet_Name_Here['Страховой полис'])

    # Filled NaN values in 1 columns in Лист1_1
    columns_to_fill_nan = ['Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Страховой полис to dtype str
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Unnamed: 0': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def uni_medica(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Changed № карты to dtype int
    Sheet_Name_Here['№ карты'] = Sheet_Name_Here['№ карты'].fillna(0).astype('int')

    # Changed Дата приема to dtype datetime
    Sheet_Name_Here['Дата приема'] = pd.to_datetime(Sheet_Name_Here['Дата приема'], infer_datetime_format=True,
                                                    errors='coerce')

    # Deleted columns Категория
    Sheet_Name_Here.drop(['Категория'], axis=1, inplace=True)

    # Changed Сумма to dtype float
    Sheet_Name_Here['Сумма'] = Sheet_Name_Here['Сумма'].astype('float')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Доктор': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'№ карты': '№ истории болезни'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата приема': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nadezhda(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис ID': 'Страховой полис'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed № истории болезни to dtype str
    Sheet_Name_Here['№ истории болезни'] = Sheet_Name_Here['№ истории болезни'].astype('str')

    # Filtered Дата начала оказания услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата начала оказания услуги'].notnull()]

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vita_smile(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О. пациента': 'ФИО пациента'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Гарантийное письмо': '№ ГП'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата лечения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Манипуляции, материалы': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Deleted columns Unnamed: 2
    Sheet_Name_Here.drop(['Unnamed: 2'], axis=1, inplace=True)

    # Deleted columns Unnamed: 3
    Sheet_Name_Here.drop(['Unnamed: 3'], axis=1, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Deleted columns Unnamed: 5
    Sheet_Name_Here.drop(['Unnamed: 5'], axis=1, inplace=True)

    # Deleted columns Unnamed: 6
    Sheet_Name_Here.drop(['Unnamed: 6'], axis=1, inplace=True)

    # Deleted columns Unnamed: 7
    Sheet_Name_Here.drop(['Unnamed: 7'], axis=1, inplace=True)

    # Deleted columns Unnamed: 8
    Sheet_Name_Here.drop(['Unnamed: 8'], axis=1, inplace=True)

    # Deleted columns Unnamed: 9
    Sheet_Name_Here.drop(['Unnamed: 9'], axis=1, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Deleted columns Unnamed: 11
    Sheet_Name_Here.drop(['Unnamed: 11'], axis=1, inplace=True)

    # Deleted columns Unnamed: 12
    Sheet_Name_Here.drop(['Unnamed: 12'], axis=1, inplace=True)

    # Deleted columns Unnamed: 13
    Sheet_Name_Here.drop(['Unnamed: 13'], axis=1, inplace=True)

    # Deleted columns Unnamed: 14
    Sheet_Name_Here.drop(['Unnamed: 14'], axis=1, inplace=True)

    # Deleted columns Unnamed: 15
    Sheet_Name_Here.drop(['Unnamed: 15'], axis=1, inplace=True)

    # Deleted columns Unnamed: 16
    Sheet_Name_Here.drop(['Unnamed: 16'], axis=1, inplace=True)

    # Deleted columns Unnamed: 18
    Sheet_Name_Here.drop(['Unnamed: 18'], axis=1, inplace=True)

    # Deleted columns Unnamed: 19
    Sheet_Name_Here.drop(['Unnamed: 19'], axis=1, inplace=True)

    # Deleted columns Unnamed: 20
    Sheet_Name_Here.drop(['Unnamed: 20'], axis=1, inplace=True)

    # Deleted columns Unnamed: 22
    Sheet_Name_Here.drop(['Unnamed: 22'], axis=1, inplace=True)

    # Deleted columns Unnamed: 23
    Sheet_Name_Here.drop(['Unnamed: 23'], axis=1, inplace=True)

    # Deleted columns Unnamed: 24
    Sheet_Name_Here.drop(['Unnamed: 24'], axis=1, inplace=True)

    # Deleted columns Unnamed: 26
    Sheet_Name_Here.drop(['Unnamed: 26'], axis=1, inplace=True)

    # Deleted columns Unnamed: 27
    Sheet_Name_Here.drop(['Unnamed: 27'], axis=1, inplace=True)

    # Deleted columns Unnamed: 28
    Sheet_Name_Here.drop(['Unnamed: 28'], axis=1, inplace=True)

    # Deleted columns Unnamed: 30
    Sheet_Name_Here.drop(['Unnamed: 30'], axis=1, inplace=True)

    # Deleted columns Unnamed: 31
    Sheet_Name_Here.drop(['Unnamed: 31'], axis=1, inplace=True)

    # Deleted columns Unnamed: 32
    Sheet_Name_Here.drop(['Unnamed: 32'], axis=1, inplace=True)

    # Deleted columns Unnamed: 34
    Sheet_Name_Here.drop(['Unnamed: 34'], axis=1, inplace=True)

    # Deleted columns Unnamed: 35
    Sheet_Name_Here.drop(['Unnamed: 35'], axis=1, inplace=True)

    # Deleted columns Unnamed: 36
    Sheet_Name_Here.drop(['Unnamed: 36'], axis=1, inplace=True)

    # Deleted columns Unnamed: 38
    Sheet_Name_Here.drop(['Unnamed: 38'], axis=1, inplace=True)

    # Deleted columns Unnamed: 39
    Sheet_Name_Here.drop(['Unnamed: 39'], axis=1, inplace=True)

    # Deleted columns Unnamed: 40
    Sheet_Name_Here.drop(['Unnamed: 40'], axis=1, inplace=True)

    # Deleted columns Unnamed: 42
    Sheet_Name_Here.drop(['Unnamed: 42'], axis=1, inplace=True)

    # Deleted columns Unnamed: 43
    Sheet_Name_Here.drop(['Unnamed: 43'], axis=1, inplace=True)

    # Deleted columns Unnamed: 44
    Sheet_Name_Here.drop(['Unnamed: 44'], axis=1, inplace=True)

    # Deleted columns Unnamed: 46
    Sheet_Name_Here.drop(['Unnamed: 46'], axis=1, inplace=True)

    # Deleted columns Unnamed: 47
    Sheet_Name_Here.drop(['Unnamed: 47'], axis=1, inplace=True)

    # Deleted columns Unnamed: 48
    Sheet_Name_Here.drop(['Unnamed: 48'], axis=1, inplace=True)

    # Deleted columns Unnamed: 50
    Sheet_Name_Here.drop(['Unnamed: 50'], axis=1, inplace=True)

    # Deleted columns Unnamed: 51
    Sheet_Name_Here.drop(['Unnamed: 51'], axis=1, inplace=True)

    # Deleted columns Unnamed: 52
    Sheet_Name_Here.drop(['Unnamed: 52'], axis=1, inplace=True)

    # Deleted columns Unnamed: 54
    Sheet_Name_Here.drop(['Unnamed: 54'], axis=1, inplace=True)

    # Deleted columns Unnamed: 55
    Sheet_Name_Here.drop(['Unnamed: 55'], axis=1, inplace=True)

    # Deleted columns Unnamed: 56
    Sheet_Name_Here.drop(['Unnamed: 56'], axis=1, inplace=True)

    # Deleted columns Unnamed: 58
    Sheet_Name_Here.drop(['Unnamed: 58'], axis=1, inplace=True)

    # Deleted columns Unnamed: 59
    Sheet_Name_Here.drop(['Unnamed: 59'], axis=1, inplace=True)

    # Deleted columns Unnamed: 60
    Sheet_Name_Here.drop(['Unnamed: 60'], axis=1, inplace=True)

    # Deleted columns Unnamed: 62
    Sheet_Name_Here.drop(['Unnamed: 62'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Unnamed: 63
    Sheet_Name_Here.drop(['Unnamed: 63'], axis=1, inplace=True)

    # Deleted columns Unnamed: 64
    Sheet_Name_Here.drop(['Unnamed: 64'], axis=1, inplace=True)

    # Deleted columns К оплате страховой компанией
    Sheet_Name_Here.drop(['К оплате страховой компанией'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Filled NaN values in 8 columns in Лист1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', '№ ГП', 'Дата начала оказания услуги', 'Врач (ФИО)',
                           'Номер зуба (для стоматологических услуг)', 'Код МКБ-10', 'Диагноз']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Added column new-column-wksz
    Sheet_Name_Here.insert(2, 'new-column-wksz', 0)

    # Set formula of new-column-wksz
    Sheet_Name_Here['new-column-wksz'] = SUBSTITUTE(Sheet_Name_Here['Страховой полис'], '№ ', '')

    # Deleted columns Страховой полис
    Sheet_Name_Here.drop(['Страховой полис'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-wksz': 'Страховой полис'}, inplace=True)

    # Added column new-column-b7k6
    Sheet_Name_Here.insert(7, 'new-column-b7k6', 0)

    # Set formula of new-column-b7k6
    Sheet_Name_Here['new-column-b7k6'] = CLEAN(Sheet_Name_Here['Номер зуба (для стоматологических услуг)'])

    # Deleted columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.drop(['Номер зуба (для стоматологических услуг)'], axis=1, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'new-column-b7k6': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nmic_medica_mente(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Added column new-column-xbf0
    Sheet_Name_Here.insert(4, 'new-column-xbf0', 0)

    # Set formula of new-column-xbf0
    Sheet_Name_Here['new-column-xbf0'] = CONCAT(Sheet_Name_Here[2], ' ', Sheet_Name_Here[3], ' ', Sheet_Name_Here[4])

    # Deleted columns 5
    Sheet_Name_Here.drop([5], axis=1, inplace=True)

    # Deleted columns 2
    Sheet_Name_Here.drop([2], axis=1, inplace=True)

    # Deleted columns 3
    Sheet_Name_Here.drop([3], axis=1, inplace=True)

    # Deleted columns 4
    Sheet_Name_Here.drop([4], axis=1, inplace=True)

    # Changed 10 to dtype float
    Sheet_Name_Here[10] = to_float_series(Sheet_Name_Here[10])

    # Changed 11 to dtype int
    Sheet_Name_Here[11] = Sheet_Name_Here[11].fillna(0).astype('int')

    # Deleted columns 12
    Sheet_Name_Here.drop([12], axis=1, inplace=True)

    # Filtered 9
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here[9].notnull()]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-xbf0': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={6: 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={1: 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={7: 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={8: 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={9: 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={10: 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={11: 'Количество'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def vdc_med_center(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Серия, номер полиса/гарант.письма': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Колич.': 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз МКБ 10': 'Код МКБ-10'}, inplace=True)

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Место работы
    Sheet_Name_Here.drop(['Место работы'], axis=1, inplace=True)

    # Deleted columns № амб. карты
    Sheet_Name_Here.drop(['№ амб. карты'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Filled NaN values in 1 columns in TDSheet
    columns_to_fill_nan = ['ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mc_naedine_n(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[-1]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={1: 'Страховой полис',
                                    2: 'ФИО пациента',
                                    3: 'Код МКБ-10',
                                    5: 'Дата начала оказания услуги',
                                    6: 'Код услуги',
                                    7: 'Наименование услуги',
                                    8: 'Количество',
                                    9: 'Цена услуги'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop([10, 0, 4], axis=1, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Количество'].notnull()) &
                                      (~Sheet_Name_Here['Количество'].str.contains('Кол-во услуг', na=False))]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    # Changed dtype
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def aldent(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Filled NaN values in 3 columns in Лист1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Changed Номер зуба (для стоматологических услуг) to dtype float
    Sheet_Name_Here['Номер зуба (для стоматологических услуг)'] = to_float_series(
        Sheet_Name_Here['Номер зуба (для стоматологических услуг)'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ssk_zubastik(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(3, 'new-column-joro', 0)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    # Renamed columns
    Sheet_Name_Here.rename(columns={' № Полиса': 'Страховой полис',
                                    'Дата  услуги': 'Дата начала оказания услуги',
                                    'Сумма': 'Цена услуги',
                                    'KодMKБ': 'Код МКБ-10',
                                    'Номер зуба': 'Номер зуба (для стоматологических услуг)',
                                    'Наименование услуг': 'Наименование услуги',
                                    'new-column-joro': 'ФИО пациента'}, inplace=True)
    # Filled NaN values
    columns_to_fill_nan = ['ФАMИЛИЯ', 'Имя', 'Отчество', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Set formula
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['ФАMИЛИЯ'], ' ', Sheet_Name_Here['Имя'], ' ',
                                             Sheet_Name_Here['Отчество'])
    # Deleted columns
    Sheet_Name_Here.drop(['ФАMИЛИЯ', 'Имя', 'Отчество'], axis=1, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ommc_im_cv_luki(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Фамилия Имя Отчетсво': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование мед. услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Исполнитель': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Filled NaN values in 1 columns in TDSheet
    columns_to_fill_nan = ['ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Added column new-column-ja1k
    Sheet_Name_Here.insert(2, 'new-column-ja1k', 0)

    # Set formula of new-column-ja1k
    Sheet_Name_Here['new-column-ja1k'] = SUBSTITUTE(Sheet_Name_Here['Страховой полис'], '№ ', '')

    # Deleted columns Страховой полис
    Sheet_Name_Here.drop(['Страховой полис'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-ja1k': 'Страховой полис'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def norma_xxi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[-1]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Changed dtype
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filled NaN values in 3 columns in ЯНВАРЬ_2023_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата оказания услуги', 'Код МКБ-10', 'Диагноз']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed dtype
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Renamed columns
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги',
                                    'Кол-во': 'Количество'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def noy(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns 1
    Sheet_Name_Here.drop([1], axis=1, inplace=True)

    # Filtered 5
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here[5].notnull()]

    # Changed 9 to dtype datetime
    Sheet_Name_Here[9] = pd.to_datetime(Sheet_Name_Here[9], infer_datetime_format=True, errors='coerce')

    # Filled NaN values in 1 columns in TDSheet_1
    columns_to_fill_nan = [2]
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 2 columns in TDSheet_1
    columns_to_fill_nan = [2, 3]
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 3 columns in TDSheet_1
    columns_to_fill_nan = [2, 3, 9]
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed 6 to dtype int
    Sheet_Name_Here[6] = Sheet_Name_Here[6].fillna(0).astype('int')

    # Deleted columns 8
    Sheet_Name_Here.drop([8], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={2: 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={3: 'ФИО пациента'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={4: 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={5: 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={6: 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={7: 'Цена услуги'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={9: 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={11: 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={12: 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Unnamed: 11': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatolog_i_ya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Поверхность зуба
    Sheet_Name_Here.drop(['Поверхность зуба'], axis=1, inplace=True)

    # Changed Количество оказанных услуг to dtype int
    Sheet_Name_Here['Количество оказанных услуг'] = Sheet_Name_Here['Количество оказанных услуг'].fillna(0).astype(
        'int')

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Deleted columns Полная стоимость
    Sheet_Name_Here.drop(['Полная стоимость '], axis=1, inplace=True)

    # Changed Дата оказания услуги to dtype datetime
    Sheet_Name_Here['Дата оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата оказания услуги'],
                                                             infer_datetime_format=True, errors='coerce')

    # Filtered Полное наименование услуги по прейскуранту
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Полное наименование услуги по прейскуранту'].notnull()]

    # Filled NaN values in 3 columns in Лицевой_счет_1
    columns_to_fill_nan = ['Полное наименование услуги по прейскуранту', 'ФИО застрахованного (полностью)', '№ Полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО застрахованного (полностью)': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ Полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба (по международной нумерации) ': 'Номер зуба (для стоматологических услуг)'},
                           inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Диагноз (по МКБ 10)': 'Диагноз'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту ЛПУ': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Полное наименование услуги по прейскуранту': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Количество оказанных услуг': 'Количество'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medosmotr(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Changed Дата to dtype datetime
    Sheet_Name_Here['Дата'] = pd.to_datetime(Sheet_Name_Here['Дата'], infer_datetime_format=True, errors='coerce')

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость': 'Цена услуги'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ip_korotchik_u_o(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Лист1_1
    columns_to_fill_nan = ['ФИО пациента', 'Номер индивидуальной карты (ID) Пациента', 'дата оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Номер индивидуальной карты (ID) Пациента to dtype str
    Sheet_Name_Here['Номер индивидуальной карты (ID) Пациента'] = Sheet_Name_Here[
        'Номер индивидуальной карты (ID) Пациента'].astype(
        'str')

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер индивидуальной карты (ID) Пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ulibka_plus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed № истории болезни to dtype int
    Sheet_Name_Here['№ истории болезни'] = Sheet_Name_Here['№ истории болезни'].fillna(0).astype('int')

    # Filtered Дата начала оказания услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Дата начала оказания услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def a_stom(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=12)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз (код по МКБ-10)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Общая сумма
    Sheet_Name_Here.drop(['Общая сумма'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in Sheet1_1
    columns_to_fill_nan = ['Страховой полис', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Sheet1_1
    columns_to_fill_nan = ['Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def klinika_semeynoy_medecini(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Страховой полис': '№ ГП'}, inplace=True)

    # Deleted columns Срок действия полиса
    Sheet_Name_Here.drop(['Срок действия полиса'], axis=1, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата (период) лечения\t': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns 	Итого, руб.
    Sheet_Name_Here.drop(['\tИтого, руб.'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость, руб.': 'Цена услуги'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def medgard(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Added column new-column-39dl
    Sheet_Name_Here.insert(2, 'new-column-39dl', 0)

    # Added column new-column-035v
    Sheet_Name_Here.insert(2, 'new-column-035v', 0)

    # Added column new-column-fnvq
    Sheet_Name_Here.insert(2, 'new-column-fnvq', 0)

    # Added column new-column-wm8o
    Sheet_Name_Here.insert(3, 'new-column-wm8o', 0)

    # Added column new-column-f0rx
    Sheet_Name_Here.insert(4, 'new-column-f0rx', 0)

    # Added column new-column-fjn5
    Sheet_Name_Here.insert(5, 'new-column-fjn5', 0)

    # Added column new-column-hi66
    Sheet_Name_Here.insert(8, 'new-column-hi66', 0)

    # Added column new-column-gfb0
    Sheet_Name_Here.insert(8, 'new-column-gfb0', 0)

    # Renamed columns ФИО, Номер полиса, Дата услуги, Код услуги, Наименование услуги, Цена, Кол-во, МКБ
    Sheet_Name_Here.rename(columns={'new-column-035v': 'ФИО',
                                    'new-column-39dl': 'Номер полиса',
                                    2: 'Дата услуги',
                                    3: 'Код услуги',
                                    4: 'Наименование услуги',
                                    5: 'Цена',
                                    6: 'Кол-во',
                                    8: 'МКБ',
                                    'new-column-fnvq': 'FIO1',
                                    'new-column-wm8o': 'FIO2',
                                    'new-column-f0rx': 'FIO3',
                                    'new-column-fjn5': 'FIO4',
                                    'new-column-gfb0': 'pol1',
                                    'new-column-hi66': 'pol2'}, inplace=True)

    # Renamed columns № зуба
    try:
        Sheet_Name_Here.rename(columns={9: '№ зуба'}, inplace=True)
    except:
        pass

    # Set formula of FIO1
    Sheet_Name_Here['FIO1'] = IF(
        AND(TYPE(Sheet_Name_Here[1]) != 'NaN', Sheet_Name_Here[1] != ' '),
        PROPER(SUBSTITUTE(Sheet_Name_Here[1], LEFT(Sheet_Name_Here[1], 3), '')), None)

    # Set formula of FIO2
    Sheet_Name_Here['FIO2'] = IF(
        AND(TYPE(Sheet_Name_Here['FIO1']) != 'NaN', TYPE(Sheet_Name_Here['FIO1']) != 'object'),
        SUBSTITUTE(Sheet_Name_Here['FIO1'], SUBSTITUTE(CLEAN(Sheet_Name_Here['FIO1']), ' ', ''), ''),
        None)

    # Set formula of FIO3
    Sheet_Name_Here['FIO3'] = IF(
        AND(TYPE(Sheet_Name_Here['FIO2']) != 'NaN', TYPE(Sheet_Name_Here['FIO2']) != 'object',
            FIND(Sheet_Name_Here['FIO2'], '-') <= 2), Sheet_Name_Here['FIO2'], None)

    # Set formula of FIO4
    Sheet_Name_Here['FIO4'] = IF(
        AND(TYPE(Sheet_Name_Here['FIO3']) == 'object', TYPE(Sheet_Name_Here['FIO2']) != 'NaN',
            TYPE(Sheet_Name_Here['FIO2']) != 'object',
            FIND(Sheet_Name_Here['FIO2'], 'По Пациенту') < 2),
        SUBSTITUTE(Sheet_Name_Here['FIO2'], RIGHT(Sheet_Name_Here['FIO2'], 15), ''), None)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = FILLNAN(Sheet_Name_Here['FIO3'], Sheet_Name_Here['FIO4'])

    # Set formula of pol1
    Sheet_Name_Here['pol1'] = IF(
        AND(TYPE(Sheet_Name_Here[1]) != 'NaN', Sheet_Name_Here[1] != ' '),
        PROPER(SUBSTITUTE(Sheet_Name_Here['FIO1'], Sheet_Name_Here['ФИО'], '')), None)

    # Set formula of pol2
    Sheet_Name_Here['pol2'] = FLOAT(Sheet_Name_Here['pol1'])

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = FILLNAN(Sheet_Name_Here['pol2'], Sheet_Name_Here['pol1'])

    # Filled NaN values in 2 columns in _Услуги_пациент_МКБ_ЭМК
    columns_to_fill_nan = ['ФИО', 'Номер полиса']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Наименование услуги'].notnull()) & (
        ~Sheet_Name_Here['Наименование услуги'].str.contains('Название услуги', na=False))]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def professionalnaya_medicinskaya_liga(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Added column
    Sheet_Name_Here.insert(4, 'new-column-qtec', 0)
    # Sheet_Name_Here.insert(33, 'new-column-fnw4', 0) # слот для скидки ( временно не нужно )
    Sheet_Name_Here.insert(33, 'new-column-c7ww', 0)

    # Renamed columns
    Sheet_Name_Here.rename(columns={'new-column-qtec': 'ФИО',
                                    7: 'Номер полиса',
                                    10: 'Отделение',
                                    11: 'МКБ',
                                    12: 'Диагноз',
                                    14: 'Врач',
                                    15: 'Специальность доктора',
                                    16: 'Код врача',
                                    17: 'Наименование услуги',
                                    19: 'Код услуги',
                                    21: '№ ГП',
                                    22: 'Дата услуги',
                                    23: 'Дата окончания',
                                    27: '№ зуба',
                                    30: 'Кол-во',
                                    'new-column-c7ww': 'Цена',
                                    # 'new-column-fnw4': 'Скидка' # слот для скидки ( временно не нужно )
                                    }, inplace=True)

    # Set formula
    Sheet_Name_Here['ФИО'] = IF(AND(TYPE(Sheet_Name_Here[1]) != 'NaN', TYPE(Sheet_Name_Here[3]) != 'NaN'),
                                SUBSTITUTE(
                                    CONCAT(Sheet_Name_Here[1], ' ', Sheet_Name_Here[2], '  ', Sheet_Name_Here[3]), '  ',
                                    ' '), None)
    # Sheet_Name_Here['Скидка'] = IF(TYPE(Sheet_Name_Here[37]) != 'NaN', Sheet_Name_Here[37], None)
    Sheet_Name_Here['Цена'] = IF(AND(TYPE(Sheet_Name_Here[36]) != 'NaN', TYPE(Sheet_Name_Here['Кол-во']) != 'NaN'),
                                 (FLOAT(Sheet_Name_Here[36]) - FLOAT(Sheet_Name_Here[32])) / FLOAT(
                                     Sheet_Name_Here['Кол-во']), None)

    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Кол-во'].notnull()) & (~Sheet_Name_Here['Кол-во'].str.contains('Количество', na=False))]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here[29].isna()]

    # Changed dtype
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_kkb_2(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Added column
    Sheet_Name_Here.insert(2, 'new-column-ns32', 0)
    Sheet_Name_Here.insert(2, 'new-column-jjlh', 0)
    Sheet_Name_Here.insert(2, 'new-column-28g5', 0)
    Sheet_Name_Here.insert(2, 'new-column-k0nh', 0)
    Sheet_Name_Here.insert(2, 'new-column-qsg3', 0)
    Sheet_Name_Here.insert(2, 'new-column-fpma', 0)

    # Renamed columns
    Sheet_Name_Here.rename(columns={'new-column-ns32': 'Цена',
                                    'new-column-28g5': 'Дата услуги',
                                    'new-column-jjlh': 'Диагноз',
                                    'new-column-fpma': 'ФИО',
                                    'new-column-qsg3': 'Номер полиса',
                                    'new-column-k0nh': '№ ГП',
                                    1: 'Отделение',
                                    2: 'faosgyui',
                                    3: 'Код услуги',
                                    4: 'Наименование услуги',
                                    5: 'МКБ',
                                    7: '№ зуба',
                                    8: 'Кол-во',
                                    9: 'ASFVHUKJ'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(
        AND(TYPE(Sheet_Name_Here['Отделение']) != 'NaN', FIND(Sheet_Name_Here['Отделение'], 'Пол:') > 1), PROPER(
            SUBSTITUTE(Sheet_Name_Here['Отделение'], RIGHT(Sheet_Name_Here['Отделение'],
                                                           LEN(Sheet_Name_Here['Отделение']) - FIND(
                                                               Sheet_Name_Here['Отделение'], 'Пол:') + 2), '')), None)

    # Set formula of Номер полиса
    Sheet_Name_Here['Номер полиса'] = IF(
        AND(TYPE(Sheet_Name_Here['Отделение']) != 'NaN', FIND(Sheet_Name_Here['Отделение'], 'Полис') > 1), FLOAT(LEFT(
            SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['Отделение'],
                                  LEFT(Sheet_Name_Here['Отделение'], INT(FIND(Sheet_Name_Here['Отделение'], 'Полис'))),
                                  ''), 'олис № ', ''), INT(FIND(SUBSTITUTE(SUBSTITUTE(Sheet_Name_Here['Отделение'],
                                                                                      LEFT(Sheet_Name_Here['Отделение'],
                                                                                           INT(FIND(Sheet_Name_Here[
                                                                                                        'Отделение'],
                                                                                                    'Полис'))), ''),
                                                                           'олис № ', ''), '№') - 2))), None)

    # Set formula of № ГП
    Sheet_Name_Here['№ ГП'] = IF(
        AND(TYPE(Sheet_Name_Here['Отделение']) != 'NaN', FIND(Sheet_Name_Here['Отделение'], 'Полис') > 1), SUBSTITUTE(
            SUBSTITUTE(Sheet_Name_Here['Отделение'],
                       LEFT(Sheet_Name_Here['Отделение'], INT(FIND(Sheet_Name_Here['Отделение'], 'письма'))), ''),
            'исьма: ', ''), None)

    # Set formula of Дата услуги
    Sheet_Name_Here['Дата услуги'] = IF(AND(TYPE(Sheet_Name_Here['faosgyui']) != 'Nan',
                                            LEN(SUBSTITUTE(CLEAN(Sheet_Name_Here['faosgyui']), ' ', '')) > INT(
                                                Sheet_Name_Here[0])), CLEAN(Sheet_Name_Here['faosgyui']), None)

    # Set formula of Диагноз
    Sheet_Name_Here['Диагноз'] = IF(
        AND(TYPE(Sheet_Name_Here['faosgyui']) != 'NaN', TYPE(Sheet_Name_Here['Дата услуги']) == 'object'),
        Sheet_Name_Here['faosgyui'], None)

    # Filled NaN values in 4 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО', 'Номер полиса', '№ ГП', 'Диагноз']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='bfill')

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Filtered Номер полиса
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Номер полиса'].notnull()]

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Set formula of Цена
    Sheet_Name_Here['Цена'] = FLOAT(SUBSTITUTE(Sheet_Name_Here['ASFVHUKJ'], ' ', ''))

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def bu_rlg_dvv_min_chuvash(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(11, 'new-column-wi77', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Фамилия имя отчество застрахованного                         ': 'ФИО',
                                    '№ гарантийнного письма': '№ ГП',
                                    'Полис застрахованного': 'Страховой полис',
                                    'Дата оказания мед.услуги': 'Дата начала оказания услуги',
                                    'Диагноз по МКБ10': 'Код МКБ-10',
                                    'Наименование мед.услуги': 'Наименование услуги',
                                    'Цена (руб)': 'Цена услуги',
                                    'new-column-wi77': 'Количество'},
                           inplace=True)
    # Set formula
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Стоимость (руб)'] / Sheet_Name_Here['Цена услуги']
    # Deleted columns
    Sheet_Name_Here.drop(['№ п/п', 'Дата рождения', 'Кол-во', 'Стоимость (руб)'], axis=1, inplace=True)
    # Filled NaN values
    columns_to_fill_nan = ['ФИО', '№ ГП', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги '].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def npo_vcpb_ugmed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента',
                                    'Номер': 'Страховой полис',
                                    'Гарантийное письмо': '№ ГП',
                                    'Дата док': 'Дата начала оказания услуги',
                                    'Код': 'Код услуги',
                                    'Услуга': 'Наименование услуги',
                                    'Принял': 'Врач',
                                    'МКБ10': 'Код МКБ-10',
                                    'Unnamed: 18': 'Количество',
                                    'Unnamed: 20': 'Цена услуги'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['№ п/п', 'Номер док', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4', 'СКА', 'С', 'По',
                          'Тип услуги', 'Unnamed: 19', 'Unnamed: 21', 'Направил'], axis=1, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['ФИО пациента'].notnull()]
    # Changed dtype
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ugmed_minus_m(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента',
                                    'Номер': 'Страховой полис',
                                    'Гарантийное письмо': '№ ГП',
                                    'Дата док': 'Дата начала оказания услуги',
                                    'Код': 'Код услуги',
                                    'Услуга': 'Наименование услуги',
                                    'Принял': 'Врач',
                                    'МКБ10': 'Код МКБ-10',
                                    'Unnamed: 19': 'Количество',
                                    'Unnamed: 21': 'Цена услуги'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['№ п/п', 'Номер док', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4', 'СКА', 'С', 'Unnamed: 7',
                          'По', 'Тип услуги', 'Unnamed: 22', 'Unnamed: 20', 'Направил'], axis=1, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['ФИО пациента'].notnull()]
    # Changed dtype
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def rzd_semashko(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name],
                                       skiprows=5)  # 5 , header=None
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-gixn
    Sheet_Name_Here.insert(3, 'new-column-gixn', 0)
    Sheet_Name_Here.insert(6, 'new-column-oubr', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-gixn': 'ФИО пациента',
                                    'new-column-oubr': 'Страховой полис',
                                    '№ полиса\nДМС/наим-\nенование предприятия\n': '№ полисаДМС',
                                    'Кол-во услуг или \nкой-ко дней\n': 'Количество',
                                    'Стои\nмость услуг \nили \nкой-ко дня\n': 'Цена услуги',
                                    'Дата оказания услуг или начала и окончания госпитали-\nзации\n': 'Дата начала оказания услуги',
                                    'Код отделе-\nния\n': 'Код отделения',
                                    'Шифр заболе-\nвания по МКБ Х\n': 'Код МКБ-10'}, inplace=True)
    # Changed dtype
    Sheet_Name_Here['ФИО пациента'] = Sheet_Name_Here['ФИО пациента'].astype('str')
    # Set formula
    Sheet_Name_Here['ФИО пациента'] = IF(
        AND(TYPE(1) != 'NaN', SUBSTITUTE(Sheet_Name_Here['ФИО застрахованного'], "  ", " ") != " "),
        Sheet_Name_Here['ФИО застрахованного'], None)
    Sheet_Name_Here['Страховой полис'] = IF(
        AND(TYPE(1) != 'NaN', SUBSTITUTE(Sheet_Name_Here['№ полисаДМС'], "  ", " ") != " "),
        Sheet_Name_Here['№ полисаДМС'], None)
    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    # Deleted columns
    Sheet_Name_Here.drop(['Unnamed: 0', '№  ИБ', 'ФИО застрахованного', 'Дата рождения', '№ полисаДМС',
                          'Итого\n(руб.)\n', 'Итого со скидкой\n(руб.)\n'], axis=1, inplace=True)
    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def avicenna(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'ID карта пациента': 'Страховой полис',
                                    'Дата оказания услуги': 'Дата начала оказания услуги',
                                    'Кол-во': 'Количество'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['Дата\nрождения', 'Стоимость', 'Стоимость со скидкой', 'Код/название клиники',
                          '№ ГП'], axis=1, inplace=True)
    # Changed dtype
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код \nуслуги'].notnull()]
    # Changed dtype
    Sheet_Name_Here['Страховой полис'] = to_float_series(Sheet_Name_Here['Страховой полис'])
    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Changed dtype
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')
    Sheet_Name_Here['Страховой полис'] = to_int_series(Sheet_Name_Here['Страховой полис'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mc_cup_medprofi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={' Полис ID': 'Страховой полис',
                                    'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)
    Sheet_Name_Here.drop(['Дата рождения', 'Стоимость'], axis=1, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def clinika_persona(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={' Полис ID': 'Страховой полис',
                                    'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)
    Sheet_Name_Here.drop(['Дата рождения', 'Стоимость'], axis=1, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ivstroiteh(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'№ полиса ': 'Страховой полис',
                                    'Дата оказания услуги': 'Дата начала оказания услуги',
                                    ' Код по МКБ-10': 'Код МКБ-10',
                                    'Количесство': 'Количество'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['Дата рождения', 'Стоимость', 'Стоимость со скидкой', '№ истории болезни'], axis=1,
                         inplace=True)
    # Changed dtype
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def zdorovie_ludi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Номер страхового полиса': 'Страховой полис',
                                    'Код заболевания (МКБ)': 'Код МКБ-10',
                                    'Дата услуги': 'Дата начала оказания услуги',
                                    'Кол-во услуг': 'Количество',
                                    'Цена, руб.': 'Цена услуги'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['Unnamed: 0', '№ п/п', 'Unnamed: 4', 'Unnamed: 9', 'Unnamed: 10',
                          'Стоимость, руб.'], axis=1, inplace=True)
    # Filled NaN values
    columns_to_fill_nan = ['ФИО Пациента', 'Страховой полис', 'Код МКБ-10']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def sovermed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Номер страхового полиса': 'Страховой полис',
                                    'Код заболевания (МКБ)': 'Код МКБ-10',
                                    'Номер зуба': 'Номер зуба (для стоматологических услуг)',
                                    'Дата услуги': 'Дата начала оказания услуги',
                                    'Кол-во услуг': 'Количество',
                                    'Цена, руб.': 'Цена услуги'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['Unnamed: 0', '№ п/п', 'Дата рождения ', 'Unnamed: 5', 'Unnamed: 11', 'Unnamed: 12',
                          'Стоимость, руб.'], axis=1, inplace=True)
    # Filled NaN values
    columns_to_fill_nan = ['ФИО Пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ldc_mibs_cheboksari(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns
    Sheet_Name_Here.drop(['№пп', '  Сумма'], axis=1, inplace=True)
    # Renamed columns
    Sheet_Name_Here.rename(columns={'Номер         страхового           полиса (карты)': 'Страховой полис',
                                    '    Фамилия И.О.       застрахованного': 'ФИО пациента',
                                    'код диагноза': 'Код МКБ-10',
                                    '  Дата оказания       услуги ': 'Дата начала оказания услуги',
                                    'Код услуги по       прейску-ранту  ': 'Код услуги',
                                    'Кол-во услуг': 'Количество',
                                    '  Цена': 'Цена услуги',
                                    '      Наименование услуги': 'Наименование услуги'}, inplace=True)
    # Changed dtype
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mmc_urp_pro(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={'ID номер': 'Страховой полис',
                                    'Ф.И.О (полностью)': 'ФИО пациента',
                                    'Дата оказания услуги': 'Дата начала оказания услуги',
                                    'Код услуги по прейскуранту': 'Код услуги',
                                    'Наименование услуги по прейскуранту ': 'Наименование услуги',
                                    'Диагноз': 'Код МКБ-10',
                                    'Кол-во': 'Количество',
                                    'Цена услуги (руб.)': 'Цена услуги'}, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['№п/п', 'Ст-ть услуг (руб.)'], axis=1, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def razu_dent(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'номер ID': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'количество': 'Количество'}, inplace=True)

    # Renamed columns Скидка, %
    Sheet_Name_Here.rename(columns={'скидка %': 'Скидка, %'}, inplace=True)

    # Renamed columns Код филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Код филиала клиники ': 'Код филиала клиники (точки)'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Наименование филиала клиники ': 'Наименование филиала клиники (точки)'},
                           inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'№ГП': '№ ГП'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач ФИО': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код отделения
    Sheet_Name_Here.rename(columns={'код отделения ': 'Код отделения'}, inplace=True)

    # Renamed columns Наименование отделения
    Sheet_Name_Here.rename(columns={'наименование отделения': 'Наименование отделения'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns стоимость
    Sheet_Name_Here.drop(['стоимость '], axis=1, inplace=True)

    # Deleted columns стоимость со скидкой
    Sheet_Name_Here.drop(['стоимость со скидкой'], axis=1, inplace=True)

    # Added column new-column-aalc
    Sheet_Name_Here.insert(8, 'new-column-aalc', 0)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'new-column-aalc': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Set formula of Номер зуба (для стоматологических услуг)
    Sheet_Name_Here['Номер зуба (для стоматологических услуг)'] = IF(
        AND(TYPE(Sheet_Name_Here['Номер зуба ']) != 'NaN', FIND(Sheet_Name_Here['Номер зуба '], ':') > 1),
        SUBSTITUTE(Sheet_Name_Here['Номер зуба '],
                   LEFT(Sheet_Name_Here['Номер зуба '], INT(FIND(Sheet_Name_Here['Номер зуба '], ':') + 1)), ''), None)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Deleted columns Номер зуба
    Sheet_Name_Here.drop(['Номер зуба '], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def lpc_dent_liniya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата и время': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полисы': 'Страховой полис'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Зуб': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуги': 'Наименование услуги'}, inplace=True)

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomatologiya_2(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Гарантийное письмо, №
    Sheet_Name_Here.drop(['Гарантийное письмо, №'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Сумма предъяв- ленная заказчику
    Sheet_Name_Here.drop(['Сумма предъяв- ленная заказчику'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата   лечения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Ф.И.О. врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба (стом)': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Номер зуба (для стоматологических услуг)': 'Диагноз'}, inplace=True)

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in февраль_2
    columns_to_fill_nan = ['Страховой полис', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    # Changed Страховой полис to dtype str
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')

    # Filled NaN values in 1 columns in февраль_2
    Sheet_Name_Here.fillna({'Страховой полис': 0}, inplace=True)

    # Filled NaN values in 3 columns in февраль_2
    columns_to_fill_nan = ['Страховой полис', 'Врач (ФИО)', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def keller_stachki(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'№ И.Б.': '№ истории болезни'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ страхового полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО застрахованного': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Deleted columns Поверхность зуба
    Sheet_Name_Here.drop(['Поверхность зуба'], axis=1, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Cт-ть услуг,руб
    Sheet_Name_Here.drop(['Cт-ть услуг,руб'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(
        Sheet_Name_Here['Дата начала оказания услуги'], infer_datetime_format=True, errors='coerce')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered ФИО пациента
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['ФИО пациента'].notnull()]

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def fbuz_msch_9_fbma_r_dubna(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата\nоказания\nуслуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер страхового полиса': 'Страховой полис'}, inplace=True)

    # Deleted columns Unnamed: 2
    Sheet_Name_Here.drop(['Unnamed: 2'], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз': 'Код МКБ-10'}, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.\nзастрахованного': 'ФИО пациента'}, inplace=True)

    # Deleted columns Unnamed: 6
    Sheet_Name_Here.drop(['Unnamed: 6'], axis=1, inplace=True)
    # Deleted columns
    Sheet_Name_Here.drop(['Номер\nлицевого\nсчета'], axis=1, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Deleted columns Unnamed: 11
    Sheet_Name_Here.drop(['Unnamed: 11'], axis=1, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во\nуслуг': 'Количество'}, inplace=True)

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Отделение/Врач': 'Врач (ФИО)'}, inplace=True)

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def mc_kirli(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обраще- ния': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги по прайсу': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол- во услуг': 'Количество'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость '], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер и дата гарантийного письма': '№ ГП'}, inplace=True)

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered ФИО пациента
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['ФИО пациента'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def stomat_center_vladmiva(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО (застрахованного)': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба ': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Deleted columns Количество
    Sheet_Name_Here.drop(['Количество'], axis=1, inplace=True)

    # Deleted columns Поверхность зуба
    Sheet_Name_Here.drop(['Поверхность зуба'], axis=1, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование услуги по прейскуранту': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Количество оказанных услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Deleted columns Полная стоимость
    Sheet_Name_Here.drop(['Полная стоимость'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Реестр_услуг_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nac_dia_centr(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата ': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ Полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ (диагноз)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Deleted columns Страховая компания
    Sheet_Name_Here.drop(['Страховая компания'], axis=1, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Changed Номер зуба (для стоматологических услуг) to dtype float
    Sheet_Name_Here['Номер зуба (для стоматологических услуг)'] = to_float_series(
        Sheet_Name_Here['Номер зуба (для стоматологических услуг)'])

    # Added column new-column-a976
    Sheet_Name_Here.insert(7, 'new-column-a976', 0)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'new-column-a976': 'Количество'}, inplace=True)

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = Sheet_Name_Here['Цена услуги'].astype('float')

    # Filled NaN values in 1 columns in Лист1
    Sheet_Name_Here.fillna({'Количество': 1}, inplace=True)

    # Set formula of Количество
    Sheet_Name_Here['Количество'] = 1

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def nasha_klinika(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО застрахованного': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоим. услуги ': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во ': 'Количество'}, inplace=True)

    # Deleted columns Сумма по услугам
    Sheet_Name_Here.drop(['Сумма по услугам '], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ услуги': 'Код МКБ-10'}, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Filled NaN values in 3 columns in Лист1_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def klinika_sov_med_nab_chelni(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Застрахованный': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена, руб.': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз или код (по МКБ-10)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код врача
    Sheet_Name_Here.rename(columns={'Код врача (или ФИО)': 'Врач (ФИО)'}, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Страховой полис', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Deleted 1 row in Sheet_Name_Here
    Sheet_Name_Here.drop(labels=[0], inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Deleted columns Стоимость, руб.
    Sheet_Name_Here.drop(['Стоимость, руб.'], axis=1, inplace=True)

    # Deleted columns Примечание
    Sheet_Name_Here.drop(['Примечание'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def spb_gbuz_gkdc_1(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Доктор': 'Врач (ФИО)'}, inplace=True)

    # Deleted columns Код пациента
    Sheet_Name_Here.drop(['Код пациента'], axis=1, inplace=True)

    # Deleted columns Дата рожд.
    Sheet_Name_Here.drop(['Дата рожд.'], axis=1, inplace=True)

    # Deleted columns Стоимость (руб)
    Sheet_Name_Here.drop(['Стоимость (руб)'], axis=1, inplace=True)

    # Deleted columns Сумма по прейскуранту
    Sheet_Name_Here.drop(['Сумма по прейскуранту'], axis=1, inplace=True)

    # Deleted columns Место работы
    Sheet_Name_Here.drop(['Место работы'], axis=1, inplace=True)

    # Deleted columns Программа прикрепления
    Sheet_Name_Here.drop(['Программа прикрепления'], axis=1, inplace=True)

    # Deleted columns Место обслуживания
    Sheet_Name_Here.drop(['Место обслуживания'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = Sheet_Name_Here['Цена услуги'].astype('float')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def chuz_poli_ovum(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns № пп
    Sheet_Name_Here.drop(['№ пп'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ договора (полиса)': 'Страховой полис'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Список анализов': 'Наименование услуги'}, inplace=True)

    # Deleted columns Общая стоимость
    Sheet_Name_Here.drop(['Общая стоимость'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Changed Дата to dtype datetime
    Sheet_Name_Here['Дата'] = pd.to_datetime(Sheet_Name_Here['Дата'], infer_datetime_format=True, errors='coerce')

    # Filled NaN values in 1 columns in БестДоктор_1
    columns_to_fill_nan = ['Дата']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gauz_rkod_mz_rt_im_m_z_sigala(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID карта пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Количество услуг': 'Количество'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Название клиники': 'Наименование филиала клиники (точки)'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач (исполнитель)': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'№ГП': '№ ГП'}, inplace=True)

    # Filled NaN values in 4 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Дата начала оказания услуги', 'Страховой полис', 'Код МКБ-10']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Deleted columns №п/п
    Sheet_Name_Here.drop(['№п/п'], axis=1, inplace=True)

    # Deleted columns Код по номенклатуре
    Sheet_Name_Here.drop(['Код по номенклатуре'], axis=1, inplace=True)

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gauz_cgrb_18(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №п/п
    Sheet_Name_Here.drop(['№п/п'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер ID карты': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Количество': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Цена': 'Количество'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 1 columns in df_06_07_2_1
    columns_to_fill_nan = ['Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def prime_stomatology(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №№ пп
    Sheet_Name_Here.drop(['№№ пп'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Фамилия, имя, отчество ': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Цена услуги по прейскуранту
    Sheet_Name_Here.drop(['Цена услуги по прейскуранту'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуг': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Deleted columns Поверхность зуба
    Sheet_Name_Here.drop(['Поверхность зуба'], axis=1, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Дата и номер гарантийного письма': '№ ГП'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Changed Номер зуба (для стоматологических услуг) to dtype int
    Sheet_Name_Here['Номер зуба (для стоматологических услуг)'] = Sheet_Name_Here[
        'Номер зуба (для стоматологических услуг)'].fillna(
        0).astype('int')

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Filtered ФИО пациента
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['ФИО пациента'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def denterprice(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ10  Диагноз': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Added column new-column-n4ux
    Sheet_Name_Here.insert(4, 'new-column-n4ux', 0)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-n4ux': 'Страховой полис'}, inplace=True)

    # Changed Страховой полис to dtype float
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('float')

    # Set formula of Страховой полис
    Sheet_Name_Here['Страховой полис'] = SUBSTITUTE(Sheet_Name_Here['№ полиса'], "№ СП: ", "")

    # Changed Страховой полис to dtype float
    Sheet_Name_Here['Страховой полис'] = to_float_series(Sheet_Name_Here['Страховой полис'])

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Deleted columns №№ пп
    Sheet_Name_Here.drop(['№№ пп'], axis=1, inplace=True)

    # Deleted columns № полиса
    Sheet_Name_Here.drop(['№ полиса'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def gbuz_lo_kirishskaya_kmb(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Unnamed: 1': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Unnamed: 2': 'Страховой полис'}, inplace=True)

    # Deleted columns Unnamed: 3
    Sheet_Name_Here.drop(['Unnamed: 3'], axis=1, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Added column new-column-pgjz
    Sheet_Name_Here.insert(3, 'new-column-pgjz', 0)

    # Deleted columns Unnamed: 5
    Sheet_Name_Here.drop(['Unnamed: 5'], axis=1, inplace=True)

    # Deleted columns new-column-pgjz
    Sheet_Name_Here.drop(['new-column-pgjz'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 6': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Unnamed: 7, Unnamed: 8
    Sheet_Name_Here.drop(['Unnamed: 7', 'Unnamed: 8'], axis=1, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'НАЗВАНИЕ': 'Наименование услуги'}, inplace=True)

    # Deleted columns Unnamed: 11
    Sheet_Name_Here.drop(['Unnamed: 11'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'ЦЕНА': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'КОЛ-ВО': 'Количество'}, inplace=True)

    # Deleted columns СУММА
    Sheet_Name_Here.drop(['СУММА'], axis=1, inplace=True)

    # Deleted columns Unnamed: 15
    Sheet_Name_Here.drop(['Unnamed: 15'], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Unnamed: 16': 'Код МКБ-10'}, inplace=True)

    # Deleted columns Unnamed: 18, Unnamed: 19, Unnamed: 20, Unnamed: 21, Unnamed: 22
    Sheet_Name_Here.drop(['Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20', 'Unnamed: 21', 'Unnamed: 22'], axis=1,
                         inplace=True)

    # Deleted columns Unnamed: 17
    Sheet_Name_Here.drop(['Unnamed: 17'], axis=1, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Unnamed: 23': '№ ГП'}, inplace=True)

    # Deleted columns Unnamed: 24
    Sheet_Name_Here.drop(['Unnamed: 24'], axis=1, inplace=True)

    # Deleted columns Unnamed: 25, Unnamed: 26, Unnamed: 27, ФАМ.
    Sheet_Name_Here.drop(['Unnamed: 25', 'Unnamed: 26', 'Unnamed: 27', 'ФАМ.'], axis=1, inplace=True)

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def on_clinic_ryazan(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Unnamed: 3
    Sheet_Name_Here.drop(['Unnamed: 3'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Deleted columns Unnamed: 5
    Sheet_Name_Here.drop(['Unnamed: 5'], axis=1, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование услуги по прейскуранту': 'Наименование услуги'}, inplace=True)

    # Deleted columns Unnamed: 9
    Sheet_Name_Here.drop(['Unnamed: 9'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма оказанных': 'Цена услуги'}, inplace=True)

    # Deleted columns Unnamed: 11
    Sheet_Name_Here.drop(['Unnamed: 11'], axis=1, inplace=True)

    # Deleted columns Общая стоимость
    Sheet_Name_Here.drop(['Общая стоимость'], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Ф.И.О.врача': 'Врач (ФИО)'}, inplace=True)

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Ф.И.О.пациента
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Ф.И.О.пациента'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def clinica_sov_stom_art_c(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns № пп.
    Sheet_Name_Here.drop(['№ пп.'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О. пациента': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns стоимость
    Sheet_Name_Here.drop(['стоимость'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Ст-ть манип.': 'Цена услуги'}, inplace=True)

    # Deleted columns количество
    Sheet_Name_Here.drop(['количество'], axis=1, inplace=True)

    # Deleted columns Общая стоимость
    Sheet_Name_Here.drop(['Общая стоимость'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Renamed columns Код филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'код филиала': 'Код филиала клиники (точки)'}, inplace=True)

    # Renamed columns Код врача
    Sheet_Name_Here.rename(columns={'код врача': 'Код врача'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Специальность врача
    Sheet_Name_Here.rename(columns={'Специальность': 'Специальность врача'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filled NaN values in 3 columns in Лист1_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 4 columns in Лист1_1
    columns_to_fill_nan = ['Код МКБ-10', 'Диагноз', 'Номер зуба (для стоматологических услуг)',
                           'Код филиала клиники (точки)']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 2 columns in Лист1_1
    columns_to_fill_nan = ['Врач (ФИО)', 'Специальность врача']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def ldc_razumed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис ID': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Filled NaN values in 10 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги', 'Код МКБ-10', 'Диагноз',
                           '№ ГП', 'Врач (ФИО)', 'Специальность врача', 'Специальность направившего врача',
                           '№ истории болезни']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)

def ssmc_region_2_ryazan(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование выполненной услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Сумма рублей
    Sheet_Name_Here.drop(['Сумма рублей'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def ssmc_region_2_tula(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование выполненной услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Сумма рублей
    Sheet_Name_Here.drop(['Сумма рублей'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def clinic_uralskaya(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО застрахованного': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Данные полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обследования': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена за единицу': 'Цена услуги'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Deleted columns Всего
    Sheet_Name_Here.drop(['Всего'], axis=1, inplace=True)

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def centr_prof_med_ultramed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Фамилия   Имя   Отчество ': 'ФИО пациента'}, inplace=True)
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата нач.\n леч-я': 'Дата начала оказания услуги'}, inplace=True)
    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оконч.\n леч-я': 'Дата окончания оказания услуги'}, inplace=True)
    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)
    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'код услуг': 'Код услуги'}, inplace=True)
    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'врач': 'Врач (ФИО)'}, inplace=True)
    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ 10': 'Код МКБ-10'}, inplace=True)
    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)
    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п '], axis=1, inplace=True)
    # Deleted columns Год рождения
    Sheet_Name_Here.drop(['Год рождения'], axis=1, inplace=True)
    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]
    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')
    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def med_art_tomsk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Клиент Ф.И.О. застрахованного': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Сотрудник': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Клиника - Исполнитель': 'Наименование филиала клиники (точки)'}, inplace=True)

    # Deleted columns N
    Sheet_Name_Here.drop(['N'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Профиль оказанной медицинской помощи
    Sheet_Name_Here.drop(['Профиль оказанной медицинской помощи'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def stomatology_clinic_karat(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Серия, номер\nполиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач/лаборант': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Зубы': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз / Код МКБ10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Процедура': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Added column new-column-c1oq
    Sheet_Name_Here.insert(11, 'new-column-c1oq', 0)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'new-column-c1oq': 'Цена услуги'}, inplace=True)

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = Sheet_Name_Here['Цена услуги'].astype('float')
    # Changed Стоимость
    Sheet_Name_Here['Стоимость\nприема'] = to_float_series(Sheet_Name_Here['Стоимость\nприема'])

    # Renamed columns Стоимость
    Sheet_Name_Here.rename(columns={'Стоимость\nприема': 'Стоимость'}, inplace=True)

    # Filled NaN values in 6 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги', 'Врач (ФИО)',
                           'Номер зуба (для стоматологических услуг)', 'Код МКБ-10']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Количество to dtype float
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].astype('float')

    # Set formula of Цена услуги
    Sheet_Name_Here['Цена услуги'] = IF(
        AND(TYPE(Sheet_Name_Here['Стоимость']) != 'NaN', TYPE(Sheet_Name_Here['Количество']) != 'NaN'),
        Sheet_Name_Here['Стоимость'] / Sheet_Name_Here['Количество'], None)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)
    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма\nпо ДМС,\nдоговору'], axis=1, inplace=True)

    # Deleted columns Unnamed: 2
    Sheet_Name_Here.drop(['Unnamed: 2'], axis=1, inplace=True)

    # Deleted columns Поверхность
    Sheet_Name_Here.drop(['Поверхность'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def gauz_no_gcp_1(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол- во': 'Количество'}, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in Лист1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def fgau_nmic_nero_im_burdenko_min_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=22)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Гарантийное письмо': '№ ГП'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Исполнитель': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype(
        'int')

    # Filled NaN values in 2 columns in Реестр_132_от_13_03_2023_ООО__Б_1
    columns_to_fill_nan = ['№ ГП', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[
        columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[
        Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def med_centr_nadezhda(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Unnamed: 1': '№ ГП'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Unnamed: 2': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Unnamed: 3': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Unnamed: 4': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 5': 'Наименование услуги'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 6': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Unnamed: 7
    Sheet_Name_Here.drop(['Unnamed: 7'], axis=1, inplace=True)

    # Reordered column Unnamed: 8
    Sheet_Name_Here_columns = [col for col in Sheet_Name_Here.columns if col != 'Unnamed: 8']
    Sheet_Name_Here_columns.insert(5, 'Unnamed: 8')
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here_columns]

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 8': 'Цена услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Unnamed: 11': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Unnamed: 9': 'Количество'}, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(
        Sheet_Name_Here['Дата начала оказания услуги'], infer_datetime_format=True, errors='coerce')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filled NaN values in 3 columns in Бест_Доктор_ФЕВРАЛЬ_1
    columns_to_fill_nan = ['№ ГП', 'ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Бест_Доктор_ФЕВРАЛЬ_1
    columns_to_fill_nan = ['Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def gippokrat(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=1)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по Прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена услуги по Прейскуранту': 'Цена услуги'}, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Added column new-column-tdza
    Sheet_Name_Here.insert(4, 'new-column-tdza', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-tdza': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['Фамилия'], " ", Sheet_Name_Here['Имя'], " ",
                                             Sheet_Name_Here['Отчество'])

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Deleted columns Фамилия
    Sheet_Name_Here.drop(['Фамилия'], axis=1, inplace=True)

    # Deleted columns Имя
    Sheet_Name_Here.drop(['Имя'], axis=1, inplace=True)

    # Deleted columns Отчество
    Sheet_Name_Here.drop(['Отчество'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def dent_art_novokuzneck(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=15)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ ID карты': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Диагноз (код по МКБ-10)': 'Диагноз'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол- во услуг': 'Количество'}, inplace=True)

    # Deleted columns Общая сумма
    Sheet_Name_Here.drop(['Общая сумма'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in Sheet1_1
    columns_to_fill_nan = ['Страховой полис', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Страховой полис to dtype int
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def gbauz_so_gkb_40(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Направление': '№ ГП'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={' Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата поступления': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата выписки': 'Дата окончания оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Артикул': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование услуг': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Наименование отделения
    Sheet_Name_Here.rename(columns={'Отделение': 'Наименование отделения'}, inplace=True)

    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Адрес
    Sheet_Name_Here.drop(['Адрес'], axis=1, inplace=True)

    # Deleted columns Коэффициент
    Sheet_Name_Here.drop(['Коэффициент'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма '], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def ao_kamskiy_dia_cent_medikam(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={' Полис ID': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def clinic_dnk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Changed Номер индивидуальной карты (ID) Пациента to dtype int
    Sheet_Name_Here['Номер индивидуальной карты (ID) Пациента'] = Sheet_Name_Here['Номер индивидуальной карты (ID) Пациента'].fillna(
        0).astype('int')

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер индивидуальной карты (ID) Пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def mega_cent(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Услуга': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Услуга.1': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагнозы': 'Код МКБ-10'}, inplace=True)

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns День рождения
    Sheet_Name_Here.drop(['День рождения'], axis=1, inplace=True)

    # Deleted columns Код
    Sheet_Name_Here.drop(['Код'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def ohta_dental_plus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена,руб.': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'кол-во': 'Количество'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач': 'Врач (ФИО)'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Дата и № ГП
    Sheet_Name_Here.drop(['Дата и № ГП'], axis=1, inplace=True)

    # Deleted columns Сумма, руб
    Sheet_Name_Here.drop(['Сумма, руб'], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def raduga_med(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={' Полис ID': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Лист1_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def sana_ko(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ПОЛИС': 'Страховой полис'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'КОД УСЛУГИ': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'НАИМЕНОВАНИЕ УСЛУГИ': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'ЦЕНА': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'КОЛ.  УСЛУГ': 'Количество'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'ДАТА ОКАЗАНИЯ УСЛУГИ ': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код врача
    Sheet_Name_Here.rename(columns={'КОД врача': 'Код врача'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Deleted columns СТОИМОСТЬ
    Sheet_Name_Here.drop(['СТОИМОСТЬ'], axis=1, inplace=True)

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def anay_ribinsk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'код услуги': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена услуги'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'ФИО врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость (руб)
    Sheet_Name_Here.drop(['Стоимость (руб)'], axis=1, inplace=True)

    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')

    # Changed Количество to dtype float
    Sheet_Name_Here['Количество'] = to_float_series(Sheet_Name_Here['Количество'])

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def dc_clinic_city(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Unnamed: 2
    Sheet_Name_Here.drop(['Unnamed: 2'], axis=1, inplace=True)

    # Deleted columns Unnamed: 3
    Sheet_Name_Here.drop(['Unnamed: 3'], axis=1, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'Медицинская карта': '№ истории болезни'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Deleted columns Unnamed: 6
    Sheet_Name_Here.drop(['Unnamed: 6'], axis=1, inplace=True)

    # Deleted columns Unnamed: 7
    Sheet_Name_Here.drop(['Unnamed: 7'], axis=1, inplace=True)

    # Deleted columns Паспорт, Unnamed: 9, Unnamed: 10
    Sheet_Name_Here.drop(['Паспорт', 'Unnamed: 9', 'Unnamed: 10'], axis=1, inplace=True)

    # Deleted columns Адрес, Unnamed: 12, Unnamed: 13
    Sheet_Name_Here.drop(['Адрес', 'Unnamed: 12', 'Unnamed: 13'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Данные полиса': 'Страховой полис'}, inplace=True)

    # Deleted columns Unnamed: 15, Unnamed: 16, Unnamed: 17, Unnamed: 18
    Sheet_Name_Here.drop(['Unnamed: 15', 'Unnamed: 16', 'Unnamed: 17', 'Unnamed: 18'], axis=1, inplace=True)

    # Deleted columns Unnamed: 20, Unnamed: 21
    Sheet_Name_Here.drop(['Unnamed: 20', 'Unnamed: 21'], axis=1, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)

    # Deleted columns Unnamed: 23, Unnamed: 24, Unnamed: 25
    Sheet_Name_Here.drop(['Unnamed: 23', 'Unnamed: 24', 'Unnamed: 25'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата вып-ия': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Unnamed: 27, Unnamed: 28
    Sheet_Name_Here.drop(['Unnamed: 27', 'Unnamed: 28'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Deleted columns Цена услуги
    Sheet_Name_Here.drop(['Цена услуги'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def moy_zubnoi(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=11)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Unnamed: 0': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 1': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Unnamed: 2': 'Страховой полис'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Unnamed: 3': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'МКБ10 Диагноз': 'Диагноз'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Filled NaN values in 4 columns in Лист3_1
    columns_to_fill_nan = ['ФИО пациента', 'Дата начала оказания услуги', 'Страховой полис', 'Врач (ФИО)']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def medin(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена, руб': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Unnamed: 2
    Sheet_Name_Here.drop(['Unnamed: 2'], axis=1, inplace=True)

    # Deleted columns Unnamed: 3
    Sheet_Name_Here.drop(['Unnamed: 3'], axis=1, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Deleted columns Unnamed: 5
    Sheet_Name_Here.drop(['Unnamed: 5'], axis=1, inplace=True)

    # Deleted columns Unnamed: 8
    Sheet_Name_Here.drop(['Unnamed: 8'], axis=1, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Deleted columns Unnamed: 12
    Sheet_Name_Here.drop(['Unnamed: 12'], axis=1, inplace=True)

    # Deleted columns Unnamed: 13
    Sheet_Name_Here.drop(['Unnamed: 13'], axis=1, inplace=True)

    # Deleted columns Unnamed: 15
    Sheet_Name_Here.drop(['Unnamed: 15'], axis=1, inplace=True)

    # Deleted columns Unnamed: 16
    Sheet_Name_Here.drop(['Unnamed: 16'], axis=1, inplace=True)

    # Deleted columns Unnamed: 17
    Sheet_Name_Here.drop(['Unnamed: 17'], axis=1, inplace=True)

    # Deleted columns Unnamed: 18
    Sheet_Name_Here.drop(['Unnamed: 18'], axis=1, inplace=True)

    # Deleted columns Unnamed: 19
    Sheet_Name_Here.drop(['Unnamed: 19'], axis=1, inplace=True)

    # Deleted columns Unnamed: 20
    Sheet_Name_Here.drop(['Unnamed: 20'], axis=1, inplace=True)

    # Deleted columns Сумма, руб
    Sheet_Name_Here.drop(['Сумма, руб'], axis=1, inplace=True)

    # Changed Страховой полис to dtype float
    Sheet_Name_Here['Страховой полис'] = to_float_series(Sheet_Name_Here['Страховой полис'])

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def new_dent_toliati(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID карта пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код \nуслуги': 'Код услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена \nуслуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Код/название клиники': 'Наименование филиала клиники (точки)'}, inplace=True)
    # Deleted columns Дата
    Sheet_Name_Here.drop(['Дата\nрождения'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')

    # Changed Страховой полис to dtype str
    Sheet_Name_Here['Страховой полис'] = Sheet_Name_Here['Страховой полис'].astype('str')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def chuz_rzd_med_arhangelsk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО застрахованных': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(
        columns={'Наименование медицинской услуги или ее кода по прейскуранту': 'Наименование услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата начала оказания мед. услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Дата окончания оказания услуги
    Sheet_Name_Here.rename(columns={'Дата конца оказания мед. услуги': 'Дата окончания оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Количество услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги, руб': 'Цена услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Специалист': 'Врач (ФИО)'}, inplace=True)

    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Deleted columns № п.п
    Sheet_Name_Here.drop(['№ п.п'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Общая сумма, руб
    Sheet_Name_Here.drop(['Общая сумма, руб'], axis=1, inplace=True)

    # Deleted columns Врач (ФИО)
    Sheet_Name_Here.drop(['Врач (ФИО)'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def poly_1_vita_medikus(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №п/п
    Sheet_Name_Here.drop(['№п/п'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО страхового пациента': 'ФИО пациента'}, inplace=True)

    # Deleted columns Дата рождения пациента
    Sheet_Name_Here.drop(['Дата рождения пациента'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ страх. полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код статистики': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование оказанной услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Количество услуги': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена за 1 усл. Страховая 1': 'Цена услуги'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Плательщик
    Sheet_Name_Here.drop(['Плательщик'], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Диагноз
    Sheet_Name_Here.rename(columns={'Клинический диагноз': 'Диагноз'}, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def med_centr_sel(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Unnamed: 0
    Sheet_Name_Here.drop(['Unnamed: 0'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Unnamed: 1': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Unnamed: 2': 'Страховой полис'}, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'Unnamed: 3': '№ истории болезни'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 4': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Unnamed: 5': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Cумма
    Sheet_Name_Here.drop(['Cумма'], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Unnamed: 10': 'Код МКБ-10'}, inplace=True)

    # Deleted columns (амб, сто)
    Sheet_Name_Here.drop(['(амб, сто)'], axis=1, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Фамилия': 'Врач (ФИО)'}, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                            infer_datetime_format=True, errors='coerce')

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    # Changed № истории болезни to dtype int
    Sheet_Name_Here['№ истории болезни'] = Sheet_Name_Here['№ истории болезни'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def med_cent_stolica(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'номер страхового полиса (гарантийного письма)': 'Страховой полис'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'наименование услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'код услуги': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз': 'Код МКБ-10'}, inplace=True)

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filtered Страховой полис
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Страховой полис'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def mlcd_diagnost(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент\n(ФИО)': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения\n(сроки лечения)': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз (код по МКБ-10)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код\nуслуги': 'Код услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Added column new-column-2iru
    Sheet_Name_Here.insert(3, 'new-column-2iru', 0)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-2iru': 'Страховой полис'}, inplace=True)

    # Set formula of Страховой полис
    Sheet_Name_Here['Страховой полис'] = SUBSTITUTE(Sheet_Name_Here['№ полиса '], "ID ", '')

    # Changed Страховой полис to dtype float
    Sheet_Name_Here['Страховой полис'] = to_float_series(Sheet_Name_Here['Страховой полис'])

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns № полиса
    Sheet_Name_Here.drop(['№ полиса '], axis=1, inplace=True)

    # Deleted columns Итого
    Sheet_Name_Here.drop(['Итого'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def gauz_gkb_7_kazani(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={2: 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={3: 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={4: 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'6': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={7: 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={8: 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={9: 'Количество'}, inplace=True)

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Deleted columns 1
    Sheet_Name_Here.drop([1], axis=1, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={5: 'Код МКБ-10'}, inplace=True)

    # Changed Код МКБ-10 to dtype str
    Sheet_Name_Here['Код МКБ-10'] = Sheet_Name_Here['Код МКБ-10'].astype('str')

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Deleted columns 10
    Sheet_Name_Here.drop([10], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def mc_medeor(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=12)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-t1wm
    Sheet_Name_Here.insert(1, 'new-column-t1wm', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-t1wm': 'ФИО пациента'}, inplace=True)

    # Added column new-column-vhlo
    Sheet_Name_Here.insert(2, 'new-column-vhlo', 0)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-vhlo': 'Страховой полис'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = LEFT(Sheet_Name_Here['Клиент'], INT(FIND(Sheet_Name_Here['Клиент'], 'Номер') - 1))

    # Set formula of Страховой полис
    Sheet_Name_Here['Страховой полис'] = SUBSTITUTE(Sheet_Name_Here['Клиент'], LEFT(Sheet_Name_Here['Клиент'],
                                                                                    INT(FIND(Sheet_Name_Here['Клиент'],
                                                                                             'Номер ID карты:') + 15)),
                                                    '')

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Номенклатура': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'N Артикула прайса': 'Код услуги'}, inplace=True)

    # Deleted columns Клиент
    Sheet_Name_Here.drop(['Клиент'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Количество to dtype float
    Sheet_Name_Here['Количество'] = to_float_series(Sheet_Name_Here['Количество'])

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def medgit(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Фамилия, имя, отчество': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ Полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код  диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Гарантийное письмо': '№ ГП'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код услуги по прейскуранту': 'Код услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена услуги (руб.)': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns №п/п
    Sheet_Name_Here.drop(['№п/п'], axis=1, inplace=True)

    # Changed Количество to dtype float
    Sheet_Name_Here['Количество'] = to_float_series(Sheet_Name_Here['Количество'])

    # Deleted columns Сума к оплате (руб.)
    Sheet_Name_Here.drop(['Сума к оплате (руб.)'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def nmc_vash_doctor(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Направление\n(№ ID карты)': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'№ по прейску-ранту': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Лечение': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол – во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена, руб.': 'Цена услуги'}, inplace=True)

    # Deleted columns Стоимость, руб.
    Sheet_Name_Here.drop(['Стоимость, руб.'], axis=1, inplace=True)

    # Filled NaN values in 4 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Дата начала оказания услуги', 'ФИО пациента', 'Страховой полис', 'Код МКБ-10']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def stomatolog_kaluga(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=8)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Гарантийное письмо, №': '№ ГП'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата   лечения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Ф.И.О. врача': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба (стом)': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Changed Номер зуба (для стоматологических услуг) to dtype float
    Sheet_Name_Here['Номер зуба (для стоматологических услуг)'] = to_float_series(
        Sheet_Name_Here['Номер зуба (для стоматологических услуг)'])

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Сумма предъяв- ленная заказчику
    Sheet_Name_Here.drop(['Сумма предъяв- ленная заказчику'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Страховой полис', '№ ГП', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def clinica_pozvonochnika_aviakonstruktorov(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=4)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О\nпациента': 'ФИО пациента'}, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'№ Истории болезни': '№ истории болезни'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ \nСтрахового\nполиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз\nМКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата\n(период оказания услуг)': 'Дата начала оказания услуги'}, inplace=True)

    # Added column new-column-t4j6
    Sheet_Name_Here.insert(9, 'new-column-t4j6', 0)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'new-column-t4j6': 'Цена услуги'}, inplace=True)

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = Sheet_Name_Here['Цена услуги'].astype('float')

    # Set formula of Цена услуги
    Sheet_Name_Here['Цена услуги'] = SUBSTITUTE(Sheet_Name_Here['Стоимость'], ",00р.", "")

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)
    # Deleted columns №
    Sheet_Name_Here.drop(['№\nп/п'], axis=1, inplace=True)

    # Deleted columns Unnamed: 1
    Sheet_Name_Here.drop(['Unnamed: 1'], axis=1, inplace=True)

    # Deleted columns Unnamed: 11
    Sheet_Name_Here.drop(['Unnamed: 11'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', '№ истории болезни', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def fkv_clinic_primavera(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={2: 'Номер полиса',
                                    3: 'ФИО пациента',
                                    4: 'Дата начала оказания услуги',
                                    5: 'ФИО врача',
                                    6: 'Код услуги',
                                    7: 'Наименование услуги',
                                    8: 'Цена услуги',
                                    9: 'Кол-во',
                                    11: 'Код МКБ-10'}, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Кол-во'].notnull()) & (
        Sheet_Name_Here['Кол-во'].apply(lambda val: all(s not in str(val) for s in ['Кол-во', ' '])))]
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['ФИО пациента'] != '3']
    # Changed dtype
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    Sheet_Name_Here['Кол-во'] = to_int_series(Sheet_Name_Here['Кол-во'])
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def klinika_novavita(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ Полиса ДМС': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена услуги ': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={' МКБ10 ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Специальность врача
    Sheet_Name_Here.rename(columns={'Специальность врача оказавшего услугу': 'Специальность врача'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'номер гарантийного письма': '№ ГП'}, inplace=True)

    # Renamed columns № истории болезни
    Sheet_Name_Here.rename(columns={'Номер истории болезни': '№ истории болезни'}, inplace=True)

    # Added column new-column-myti
    Sheet_Name_Here.insert(4, 'new-column-myti', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-myti': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['Фамилия '], " ", Sheet_Name_Here['Имя'], " ",
                                             Sheet_Name_Here['Отчество'])

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Фамилия
    Sheet_Name_Here.drop(['Фамилия '], axis=1, inplace=True)

    # Deleted columns Имя
    Sheet_Name_Here.drop(['Имя'], axis=1, inplace=True)

    # Deleted columns Отчество
    Sheet_Name_Here.drop(['Отчество'], axis=1, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Цена по прайсу
    Sheet_Name_Here.drop(['Цена по прайсу'], axis=1, inplace=True)

    # Deleted columns Стоимость услуг
    Sheet_Name_Here.drop(['Стоимость услуг '], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def akvilio(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns № Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': '№ Страховой полис'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер и дата гарантийного письма': '№ ГП'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код  МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач выполневший работу': 'Врач (ФИО)'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)
    # Deleted columns Дата
    Sheet_Name_Here.drop(['Дата\nрождения'], axis=1, inplace=True)

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    # Filled NaN values in 4 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', '№ Страховой полис', '№ ГП', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def akvilio_nn(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Ф.И.О.': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Номер и дата гарантийного письма': '№ ГП'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код  МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач выполневший работу': 'Врач (ФИО)'}, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)
    # Deleted columns Дата
    Sheet_Name_Here.drop(['Дата\nрождения'], axis=1, inplace=True)

    # Deleted columns Сумма к оплате
    Sheet_Name_Here.drop(['Сумма к оплате'], axis=1, inplace=True)

    # Filled NaN values in 4 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', '№ ГП', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def nero_ortoped_centr(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column
    Sheet_Name_Here.insert(4, 'new-column-ord1', 0)
    # Renamed columns
    Sheet_Name_Here.rename(columns={'new-column-ord1': 'ФИО пациента',
                                    4: 'Страховой полис',
                                    5: 'Дата начала оказания услуги',
                                    6: 'Код услуги',
                                    7: 'Наименование услуги',
                                    8: 'Цена услуги',
                                    9: 'Количество',
                                    11: 'Код МКБ-10',
                                    12: '№ ГП',
                                    14: 'Врач (ФИО)'}, inplace=True)

    # Set formula
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here[1], " ", Sheet_Name_Here[2], " ", Sheet_Name_Here[3])
    # Filled NaN values
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[
        (Sheet_Name_Here['Количество'].notnull()) & (~Sheet_Name_Here['Количество'].str.contains('Кол-во', na=False))]
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'] != 9]
    # Deleted columns
    Sheet_Name_Here.drop([0, 1, 2, 3, 13], axis=1, inplace=True)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def kons_dia_poly(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-pahj
    Sheet_Name_Here.insert(4, 'new-column-pahj', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-pahj': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['2'], " ", Sheet_Name_Here['3'], " ", Sheet_Name_Here['4'])

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'5': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={7: 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'8': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={9: 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={10: 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={11: 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={13: 'Код МКБ-10'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={15: '№ ГП'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={17: 'Врач (ФИО)'}, inplace=True)

    # Deleted columns 1
    Sheet_Name_Here.drop([1], axis=1, inplace=True)

    # Deleted columns 2
    Sheet_Name_Here.drop(['2'], axis=1, inplace=True)

    # Deleted columns 3
    Sheet_Name_Here.drop(['3'], axis=1, inplace=True)

    # Deleted columns 4
    Sheet_Name_Here.drop(['4'], axis=1, inplace=True)

    # Deleted columns 12
    Sheet_Name_Here.drop([12], axis=1, inplace=True)

    # Deleted columns 16
    Sheet_Name_Here.drop([16], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def oftomolog_poly(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-chgq
    Sheet_Name_Here.insert(4, 'new-column-chgq', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-chgq': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['2'], " ", Sheet_Name_Here['3'], " ", Sheet_Name_Here['4'])

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'5': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={7: 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'8': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={9: 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={10: 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={11: 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={13: 'Код МКБ-10'}, inplace=True)

    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={15: '№ ГП'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={17: 'Врач (ФИО)'}, inplace=True)

    # Deleted columns 1
    Sheet_Name_Here.drop([1], axis=1, inplace=True)

    # Deleted columns 2
    Sheet_Name_Here.drop(['2'], axis=1, inplace=True)

    # Deleted columns 3
    Sheet_Name_Here.drop(['3'], axis=1, inplace=True)

    # Deleted columns 4
    Sheet_Name_Here.drop(['4'], axis=1, inplace=True)

    # Deleted columns 12
    Sheet_Name_Here.drop([12], axis=1, inplace=True)

    # Deleted columns 16
    Sheet_Name_Here.drop([16], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def mk_sova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Стоимость (руб)
    Sheet_Name_Here.drop(['Стоимость (руб)'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def mmk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'№ полиса': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена (руб)': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Стоимость (руб)
    Sheet_Name_Here.drop(['Стоимость (руб)'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def ldc_vitanova(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns № ГП
    Sheet_Name_Here.rename(columns={'Гарантийное письмо\n(номер, дата)': '№ ГП'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата обращения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование мед. услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена услуги\n(руб.)': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)
    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость\n(руб.)'], axis=1, inplace=True)

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', '№ ГП']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def stomat_center_nil_i_l_sterlitamakskiy(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер ID ': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def fgbu_nmic_endokrin_min_rf(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Полис': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата вып.': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'Код': 'Код услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Специалист': 'Врач (ФИО)'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол.': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Deleted columns Unnamed: 2
    Sheet_Name_Here.drop(['Unnamed: 2'], axis=1, inplace=True)

    # Deleted columns Unnamed: 3
    Sheet_Name_Here.drop(['Unnamed: 3'], axis=1, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Deleted columns Дата рожд.
    Sheet_Name_Here.drop(['Дата рожд.'], axis=1, inplace=True)

    # Deleted columns Unnamed: 6
    Sheet_Name_Here.drop(['Unnamed: 6'], axis=1, inplace=True)

    # Deleted columns Unnamed: 8
    Sheet_Name_Here.drop(['Unnamed: 8'], axis=1, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Deleted columns Unnamed: 13
    Sheet_Name_Here.drop(['Unnamed: 13'], axis=1, inplace=True)

    # Deleted columns Unnamed: 14
    Sheet_Name_Here.drop(['Unnamed: 14'], axis=1, inplace=True)

    # Deleted columns Unnamed: 15
    Sheet_Name_Here.drop(['Unnamed: 15'], axis=1, inplace=True)

    # Deleted columns Unnamed: 17
    Sheet_Name_Here.drop(['Unnamed: 17'], axis=1, inplace=True)

    # Deleted columns Коэф.
    Sheet_Name_Here.drop(['Коэф.'], axis=1, inplace=True)

    # Deleted columns Стоимость, руб
    Sheet_Name_Here.drop(['Стоимость, руб'], axis=1, inplace=True)

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Код услуги to dtype str
    Sheet_Name_Here['Код услуги'] = Sheet_Name_Here['Код услуги'].astype('str')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def slavutich(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[1]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-apfv
    Sheet_Name_Here.insert(4, 'new-column-apfv', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-apfv': 'ФИО пациента'}, inplace=True)

    # Added column new-column-glg3
    Sheet_Name_Here.insert(6, 'new-column-glg3', 0)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'new-column-glg3': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Наименование': 'Наименование услуги'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Скидка, %
    Sheet_Name_Here.rename(columns={'Скидка': 'Скидка, %'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here['Фамилия'], " ", Sheet_Name_Here['Имя'], " ",
                                             Sheet_Name_Here['Отчество'])

    # Set formula of Страховой полис
    Sheet_Name_Here['Страховой полис'] = SUBSTITUTE(Sheet_Name_Here['Серия и номер полиса'], "№", "")

    # Changed Страховой полис to dtype float
    Sheet_Name_Here['Страховой полис'] = to_float_series(Sheet_Name_Here['Страховой полис'])

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Фамилия
    Sheet_Name_Here.drop(['Фамилия'], axis=1, inplace=True)

    # Deleted columns Имя
    Sheet_Name_Here.drop(['Имя'], axis=1, inplace=True)

    # Deleted columns Отчество
    Sheet_Name_Here.drop(['Отчество'], axis=1, inplace=True)

    # Deleted columns Серия и номер полиса
    Sheet_Name_Here.drop(['Серия и номер полиса'], axis=1, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    # Deleted columns ID Прейскуранта
    Sheet_Name_Here.drop(['ID Прейскуранта'], axis=1, inplace=True)

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def sk_bolinet(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер индивидуальной карты (ID) Пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def doctor_krasnoyarsk(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги ': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Код филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Код филиала клиники': 'Код филиала клиники (точки)'}, inplace=True)

    # Renamed columns Наименование филиала клиники (точки)
    Sheet_Name_Here.rename(columns={'Наименование филиала клиники': 'Наименование филиала клиники (точки)'},
                           inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Врач ФИО': 'Врач (ФИО)'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Скидка, %
    Sheet_Name_Here.drop(['Скидка, %'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Filtered Количество
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Количество'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def karata_32(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'Номер индивидуальной карты (ID) Пациента': 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def arkhkom(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба': 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Коли-чество': 'Количество'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 4 columns in февраль_23_1
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги',
                           'Дата окончания оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def gbuz_mo_serpuhovskaya_rsp(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'номер ID': 'Страховой полис'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Страховой полис
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Страховой полис'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def semeyniy_stomat(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=6)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Filled NaN values in 3 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def nii_vitadent(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=2)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата лечения': 'Дата начала оказания услуги',
                                    'Код    услуги': 'Код услуги',
                                    'Полис                           ': 'Номер полиса',
                                    'Ф.И.О. врача': 'Врач (ФИО)',
                                    'Ф.И.О пациента': 'ФИО пациента',
                                    '№ зуба': 'Номер зуба (для стоматологических услуг)',
                                    'Объем  услуг': 'Количество'}, inplace=True)

    # Deleted columns Класс по Блеку
    Sheet_Name_Here.drop(['Класс по Блеку', 'Стоимость услуг', 'Всего', 'ИТОГО'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 4 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ФИО пациента', 'Дата начала оказания услуги', 'Врач (ФИО)',
                           'Номер полиса', 'Диагноз', 'Номер зуба (для стоматологических услуг)']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def clinic_cmd(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=12)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns 1
    Sheet_Name_Here.drop(['1'], axis=1, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'2': 'ФИО пациента'}, inplace=True)

    # Deleted columns 3
    Sheet_Name_Here.drop(['3'], axis=1, inplace=True)

    # Deleted columns 4
    Sheet_Name_Here.drop(['4'], axis=1, inplace=True)

    # Deleted columns 5
    Sheet_Name_Here.drop(['5'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'6': 'Страховой полис'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'7': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'8': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'9': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'10': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'11': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'12': 'Цена услуги'}, inplace=True)

    # Deleted columns 13
    Sheet_Name_Here.drop(['13'], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Filtered Цена услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Цена услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def med_luchevoy_center(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Тариф
    Sheet_Name_Here.drop(['Тариф'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Врач (ФИО)
    Sheet_Name_Here.rename(columns={'Доктор': 'Врач (ФИО)'}, inplace=True)

    # Changed Дата услуги to dtype datetime
    Sheet_Name_Here['Дата услуги'] = pd.to_datetime(Sheet_Name_Here['Дата услуги'], infer_datetime_format=True,
                                                      errors='coerce')

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Страховой полис
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Страховой полис'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def konsilium(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=5)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Deleted columns Дата рождения поциента
    Sheet_Name_Here.drop(['Дата рождения поциента'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуг': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код диагноза по МКБ-10': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'Название услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'Номер зуба (для стоматологии)': 'Номер зуба (для стоматологических услуг)'},
                           inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Сумма
    Sheet_Name_Here.drop(['Сумма '], axis=1, inplace=True)

    # Changed Дата начала оказания услуги to dtype datetime
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                                    infer_datetime_format=True, errors='coerce')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def tanmed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'номер ID': 'Страховой полис'}, inplace=True)

    # Deleted columns Дата рождения
    Sheet_Name_Here.drop(['Дата рождения'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата оказания услуги': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={'№ зуба (для стомат. услуг)': 'Номер зуба (для стоматологических услуг)'},
                           inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Deleted columns Стоимость со скидкой
    Sheet_Name_Here.drop(['Стоимость со скидкой'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def riat_spb(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата\nоказания': 'Дата начала оказания услуги',
                                    'Название\nуслуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент/полис': 'ФИО пациента'}, inplace=True)

    # Deleted columns Unnamed: 4
    Sheet_Name_Here.drop(['Unnamed: 4'], axis=1, inplace=True)

    # Deleted columns Unnamed: 6
    Sheet_Name_Here.drop(['Unnamed: 6'], axis=1, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Код по МКБ': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма': 'Цена услуги'}, inplace=True)

    # Deleted columns Unnamed: 10
    Sheet_Name_Here.drop(['Unnamed: 10'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')
    # Filtered Название
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filled NaN values in 2 columns in Sheet_Name_Here
    columns_to_fill_nan = ['Дата начала оказания услуги', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def ldc_mibs_tambov(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Deleted columns №
    Sheet_Name_Here.drop(['№'], axis=1, inplace=True)

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата выполнения': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Сумма, руб.': 'Цена услуги'}, inplace=True)

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def stomat_clinic_mba(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=9)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Added column new-column-36w0
    Sheet_Name_Here.insert(4, 'new-column-36w0', 0)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'new-column-36w0': 'ФИО пациента'}, inplace=True)

    # Set formula of ФИО пациента
    Sheet_Name_Here['ФИО пациента'] = CONCAT(Sheet_Name_Here[2], ' ', Sheet_Name_Here[3], ' ', Sheet_Name_Here[4])

    # Deleted columns 2, 3, 4
    Sheet_Name_Here.drop([2, 3, 4], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={1: 'Страховой полис'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={5: 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={6: 'Код МКБ-10'}, inplace=True)

    # Renamed columns Номер зуба (для стоматологических услуг)
    Sheet_Name_Here.rename(columns={7: 'Номер зуба (для стоматологических услуг)'}, inplace=True)

    # Deleted columns 8, 9, 10
    Sheet_Name_Here.drop([8, 9, 10], axis=1, inplace=True)

    # Deleted columns 11
    Sheet_Name_Here.drop([11], axis=1, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'8.1': 'Код услуги'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'9.1': 'Наименование услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'10.1': 'Количество'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'11.1': 'Цена услуги'}, inplace=True)

    # Deleted columns 12
    Sheet_Name_Here.drop([12], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 3 columns in Образец_1
    columns_to_fill_nan = ['Страховой полис', 'Дата начала оказания услуги', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    # Changed Цена услуги to dtype float
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)
def dentaservis(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=0, header=None)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns
    Sheet_Name_Here.rename(columns={0: 'Страховой полис',
                                    1: 'ФИО пациента',
                                    2: 'Дата начала оказания услуги',
                                    3: 'Наименование услуги',
                                    4: 'Код услуги',
                                    5: 'Код МКБ-10',
                                    6: 'Количество',
                                    7: 'Цена услуги',
                                    8: 'Врач (ФИО)'}, inplace=True)
    # Filtered
    Sheet_Name_Here = Sheet_Name_Here[(Sheet_Name_Here['Количество'].notnull()) & (~Sheet_Name_Here['Количество'].str.contains('Кол-во', na=False))]
    # Changed dtype
    Sheet_Name_Here['Количество'] = to_int_series(Sheet_Name_Here['Количество'])
    Sheet_Name_Here['Цена услуги'] = to_float_series(Sheet_Name_Here['Цена услуги'])
    Sheet_Name_Here['Дата начала оказания услуги'] = pd.to_datetime(Sheet_Name_Here['Дата начала оказания услуги'],
                                                          infer_datetime_format=True, errors='coerce')
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def check_collumn_name(func_list):
    test = func_list.columns  # важно название листа
    test = set(test)
    test_colums = set()
    for header_value in headers.items():
        for column in header_value:
            test_colums.add(column)
    i = test - test_colums
    print(i) # noqa


def template_qualifier(folder_name, printer=False):
    template_tuple = {
        'Акционерное Общество «Клиника К+31»': data_pandas_k_31,
            'АО "Национальный Медицинский Сервис"': data_national_medical_service,
            'ООО "Дальвен"': dalven,
            'ООО «Стоматология на Таганской»': data_pandas_stomat_tagansk,
            'ООО «Клиники Чайка»': chaika,
            'ООО "Стоматологическая поликлиника №9 Азино"': data_pandas_azino,
            'ООО Клиника "Креде эксперто"': data_pandas_kredo_experto,
            'ЧУЗ "ЦКБ "РЖД-медицина"': data_pandas_rzd,  # data_pandas_rzd_fix (Старый формат)
            'ООО «МЦ» XXI ВЕК»': vek_21,
            'ООО "СМАРТ КЛИНИК"': smart_clinic,
            'ООО «ГУТА-КЛИНИК»': guta_clinic,
            'ГБУЗ «ГКБ им. И.В. Давыдовского ДЗМ»': davidovskogo,
            'АО «Поликлиника «Медросконтракт»': medroscontract,
            'ООО «ЛиМ»': lim,
            'ООО «Клиника современных технологий «Садко»': sadoko_kst,
            'ООО «Центр развития стоматологии «Садко»': sadoko_crs,
            'ГБУЗ "ГКБ им. С.С. Юдина ДЗМ"': udina,
            'ООО «СФЕРА-МЕД»': sphera_med,
            'ГБУЗ "ГКБ им. А.К.Ерамишанцева ДЗМ"': gbuz_eramishanceva,
            'ООО "МедГард" (Самара, Тольятти)': medgard_samara_toliati,
            'ООО "МедГард" (Невский)': medgard_nevskii,
            'ООО "Медицина Альфастрахования"(Моска)': alfasrahovanie_moscow,
            'ООО «Медицина АльфаСтрахования г. Ростов-на-Дону»': alfasrahovanie_rostov_na_donu,
            'КГБУЗ КДЦ "Вивея"': vivea,
            'АО "Детская поликлиника Литфонда"': litfond,
            'ООО "Жемчужина Подолья"': zhemchuzhina_podolia,
            'ООО «Стоматология Стиль Дент»': stil_dent,
            'ООО «Клиника «Аллергомед»': allergomed,
            'Общество с ограниченной ответственностью Магазин недвижимости': magazin_nedvizhemosti,
            'ФГБУ "52 консультативно-диагностический центр" Министерства обороны РФ': fgbu_52_kdc_min_ob_rf,
            'УГМК': ugmk,
            'ООО «Эстетика»': estetika, # Ручное (частично)
            'ООО «Клиника ОсНова»': osnova,
            'ФГБУ «Клиническая больница» Управления делами Президента Российской Федерации': look_data,
            'ООО Центр современной медицины': center_sovremennoy_medecini,
            'ООО Сеть семейных медицинских центров': set_semeynih_med_centrov,
            'ФГБУ "НМИЦ ТПМ" МИНЗДРАВА РОССИИ': nmictpm_mz_rf,
            'АО "Семейный доктор"': look_data,
            '32 практика.xls': praktica_32,
            'Ависмед.xls': avismed,
            'ООО "МЦ Айболит"': aibolit,
            'ООО "Аллегро"': allegro,
            'ГБУЗ "ГКБ имени В.М Буянова ДЗМ"': buyanova,
            'ООО «МЦ «В Надежных Руках»': v_nadezhnih_rukah,
            'ООО «ФИРМА ВЕРОНИКА ЛТД»': veronika,
            'Витоника': vitonika,
            'Ворохобова': vorohobova,
            'Гк Медицина': gk_medecina,
            'ФГУП "ГлавУпДК" при МИД России': look_data,
            'ООО «ИНТАН» 1-Й ЦЕНТР ЗУБНОЙ ИМПЛАНТАЦИИ И СТОМАТОЛОГИИ»': intan,
            'АО "Ильинская больница"': ilinskaya_bolnica,
            'ООО "Здоровое поколение"': zdorovor_pokolenie,
            'ООО «Звезда»': zvezda,
            'ЗАО МЕДИ': zao_medi,
            'ООО "Гранти-мед"': granti_med,
            'Ермишанцева': ermishanceva,
            'ООО «ДЖИ ЭМ ЭС»': jms,
            'ООО Центр Стоматологии "Дентиформ"': dentiform,
            'ООО "Дентика"': dentika,
            'ООО "Камелия-Мед"': kamelia_med,
            'ООО "Стоматологическая поликлиника "Лазурь"': lazur,
            'Медас Самара': medas_samara,
            'Медас Пермь': medas_perm,
            'ФГБОУ ВО "Московский государственный университет им.М.В.Ломоносова"обособленное подразделение МНОЦ МГУ': mgu_lomonosova,
            'ООО "Немецкая клиника"': nemeckaya_clinika,
            'ООО Наука': nauka,
            'ООО МРТ24': data_pandas_24mrt,
            'ООО Медтим': medtim,
            'ООО «МЕДИЦИНСКИЙ ЦЕНТР АТЛАС»': medicinskiy_cent_atlas,
            'ООО «Профессиональная медицинская лига»': professionalnaya_medicinskaya_liga,
            'ФГБУ "НМИЦ ТО им Н.Н. Приорова" Минздрава России': pirogova,
            'АО «Поликлинический Комплекс»': policlinicheskiy_komplex,
            'ООО «Семейный Медицинский Центр »': semeiniy_med_centr,
            'ООО «Сити-Клиник»': city_clinik,
            'ООО СК "ТЕСТ"': sk_test,
            'ООО СК "Юнит"': sk_unit,
            'Общество с ограниченной ответственностью "СОН-МЕД"': son_med,
            'Стоматологическая поликлиника 9 Казань': stomat_poli_9_kazan,
            'ООО «СуперМедик»': super_medic,
            'ФГБУ ФНКЦ ФМБА России': fnkc_fmba,
            'ООО "Формула улыбки"': formula_ulibki,
            'ООО «Форум Интернэшнл Технолоджи»': forum_internationla_technolog,
            'ООО «Центр Семейной Медицины»': centr_semeynoy_med,
            'ФГБУ «ЦКБ с поликлиникой»': fgbu_ckb_s_poliklinikoy,
            'ЦНИИСиЧЛХ': look_data,
            'ООО "ЦМРТ Новослободская"': crmt_Novoslobotskaya,
            'ООО «ЦМРТ Сокольники»': crmt_Novoslobotskaya,
            'ГБУЗ НИКИО им.Л.И.Свержевского ДЗМ': gbuz_nikio_im_l_i_sverzhevckogo,
            'ОАО «Фармация»': farmaciya,
            'ОАО «Фармация» (Эскулап)': 'Ручное',
            'ООО "АПЕКС"': apex,
            'ООО "Бека-Инвест"': beka_invest,
            'ООО "ДЕНС"': dems,
            'ООО "КДС"': kds,
            'ООО "Клиника Доброго Стоматолога"': klinika_dobrogo_stomatologa,
            'ООО "МедГард"': medgard,
            'ООО "Медицинские Центры - 2"': medecinskie_centri_2,
            'ООО "Многопрофильный медицинский центр "Диалайн"': mnogoprofilnii_medecinskii_center_dialain,
            'ООО "Одинмед"': odinmed,
            'ООО "Поликлиника № 2 Вита Медикус"': poliklinika_2_vita_medicus,
            'ООО "РМА"': rma,
            'ООО "СОГАЗ-Медсервис"': sogaz_medservis,
            'ООО "Спорткомплекс "Олимпия-Пермь"': stomatkomplex_olimpia_perm,
            'ООО "Стоматбизнес Компани"': stomatbiznes_company,
            'ООО "Медики"': stomatbiznes_company,
            'ООО "Стомед"': stomed,
            'ООО "Династия"': dinastia,
            'ООО "ЭФА"': efa,
            'ООО «Боткинская 33»': botkinskaya_33,
            'ООО «ВАЛЕ-ДЕНТАЛЬ»': vale_dental,
            'ООО «Гутен Таг»': guten_tag,
            'ООО «Доктор рядом»': doctor_ryadom,
            'ООО «Европейский медицинский центр «УГМК-Здоровье»': evropeiskiy_medecinskiy_center_egmk_zdorovie,
            'ООО «Клиника немецкой стоматологии «Гутен Таг»': klinika_nemeckoy_stomatologii_guten_tag,
            'ООО «Клиника санитас в медпарке»': klinika_sanitas_v_medparke,
            'ООО «Медицина АльфаСтрахования г. Пермь»': medecina_alfastrahovaniya_perm,
            'ООО «Медицина АльфаСтрахования г. Самара»': medecina_alfastrahovaniya_samara,
            'ООО «Медицина АльфаСтрахования г. Ярославль»': medecina_alfastrahovaniya_yaroslavl,
            'ООО «Мир улыбок»': mir_ulibok,
            'ООО «Наш МЦ «Парацельс»': nash_mc_paracels,
            'ООО «Панацея»': panaceya,
            'ООО «Первый Доктор»': perviy_doctor,
            'ООО «Призвание»': prizvanie,
            'ООО «Приоритет»': prioritet,
            'ООО «Стомат. студия «Ваш Доктор»': stomat_studia_vash_doctor,
            'ООО «Стоматологическая поликлиника «Денто-Смайл»': stomatologicheskay_poliklinika_dento_smail,
            'ООО «Стоматологическая поликлиника ВИЗАВИ»': stomatologicheskay_poliklinika_vizavi,
            'ООО «Стоматоша»': stomatoshka,
            'ООО «Стоматологическая поликлиника №9 города Казани»': stomatologicheskay_poliklinika_9_goroda_kazani,
            'ООО «Стоматологический центр»': stomatologicheskiy_center,
            'АО "Самарский диагностический центр"': samarskiy_dia_centr,
            'ГКБ №1 им. "Н.И. Пирогова"': gkb_nomer_1_n_i_pirogova,
            'ООО "Витаника"': vitanika,
            'ООО "Даймонд клиник"': diamond,
            'ООО "Поликлиника консультативно-диагностическая им. Е.М. Нигинского"': poliklinaka_konsultativno_diagnosticheskaya_im_e_m_nigibskogo,
            'ООО "ЦСВМП"': csvmp,
            'ООО "поликлиника ЛАУС ДЕО"': laus_deo,
            'ООО «Блеск» (на Геодезической)': blesk_na_geodezicheskoi, # Нужен координатный морфер
            'ООО «СЦ Приор-М»': cs_prior_m,
            'ООО «Тонус+»': tonus_plus,
            'ООО «Тонус»': tonus,
            'ООО «ЦС «32 Практика»': cs_32_praktika,
            'ООО «Центр лучевой диагностики «ТОНУС ПРЕМИУМ»': centr_luchevoi_diagnostiki_tonus_premium,
            'ООО «ЭРАМед»': eramed,
            'ООО "Эра-1"': era_1,
            'ООО ЛДЦ "Казанская клиника"': ldc_kazanskaya_klinka,
            'ООО МЦ «Лотос»': lotos,
            'ФГБУ "Поликлиника № 1" УДП РФ': fgbu_policlinika_nomer_1_uod_rf,
            'ФГБУ НМИЦ "ЦНИИСиЧЛХ" Минздрава России': fgbu_nmic_cniisichlh_min_rf,
            'ЧУЗ «КБ «РЖД-МЕДИЦИНА» г. Санкт-Петербург»': rzd_spb,
            'ФГАОУ ВО Первый МГМУ им. И.М. Сеченова Минздрава России (Сеченовский Университет)': fgbu_bo_mgu_im_i_m_sechenova_mizdrava_rf,
            'ГБУЗ "Краевая клиническая больница №2"': gbuz_kkb_2,
            'ФГБУ «ОБП»': fgbu_obp,
            'ФГБУ «Клиническая больница №1"': fgb_klinicheskaya_bonica_nomer_1,
            'АО «Современные медицинские технологии»': policlinicheskiy_komplex,
            'ООО "ДАЛИЗ"': daliz,
            'ФГБУ НМИЦ онкологии им. Н.Н. Блохина Минздрава России': fgbu_nmic_onkologii_im_n_n_blohina_min_rf,
            'ООО "Парацельс" (г. Краснодар)': paracels_krasnodar,
            'ООО «Наш доктор» (п. Мехзавод)': nash_doctor_p_mehzavod,
            'ООО Альфа Мед Плюс': alfa_med,
            'ООО Альфа Мед': alfa_med,
            'ООО МаксиДент': alfa_med,
            'ООО МаксиДент+': alfa_med,
            'ГБУЗ СО ССП № 3': gbuz_so_ssp_3,
            'ООО Дент-Реал': dent_real,
            'ООО Дент-Реал Плюс': dent_real_plus,
            'ООО «МЕДИЦИНА БУДУЩЕГО»': medicina_buduschego,
            'ГБУЗ ССП № 2': gbuz_ssp_2,
            'Санкт-Петербургское государственное унитарное предприятие  «Петербургский метрополитен»': spb_gup_peterburskiy_metropoliten,
            'ООО МФЦ «Гармония»': mfc_garmoniya,
            'ООО «Медицина АльфаСтрахования г. Мурманск»': medecina_alfastrahovaniya_murmansk,
            'ООО «Медицина АльфаСтрахования г. Тюмень»': medecina_alfastrahovaniya_tumen,
            'АНМО «Ставропольский краевой клинический консультативно-диагностический центр»': stavropolskiy_kraevoy_klinicheskiy_kons_diagnos_center,
            'ГОБУЗ «Мурманский областной клинический многопрофильный центр»': gobuz_murmanskiy_olastnoy_klinicheskiy_mnogoprof_center,
            'ГБУЗ МКНЦ имени А.С. Логинова ДЗМ': gbuz_moskvi_mknpc_im_a_s_lohinova_dzgm,
            'ФГАУ НМИЦ МНТК «Микрохирургия глаза» им. акад. С.Н. Фёдорова» Минздрава России г. Краснодар': fgau_nmic_mntk_mg_im_akad_s_n_fedorova_min_rf_krasnodar,
            'ФБУЗ ПОМЦ ФМБА России': fbuz_pomc_fmba_rossii,
            ' ГБУЗ СО "Самарская городская поликлиника № 6 Промышленного района"': gbuz_so_samarskaya_gp_6_prom_r,
            'АО «К 31 Сити»': ao_k31_city,
            'ГБУЗ города Москвы «Городская клиническая больница № 15 имени О.М. Филатова ДЗМ»': gbuz_moskvi_gkb_15_im_o_m_filatova_dzm,
            'ООО "Эмаль"': emal,
            'ООО "Юни Медика"': uni_medica,
            'ООО Надежда': nadezhda,
            'ООО «ВитаСмайл»': vita_smile,
            'ООО НМИЦ МЕДИКА МЕНТЕ': nmic_medica_mente,
            'ООО ВДЦ - медицинский центр': vdc_med_center,
            'ООО МЦ «Наедине-Н»': mc_naedine_n,
            'ООО АЛДЕНТ': aldent,
            'ООО ССК ЗУБАСТИК': ssk_zubastik,
            'ООО ОММЦ им. св.Луки': ommc_im_cv_luki,
            'ООО Норма ХХI': norma_xxi,
            'ООО Ной': noy,
            'ООО Стоматолог и Я': stomatolog_i_ya,
            'ООО Медосмотр': medosmotr,
            'ИП Коротчик Ю.О.': ip_korotchik_u_o,
            'ООО Улыбка Плюс': ulibka_plus,
            'А-Стом': a_stom,
            'ООО КЛИНИКА СЕМЕЙНОЙ МЕДИЦИНЫ': klinika_semeynoy_medecini,
            'БУ «Республиканский клинический госпиталь для ветеранов войн» Минздрава Чувашии': bu_rlg_dvv_min_chuvash,
            'ООО НПО Волгоградский центр профилактики болезней ЮгМед': npo_vcpb_ugmed,
            'ООО ЮгМед-М': ugmed_minus_m,
            'ЧУЗ «КБ «РЖД-Медицина» им. Н.А. Семашко»': rzd_semashko,
            'ООО "Авиценна"': avicenna,
            'ООО МЦ «ЦУП МедПрофи»': mc_cup_medprofi,
            'ООО Клиника «Персона»': clinika_persona,
            'ООО «Ивстройтех»': ivstroiteh,
            'ООО «Здоровые люди»': zdorovie_ludi,
            'ООО «Совермед»': sovermed,
            'ООО "ЛДЦ МИБС-Чебоксары"': ldc_mibs_cheboksari,
            'ООО «ММЦ «УРО-ПРО»': mmc_urp_pro,
            'ООО Гиппократ': gippokrat,
            'ООО «РУЗА ДЕНТ»': razu_dent,
            'ООО ЛПЦ "ДентЛиния"': lpc_dent_liniya,
            'ООО «Стоматолог-2»': stomatologiya_2,
            'ООО "КЕЛЛЕР СТАЧКИ"': keller_stachki,
            'ФБУЗ «Медико-санитарная часть № 9» ФМБА России г. Дубна': fbuz_msch_9_fbma_r_dubna,
            'ООО «МЦ на Кырли»': mc_kirli,
            'ООО "Стоматологический Центр "ВладМиВа"': stomat_center_vladmiva,
            'ООО «Национальный Диагностический Центр»': nac_dia_centr,
            'ООО "Наша Клиника"': nasha_klinika,
            'ООО КЛИНИКА СОВРЕМЕННОЙ МЕДИЦИНЫ г. Набережные Челны': klinika_sov_med_nab_chelni,
            'СПб ГБУЗ «ГКДЦ №1»': spb_gbuz_gkdc_1,
            'Частное учреждение здравоохранения «Поликлиника Овум»': chuz_poli_ovum,
            'ГАУЗ «РКОД МЗ РТ им. проф. М.З.Сигала»': gauz_rkod_mz_rt_im_m_z_sigala,
            'ГАУЗ «ЦГКБ №18»': gauz_cgrb_18,
            'ООО «Прайм-Стоматология»': prime_stomatology,
            'ООО «Дентерпрайз»': denterprice,
            'ГБУЗ ЛО "Киришская КМБ"': gbuz_lo_kirishskaya_kmb,
            'ООО "Он Клиник Рязань"': on_clinic_ryazan,
            'ООО «Клиника современной стоматологии «АРТ С»': clinica_sov_stom_art_c,
            'ООО «ЛДЦ Разумед»': ldc_razumed,
            'ООО «Сеть семейных медицинских центров Регион №2» (Рязань)': ssmc_region_2_ryazan,
            'ООО «Сеть семейных медицинских центров Регион №2» (Тула)': ssmc_region_2_tula,
            'ООО «Клиника «Уральская»': clinic_uralskaya,
            'ООО  Центр профилактической медицины Ультрамед': centr_prof_med_ultramed,
            'ООО «Мед-Арт» г. Томск': med_art_tomsk,
            'ООО Стоматологическая клиника Карат': stomatology_clinic_karat,
            'ГАУЗ НО "Городская клиническая поликлиника № 1"': gauz_no_gcp_1,
            'ФГАУ "НМИЦ нейрохирургии им. ак. Н.Н. Бурденко" Минздрава России': fgau_nmic_nero_im_burdenko_min_rf,
            'ООО Медицинский центр «Надежда»': med_centr_nadezhda,
            'ООО "Дент Арт" (Новокузнецк)': dent_art_novokuzneck,
            'ГАУЗ СО «ГКБ №40»': gbauz_so_gkb_40,
            'АО «Камский диагностический центр «Медикам»': ao_kamskiy_dia_cent_medikam,
            'ООО «Клиника ДНК»': clinic_dnk,
            'ООО «Мега-Центр»': mega_cent,
            'ООО «Охта Дентал плюс»': ohta_dental_plus,
            'ООО «Радуга-Мед»': raduga_med,
            'ООО «Сана Ко»': sana_ko,
            'ООО «Аня-Рыбинск»': anay_ribinsk,
            'ООО ДЦ «Клиника-Сити»': dc_clinic_city,
            'ООО "Мой зубной"': moy_zubnoi,
            'ООО «МЕДИН»': medin,
            'ООО «НЬЮ ДЕНТ» г. Тольятти': new_dent_toliati,
            'ЧУЗ Клиническая поликлиника РЖД-Медицина г. Архангельск': chuz_rzd_med_arhangelsk,
            'ООО «Поликлиника №1 Вита Медикус»': poly_1_vita_medikus,
            'ООО Медицинский центр "СЭЛ"': med_centr_sel,
            'ООО Медицинский центр «Столица»': med_cent_stolica,
            'ООО МЛДЦ Диагност': mlcd_diagnost,
            'ГАУЗ «Городская клиническая больница №7» г. Казани': gauz_gkb_7_kazani,
            'ООО МЦ «МЕДЕОР»': mc_medeor,
            'ООО «МедГид»': medgit,
            'ООО «НМЦ «Ваш Доктор»': nmc_vash_doctor,
            'ООО «Стоматолог» г. Калуга': stomatolog_kaluga,
            'ООО Клиника Позвоночника на Авиаконструкторов': clinica_pozvonochnika_aviakonstruktorov,
            'ООО «ФКВ клиника «Примавера»': fkv_clinic_primavera,
            'ООО «Клиника НоваВита»': klinika_novavita,
            'ООО Аквилио': akvilio,
            'ООО Аквилио-НН': akvilio_nn,
            'ООО «Нейро-ортопедический центр»': nero_ortoped_centr,
            'ООО Консультативно-диагностическая поликлиника': kons_dia_poly,
            'ООО Офтальмологическая поликлиника': oftomolog_poly,
            'ООО «МК «СОВА»': mk_sova,
            'ООО «ММК»': mmk,
            'ООО «ЛДК ВитаНова»': ldc_vitanova,
            'ООО «Стоматологический центр - Ниль и К «Стерлитамакский»': stomat_center_nil_i_l_sterlitamakskiy,
            'ФГБУ НМИЦ эндокринологии Минздрава России': fgbu_nmic_endokrin_min_rf,
            'ООО Славутич': slavutich,
            'ООО СК «Болинет»': sk_bolinet,
            'ООО Доктор г. Красноярск': doctor_krasnoyarsk,
            'ООО «32 карата»': karata_32,
            'ООО "Архком"': arkhkom,
            'ГБУЗ МО «Серпуховская районная стоматологическая поликлиника»': gbuz_mo_serpuhovskaya_rsp,
            'ООО "СЕМЕЙНЫЙ СТОМАТОЛОГ"': semeyniy_stomat,
            'ООО НИИ "Витадент"': nii_vitadent,
            'ООО «Клиника ЦМД»': clinic_cmd,
            'ООО «Медицинский лучевой центр»': med_luchevoy_center,
            'ООО «Консилиум»': konsilium,
            'ООО "ТАНМЕД"': tanmed,
            'ООО «РИАТ СПб»': riat_spb,
            'ООО «ЛДЦ МИБС – Тамбов»': ldc_mibs_tambov,
            'ООО "Стоматологическая клиника МВА"': stomat_clinic_mba,
            'ООО "Дентасервис"': dentaservis,
    }
    if printer:
        summ = 0
        for i in template_tuple.keys():  # noqa: B007
            summ += 1
        return print(summ)  # noqa: T201
    if folder_name in template_tuple:
        sample = template_tuple[folder_name]
    else:
        sample = 'clinic_name'
    return sample


def get_key(dictionary, dict_value):
    for dict_key, dict_value in dictionary.items():
        if len(dict_value) > 0:
            if dict_value[0] == dict_value:
                return dict_key


def look_data(local_path, file_path):  # noqa
    wb = load_workbook(local_path, data_only=True)
    done_2 = False
    skip_row_list = []
    min_row = None
    max_row = None
    sheet_number = None
    for i in range(0, len(wb.sheetnames)):
        sheet_name = wb.sheetnames[i]
        sheet = wb[sheet_name]
        for column_number in range(1, sheet.max_column):
            done = False
            if done_2:
                break
            for row_number in range(1, (sheet.max_row + 2)):
                inform = sheet.cell(row=row_number, column=column_number)
                if not done:
                    for key in headers:
                        if inform.value in headers[key]:
                            min_row = row_number
                            sheet_number = i
                            done = True
                            break
                else:
                    if inform.value is None:
                        try:
                            sheet_df_dictonary = pd.read_excel(local_path, engine='openpyxl', sheet_name=[sheet_name],  # noqa: E501
                                                               skiprows=min_row - 1)
                            Sheet_Name_Here = sheet_df_dictonary[sheet_name]
                            check_collumn_name(Sheet_Name_Here)
                        except:  # noqa
                            pass
                        max_row = row_number - 1
                        done_2 = True
                        break
                    else:
                        continue
    return check_data(local_path, min_row, max_row, sheet_number, file_path, skip_row_list)


def check_data(local_path, min_row, max_row, sheet_number, file_path, skip_row_list=None): # noqa:  CCR001
    global key_name, inform_value
    data_list = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    wb = load_workbook(local_path, data_only=True)
    sheet_name = wb.sheetnames[sheet_number]
    sheet = wb[sheet_name]
    for column_number in range(1, sheet.max_column + 1):
        done = False
        for row_number in range(min_row, max_row + 1):
            if skip_row_list is not None:
                if row_number in skip_row_list:
                    continue
            inform = sheet.cell(row=row_number, column=column_number)
            inform_value = inform.value
            if done:
                if inform.value is not None:
                    data_list[key_name].append(str(inform_value))
                else:
                    data_list[key_name].append(None)
            if not done:
                for key in headers:
                    if inform.value in headers[key]:
                        key_name = key
                        done = True
                        break
    try:
        return clear_data(data_list, local_path, file_path)
    except:  # noqa:
        print('Смена формата не прошла', file_path)  # noqa: T201
        return filling_book(data_list, local_path, file_path)


def clear_data(data_list, local_path, file_path):  # noqa
    try:
        data_list['policy_number'] = [int(re.sub('^0{2}|^0{3}|^0{4}|^0{5}', '', x)) if x is not None else x for x in  # noqa: E501
                                      data_list['policy_number']]
    except:  # noqa
        try:
            data_list['policy_number'] = [x for x in data_list['policy_number']]  # noqa
        except:  # noqa
            print('policy_number с ошибкой\n', data_list['policy_number'])  # noqa
            pass

    try:
        data_list['date'] = [
            datetime.strptime(x, '%Y-%m-%d %H:%M:%S').date().strftime('%d.%m.%Y') if x is not None else x for x in # noqa:  E501
            data_list['date']]
    except:  # noqa: E722, B001
        try:
            data_list['date'] = [datetime.strptime(x, '%Y-%m-%d').date().strftime('%d.%m.%Y') if x is not None else x # noqa:  E501
                                 for x in data_list['date']]
        except:  # noqa: E722, B001
            try:
                data_list['date'] = [
                    datetime.strptime(x, '%d-%m-%Y').date().strftime('%d.%m.%Y') if x is not None else x for x in # noqa:  E501
                    data_list['date']]
            except:  # noqa: E722, B001
                try:
                    for x, y in enumerate(data_list['date']):  # noqa: VNE001
                        if y is None:
                            continue
                        try:
                            data_list['date'][x] = datetime.strptime(y, '%Y-%m-%d %H:%M:%S').date().strftime('%d.%m.%Y') # noqa:  E501
                        except:  # noqa: E722, B001
                            try:
                                data_list['date'][x] = datetime.strptime(y, '%d.%m.%Y').date().strftime('%d.%m.%Y') # noqa:  E501
                            except:  # noqa: E722, B001
                                continue
                except:  # noqa: E722, B001
                    try:
                        data_list['date'] = [x.split(' ', 2)[0] for x in data_list['date']]
                    except:  # noqa: E722, B001
                        try:
                            data_list['date'] = [x.split(' ', 2)[0].replace('-', '.').split('.', -1)[-1] + '.' + x.split(' ', 2)[0].replace('-', '.').split('.', -1)[-2] + '.' + x.split(' ', 2)[0].replace('-', '.').split('.', -1)[0] for x in data_list['date']]  # noqa: E501
                        except:  # noqa: E722, B001
                            print('date с ошибкой\n', data_list['date'])  # noqa: T201
    try:
        data_list['service_price'] = [float(x) for x in data_list['service_price']]
    except:  # noqa: E722, B001
        try:
            data_list['service_price'] = [float(x.replace('.', ',')) for x in data_list['service_price']]  # noqa:  E501
        except:  # noqa: E722, B001
            pass
    try:
        data_list['service_amount'] = [int(x) for x in data_list['service_amount']]
    except:  # noqa: E722, B001
        pass
    try:
        data_list['tooth_number'] = [x if x is not None else x for x in data_list['tooth_number']]
    except:  # noqa: E722, B001
        print('tooth_number с ошибкой\n', data_list['tooth_number'])  # noqa: T201
        pass
    return filling_book(data_list, local_path, file_path)


def filling_book(data_list, file_name, file_path):  # тип оплаты факт  # noqa: CCR001, CFQ001, C901
    global maximum_len, value, list_data_2, list_data_3, temp_type
    month_list = ('Ноябрь', 'октябрь', 'Июнь', 'ноябрь', 'июль', '10', 'сентябрь', '11', 'Октябрь', 'май', 'март',  # noqa:  E501
                  'август')
    wb = Workbook()
    ws = wb.active
    ws.append(shablon)
    for month in month_list:
        if month in file_path:
            new_file_folder = file_path.split('/', -1)[-3].replace('"', '').replace("""'""", '')
        else:
            new_file_folder = file_path.split('/', -1)[-2].replace('"', '').replace("""'""", '')
    new_file_name = str(new_file_folder + '_' + file_name.replace('.xlsx', '') + '_' + 'morphed.xlsx')  # noqa:  E501
    if len(data_list['fio']) == 0:
        maximum_len = len(data_list['last_name'])
    else:
        maximum_len = len(data_list['fio'])
    for column_number in range(1, (len(shablon) + 1)):
        for row_number in range(2, maximum_len + 2):
            car = get_column_letter(column_number)
            value = get_key(LOCAL_HEADERS, shablon[column_number - 1])
            if column_number == 24:
                ws[car + str(row_number)] = 'fact'
            if value == 'fio' and len(data_list['fio']) == 0:
                ws[car + str(row_number)] = data_list['last_name'][row_number - 2] + ' ' + data_list['first_name'][row_number - 2] + ' ' + data_list['middle_name'][row_number - 2]  # noqa:  E501
                continue   # noqa:  E126
            elif value == 'doctor_name' and len(data_list['doctor_name']) == 0 and (len(data_list['doctor_last_name']) + len(data_list['doctor_first_name']) + len(data_list['doctor_middle_name'])) > 0:  # noqa:  E501
                ws[car + str(row_number)] = data_list['doctor_last_name'][row_number - 2] + ' ' + data_list['doctor_first_name'][row_number - 2] + ' ' + data_list['doctor_middle_name'][row_number - 2]  # noqa:  E501
                continue  # noqa:  E126
            elif value == 'service_price' and len(data_list['total_price']) > 0:
                ws[car + str(row_number)] = data_list['total_price'][row_number - 2]
            elif value is None or data_list[value] is None or len(data_list[value]) < 1:
                continue  # noqa:  E126
            else:
                ws[car + str(row_number)] = data_list[value][row_number - 2]
    wb.save(new_file_name)
    try:
        ya_disk.remove(f'{file_name}')
    except:  # noqa: E722, B001
        pass
    new_file_path = file_path.replace(file_name, '') + new_file_name
    try:
        ya_disk.upload(f'{new_file_name}', new_file_path)
        print('success')  # noqa: T201
    except:  # noqa: E722, B001
        pass


def sql_to_base(data_list):  # noqa: CCR001, C901
    INSERT_SQL = {0: '''insert into one_time_data_load_for_merge.pp_all_date ("ФИО пациента", "Страховой полис", "Дата начала оказания услуги", "Дата окончания оказания услуги", "Код услуги", "Наименование услуги", "Код МКБ-10", "Диагноз", "Номер зуба (для стоматологических услуг)", "Цена услуги", "Количество", "Скидка, %", "Код филиала клиники (точки)", "Наименование филиала клиники (точки)", "№ ГП", "Код врача", "Врач (ФИО)", "Специальность врача", "Специальность направившего врача", "Код отделения", "Наименование отделения", "№ истории болезни", "ID Прейскуранта", "Тип оплаты", "Тип обслуживания / Вид помощи", "Медицинская услуга") values ''', 1: '''(''', 2: ''')'''}  # noqa: Q001, E501
    values = ''
    sample = ('fio', 'policy_number', 'date', 'end_date', 'service_code', 'service_name', 'mkb', 'diagnosis',  # noqa:  E501
              'tooth_number', 'service_price', 'service_amount', 'discount_sice', 'clinic_code', 'clinic_name',  # noqa:  E501
              'guarantee_letter', 'doctor_code', 'doctor_name', 'doctor_speciality', 'doctor_speciality_2',  # noqa:  E501
              'branch_code', 'branch_name', 'number_disease_history', 'price-list_id', 'service_type',  # noqa:  E501
              'service_type', 'medical_service')
    for cycle_number in range(0, len(data_list['service_amount'])):
        value_list = list()  # noqa:  C408
        for data_value in sample:
            if len(data_list[data_value]) == 0:
                value_list.append(None)
            else:
                if data_value == 'fio' and len(data_list[data_value]) == 0:
                    fio = data_list['last_name'][cycle_number] + data_list['first_name'][cycle_number] + data_list['middle_name'][cycle_number]  # noqa:  E501
                    value_list.append(fio)
                    continue  # noqa:  E126
                elif data_value == 'doctor_name' and len(data_list[data_value]) == 0:
                    doctor_fio = data_list['doctor_last_name'][cycle_number] + data_list['doctor_first_name'][  # noqa:  E501
                        cycle_number] + data_list['doctor_middle_name'][cycle_number]
                    value_list.append(doctor_fio)
                    continue  # noqa:  E126
                elif (data_value == 'date' or data_value == 'end_date') and data_list[data_value][
                    cycle_number] is not None:  # noqa: E125
                    if isinstance(data_list[data_value][cycle_number], str):
                        value_list.append('to_date(' + str(data_list[data_value][cycle_number]) + ', dd.mm.yyyy)')  # noqa:  E501
                    else:
                        value_list.append(data_list[data_value][cycle_number])
                    continue  # noqa:  E126
                else:
                    try:
                        value_list.append(data_list[data_value][cycle_number])
                    except:  # noqa: E722, B001
                        value_list.append(None)
        values += INSERT_SQL[1] + str(value_list).replace('[', '').replace(']', '') + INSERT_SQL[2] + ','  # noqa: E501
    sql = INSERT_SQL[0] + values[0:-1].replace('None', 'Null').replace("""'to_date(""", """to_date('""").replace(  # noqa:  E501
        """, dd.mm.yyyy')'""", """', 'dd.mm.yyyy')""")
    return sql


def change_format(file_path):  # noqa: CCR001
    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
        """Change the file format of an excel file from xls or xlsb to xlsx"""
        _, file_extension = os.path.splitext(file_path)
        if file_extension not in ('.xls', '.xlsb'):
            return file_path
        wb = load_workbook(filename=file_path)
        new_wb = Workbook()
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            new_sheet = new_wb.create_sheet(title=sheet_name)
            for row in sheet:
                new_row = []
                for cell in row:
                    new_row.append(cell.value)
                new_sheet.append(new_row)
        new_file_path = os.path.splitext(file_path)[0] + '.xlsx'
        new_wb.save(new_file_path)
        os.remove(file_path)
        return new_file_path
    else:
        print('Файл не найден или пустой')  # noqa: T201


def start_with_query(files_path={'path': [], 'file_name': [], 'folder_name': []}):  # noqa: CCR001, B006, E501
    DWH = DwsConn(conn_string=gp_conn)
    list_of_dirs = DWH.select('select name,path from yandex_disk.clinics_files_test as clsgl where status is null;')  # noqa
    for data_row in list_of_dirs:
        files_path['path'].append(data_row[1])
        files_path['file_name'].append(data_row[0])
        files_path['folder_name'].append(data_row[1].split('/', -1)[3])
    return files_path



def main_loop():  # noqa:  CCR001, CFQ001, C901
    file_folder_list = start_with_query()
    total_result = {'Успешно': [],
                    'Сломался на этапе смены формата': [],
                    'Сломана функция': [],
                    'Отсутсвует': [],
                    'Другое': [],
                    'Ручное': [],
                    }
    for number, reg_file in enumerate(file_folder_list['path']):
        if 'morphed' in file_folder_list['file_name'][number]:
            continue  # noqa:  E126
        if 'test' not in file_folder_list['folder_name'][number]:
            print(file_folder_list['file_name'][number])  # noqa: T201
            try:
                ya_disk.download(reg_file, file_folder_list['file_name'][number])
            except:  # noqa: E722, B001
                pass
            clean_file = template_qualifier(file_folder_list['folder_name'][number])
            try:
                if 'clinic_name' in clean_file:
                    total_result['Отсутсвует'].append(file_folder_list['path'][number])
                    continue  # noqa:  E126
                elif 'Ручное' in clean_file:
                    total_result['Ручное'].append(file_folder_list['path'][number])
                    continue  # noqa:  E126
            except:  # noqa: E722, B001
                pass
            try:
                file_folder_list['file_name'][number] = change_format(file_folder_list['file_name'][number])  # noqa:  E501
            except:  # noqa: E722, B001
                try:
                    if file_folder_list['file_name'][number].split('.', -1)[-1].lower() == 'xls':
                        new_name = file_folder_list['file_name'][number].lower() + 'x'
                        p.save_book_as(file_name=file_folder_list['file_name'][number], dest_file_name=new_name)  # noqa:  E501
                        os.remove(file_folder_list['file_name'][number])
                        file_folder_list['file_name'][number] = new_name
                    elif file_folder_list['file_name'][number].split('.', -1)[-1].lower() == 'xlsb':
                        new_name = file_folder_list['file_name'][number] = file_folder_list['file_name'][number].split('.', -1)[0] + '.xlsx'    # noqa:  E501
                        p.save_book_as(file_name=file_folder_list['file_name'][number], dest_file_name=new_name)  # noqa:  E501
                        os.remove(file_folder_list['file_name'][number])
                        file_folder_list['file_name'][number] = new_name
                except:  # noqa: E722, B001
                    total_result['Сломался на этапе смены формата'].append(file_folder_list['path'][number])  # noqa: E501
                    traceback.print_exc()
                    continue

            try:
                clean_file(file_folder_list['file_name'][number], file_folder_list['path'][number])
                total_result['Успешно'].append(file_folder_list['folder_name'][number])
            except:  # noqa: E722, B001
                total_result['Сломана функция'].append(file_folder_list['path'][number])
                traceback.print_exc()
                continue
    return print('Успешно обработано: ', len(set(total_result['Успешно'])), '\n',  # noqa: T201
                 'Сломался на этапе смены формата: ', len(set(total_result['Сломался на этапе смены формата'])), '\n',  # noqa: E501
                 'Сломана функция: ', len(set(total_result['Сломана функция'])), '\n',
                 'Отсутсвует', len(set(total_result['Отсутсвует'])), '\n',
                 'Ручное', len(set(total_result['Ручное'])), '\n',
                 'Успешно обработано: ', set(total_result['Успешно']), '\n',
                 'Сломался на этапе смены формата: ', set(total_result['Сломался на этапе смены формата']), '\n',  # noqa: E501
                 'Сломана функция: ', set(total_result['Сломана функция']), '\n',
                 'Отсутсвует', set(total_result['Отсутсвует']), '\n',
                 'Ручное', set(total_result['Ручное']))
