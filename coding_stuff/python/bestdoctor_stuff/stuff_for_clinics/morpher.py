
import math
import traceback
import time

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import pyexcel as p
from datetime import timedelta, datetime
import re
import pandas as pd
import yadisk
from mitosheet import *
import psycopg2
from typing import List,AnyStr



y = yadisk.YaDisk('406760da5b1345a88999f0acb1ef95bf', '7ecb5577edb643d3a07ce020292db4b6',
                  "AQAEA7qkBXctAAfNlAAXrxr49UN8l0uBHNszaAo")

gp_conn = gp_conn = 'postgresql://zeppelin:R63v5NspNsSEem@c-c9qbht031ah0gtrlftmj.rw.mdb.yandexcloud.net:5432/warehouse' # подлючение под пользаком zeppelin


headers = {
    'policy_number': ['Номер полиса', '№ полиса', 'Страховой полис', 'Полис', 'ID', ''''№ полиса или № по базе''',
                      'POLICY', '№ полиса ДМС', '№ полиса или № по базе', '№ полиса (карточки)', '№\nполиса',
                      'Данные полиса', 'Номер полиса ', 'Полис №', 'ID карта пациента', '№Полиса', 'Номер \nполиса',
                      '№ полиса ДМС', '№ Полиса ДМС', 'Номер полиса \nДМС\n', 'ПОЛИС',
                      'Номер полиса ДМС', 'Номер страхового полиса', '№ Полиса', 'КОД УСЛУГИ'],
    'fio': ['ФИО', 'Пациент', 'ФИО пациента', 'ФИО застрахованного', 'Ф.И.О.', 'гнитно', '\nФИО',
            ''''Ф.И.О., возраст (лет)''', 'Фамилия имя Отчество', 'ФИО                   пациента',
            'Пациент (ФИО)', 'Ф.И.О., возраст (лет) ', 'Ф.И.О. застрахованных', 'Фамилия И.О. пациента',
            'Фамилия, имя, отчество', 'ФИО \nзастрахованных (по алфавиту)\n', 'Фамилия, Имя, отчество',
            'Застрахованного', 'Фамилия И. О. № полиса', 'Фамилия И.О. \n№ полиса', 'ФИО \nзастрахованного',
            'ФИО\nпациента', 'ФИО Застрахованного', 'Ф.И.О. ', 'ФИО застрахованного (полностью)',
            'Фамилия Имя Отчество застрахованного', 'ФИО застрахован. ', 'Ф.И.О. пациента', 'Фамилия Имя Отчество',
            'ФИО Застрахованного (в одну ячейку; делить на три раздельных ячейки не надо)', 'ФИО Пациента',
            ],
    'guarantee_letter': ['№ ГП', 'ГП', '№ гарантийного письма ', 'Номер ГП', 'номер гарантийного письма ',
                         '№ и дата гарантийного письма', 'Номер и дата гарантийного письма', 'Дата и номер ГП',
                         'дата/номер гарантийного письма ', '№ гарантийного письма', 'Гарантийное письмо',
                         'Гарантийное \nписьмо \nномер', 'НОМЕР ГП', 'Номер гарантийного письма'],
    'first_name': ['Имя', 'P_NAME', 'Имя Застрахованного', 'Имя ', 'Номер гп'],
    'last_name': ['Фамилия', 'P_FAMILY', 'Фамилия Застрахованного', 'Фамилия '],
    'middle_name': ['Отчество', 'P_PATRONIMIC', 'Отчество Застрахованного ', 'Отчество '],
    'date': ['Дата', 'Дата услуги', 'Дата начала оказания услуги', ''''Дата предоставления услуги''',
             'Дата обращения', 'Дата и время\nначала', 'Начало', 'Дата окачания услуги', 'Дата оказания услуг',
             'Дата обр.', 'DATA', 'Дата приема', 'Дата оказания услуги', 'дата', 'Дата предоставления услуги',
             'Дата взятия пробы', 'Дата оказания медицинской услуги', 'Дата\nначала услуги', 'дата оказания',
             'Дата оказания услуги ', 'Дата оказания медицинской услуги ', 'оказания', 'Дата оказания',
             'ДАТА ОКАЗАНИЯ УСЛУГИ ', 'Дата поступления (госпитализация)', 'Дата начала и окончания госпитализации',
             'Дата \nоказания \nуслуги', 'дата оказания услуги', 'ДАТА ОКАЗАНИЯ УСЛУГИ'],
    'end_date': ['Дата окончания оказания услуги', 'Дата выполнения', 'Дата выезда', 'Дата окончания', 'Дата выписки',
                 'Дата и время\nвозвращения', 'Дата окончания услуги'],
    'doctor_name': ['Доктор', 'Врач', 'Ф.И.О. доктора ', 'Исполнитель', 'ФИО врача, оказавшего услугу', 'ФИО врача',
                    '\nФИО врача', 'Врач                              (ФИО)', 'ФИО доктора', 'Ф.И.О. Врача',
                    'Код врача (или ФИО)', 'ФИО лечащего врача', 'ВРАЧ'],
    'doctor_last_name': ['D_FAMILY'],
    'doctor_first_name': ['D_NAME'],
    'doctor_middle_name': ['D_PATRONIMIC'],
    'tooth_number': ['№ зуба', 'Зуб №', 'Номер зуба (для стоматологических услуг)', 'Зуб', 'Зуб - номер', '№ Зуба',
                     'Номер зуба', '№ зуба (по международной нумерации)', '№ зуба', '№ ЗУБА',
                     '№ зуба (по междун. классиф.)'],
    'mkb': ['МКБ-10', 'Код по МКБ-10', 'Код по МКБ', 'Код МКБ', 'Код МКБ-10', 'Код диагноза по МКБ10', 'МКБ',
            'Шифр заболевания по МКБ-10', 'Код диагноза по МКБ-Х или развернутый диагноз', 'Код диагноза по МКБ-Х',
            'Диагноза по МКБ-Х', 'МКБ10', 'Код диагноза по МКБ10 (не менее четырех знаков)', 'Код диагноза\n (МКБ-10)',
            'Код диагноза по МКБ-Х или развернутый диагноз ', 'Код\nдиагн.', 'Кол.\nУслуг', 'КОД ДИАГНОЗА ПО МКБ-10',
            'Код диагноза по МКБ 10  и стомат. диагноз', 'Код диагноза по МКБ-10',
            'Код диагноза по МКБ 10 и диагноз', ],
    'service_code': ['Код услуги', 'Код услуги по прейск.ЛПУ', 'Код', ''''Код услуги''', 'SRV_CODE', 'Код Услуги ',
                     'Код мед. услуги по прейскуранту', 'Код услуги по Прейскуранту', 'Код\nуслуги', 'Код  услуги',
                     'Код МЗ', 'КОД УСЛУГИ', 'Код услуги по прейскуранту', 'Код\n услуги', 'Код услуги '
                                                                                           'Код услуги по прейскуранту ЛПУ',
                     'Код медицинской услуги', 'Код услуги (Прейскурант)',
                     'Код услуги по прейск', 'Код услуги '],
    'service_name': ['Наименование услуги', 'Название услуги', 'Название мед.услуги', '\nНаименование услуги',
                     'Наименование медицинской услуги', ''''Наименование услуги''',
                     'Наименование выполненной услуги', 'Полное наименование услуги по прейскуранту',
                     'Услуга', 'SRV_NAME', 'Оказанные услуги', 'Наименование мед. услуги', 'Анализ / Профиль',
                     'Наименование медицинской услуги или ее кода по Прейскуранту', 'наименование услуги',
                     'Наименование услуги (код по Перечню услуг)', 'Название услуг', 'Наименование \nуслуги',
                     'Наименование медицинской услуги по Прейскуранту', 'Код и название мед. услуги',
                     'Наименование услуги подменная', 'Наименование медицинской услуги по Прейскуранту',
                     'оказанных услуг', 'Выполненные исследования', 'НАИМЕНОВАНИЕ УСЛУГИ', 'Наименование услуги '],
    'service_price': ['Цена, руб.', 'Стоимость', 'Стоимость (руб.)', 'Цена', 'Цена (руб)', 'Цена услуги',
                      'Цена услуги ',
                      ''''Стоимость услуги''', 'SRV_COST', 'Цена по прейскуранту', 'Цена 1-й услуги, руб.',
                      'Стоимость мед. услуги (руб.)', 'Стоимость услуги', 'Стоимость 1 услуги руб.', 'Цена ', 'Сумма',
                      'Цена услуги по прейскуранту  ', 'Стоимость \nуслуги\n', 'Стоим. за ед.', 'Стоим. за\nед.',
                      'Общая стоим.', 'ЦЕНА', 'Стоимость\nруб.', 'Цена     услуги', 'Цена услуги ', 'Цена\nуслуги',
                      'Цена за единицу', 'Цена, \nруб.', 'цена*кол-во (стоимость)'],
    'service_amount': ['Кол-во', 'Пр.', 'Количество', ''''Кол-во услуг''', 'Кол-во услуг', 'SRV_TOTAL', 'кол-во услуг',
                       'Кол-во мед. услуг', 'Количество услуг', 'К-во услуг', 'Кол-во', 'Кол-во услуг  ',
                       'Количество \nуслуг\n', 'услуг', 'Кол. услуг', 'Кол. Услуг', 'КОЛ.  УСЛУГ',
                       'Количество оказанных услуг', 'Кол-во ', 'кол-во', 'Кол- во', 'КОЛ-ВО',
                       'Кол-во оказанных услуг'],
    'total_price': ['Сумма, руб.', 'Сумма к оплате', 'Стоимость (руб.)', 'Стоимость', ''''Общая стоимость''',
                    'Сумма рублей', 'Стоимость услуг по прейскуранту без учета скидки (руб)', 'SRV_SUMM',
                    'Начислено', 'Сумма к оплате, руб.', 'Полная стоимость', 'Ст-ть услуг (руб.)', 'стоимость',
                    'Итого (руб.)', 'Общая стоимость', 'Общая сумма руб.', 'Общая Сумма', 'Ст-ть',
                    'Стоимость руб.', 'услуги,', 'СТОИМОСТЬ', 'Стоимость услуг (Цена*Количество)',
                    'Цена (руб) * кол-во услуг', 'Стоимость, руб.', 'Полная стоимость'],
    'payment_type': ['Тип оплаты'],
    'discount_sice': ['% Скидка', 'Франшиза', 'Спецсект (коэф.)', 'Коэф', 'Скидка, %', 'Скидка или надбавка, %',
                      'Скидка', 'СКИДКА'],
    'diagnosis': ['Диагноз', ''''Диагноз''', 'SRV_DIAG', 'Код диагноза', 'Диагноз клинический (текст)',
                  'Диагноз клинический', 'Код диагн.', 'ДИАГНОЗ (текст)', 'Диагноз,\n№ зуба', 'Диагноз (по МКБ 10)',
                  'Диагноз, номер зуба', 'Диагноз (код по МКБ)', 'Диагноз или код (по МКБ-10)', 'Диaгноз'],
    'clinic_code': ['Код клиники'],
    'clinic_name': ['Клиника-Исполнитель (при Сети Клиник)', 'Наименование филиала клиники'],
    'doctor_code': ['Код врача', 'Код врача, оказавшего услугу ', 'Код врача'],
    'doctor_speciality': ['Специальность доктора', 'Специальность врача назначившего обследование/лечение',
                          'Специаль-ность врача', 'Специальность врача'],
    'doctor_speciality_2': ['Специальность врача оказавшего услугу  ', 'Специальность отправителя'],
    'branch_code': ['Номер филиала'],
    'branch_name': ['Отделение ', 'Филиал', 'Учреждение', 'Отделение', 'Подразделение'
                                                                       'Отделение (поликлинический прием, лаборатория, функционалой диагностики, физиотерапия, ПНД, процедурный кабинет, дневной стационар, стоматология)',
                    'Отделение (поликлинический прием, лаборатория, функционалой диагностики, физиотерапия, ПНД, процедурный кабинет, дневной стационар, стоматология)',
                    'Отделение / Специальность', 'Наименование филиала'],
    'number_disease_history': ['Номер медицинской карты /истории болезни', 'Номер мед.карты'],
    'price-list_id': ['ID прейскуранта'],
    'service_type': ['Тип обслуживания / Вид помощи', 'Код вида обращения'],
    'medical_service': ['Медицинская услуга'],
}
shablon = ['ФИО пациента', 'Страховой полис', 'Дата начала оказания услуги', 'Дата окончания оказания услуги',
           'Код услуги', 'Наименование услуги', 'Код МКБ-10', 'Диагноз', 'Номер зуба (для стоматологических услуг)',
           'Цена услуги', 'Количество', 'Скидка, %', 'Код филиала клиники (точки)',
           'Наименование филиала клиники (точки)', '№ ГП', 'Код врача', 'Врач (ФИО)', 'Специальность врача',
           'Специальность направившего врача', 'Код отделения', 'Наименование отделения', '№ истории болезни',
           'ID Прейскуранта', 'Тип оплаты', 'Тип обслуживания / Вид помощи', 'Медицинская услуга']
LOCAL_HEADERS = {
    'policy_number': ['Страховой полис'],
    'fio': ['ФИО пациента'],
    'guarantee_letter': ['№ ГП'],
    'date': ['Дата начала оказания услуги'],
    'end_date': ['Дата окончания оказания услуги'],
    'doctor_name': ['Врач (ФИО)'],
    'tooth_number': ['Номер зуба (для стоматологических услуг)'],
    'mkb': ['Код МКБ-10', 'МКБ10'],
    'service_code': ['Код услуги'],
    'service_name': ['Наименование услуги'],
    'service_price': ['Цена услуги'],
    'service_amount': ['Количество'],
    'total_price': ['Сумма, руб.'],
    'discount_sice': ['Скидка, %'],
    'payment_type': ['Тип оплаты'],
    'diagnosis': ['Диагноз'],
    'clinic_code': ['Код филиала клиники (точки)'],
    'clinic_name': ['Наименование филиала клиники (точки)'],
    'doctor_code': ['Код врача'],
    'doctor_speciality': ['Специальность врача'],
    'doctor_speciality_2': ['Специальность направившего врача'],
    'branch_code': ['Код отделения'],
    'branch_name': ['Наименование отделения'],
    'number_disease_history': ['№ истории болезни'],
    'price-list_id': ['ID Прейскуранта'],
    'service_type': ['Тип обслуживания / Вид помощи'],
    'medical_service': ['Медицинская услуга'],
}


class DwsConn:
    def __init__(self, conn_string):
        self.conn_string = conn_string

    def select(self, query: str) -> List[tuple]:
        with psycopg2.connect(self.conn_string) as conn:
            with conn.cursor() as cursor:
                cursor.execute(query)
                return cursor.fetchall()

    def execute(self, query: str) -> None:
        with psycopg2.connect(self.conn_string) as conn:
            with conn.cursor() as cursor:
                cursor.execute(query)
                conn.commit()

DWH = DwsConn(conn_string=gp_conn)
data = DWH.select('select 1')
class DwsConn:
    def __init__(self, conn_string):
        self.conn_string = conn_string

    def select(self, query: str) -> List[tuple]:
        with psycopg2.connect(self.conn_string) as conn:
            with conn.cursor() as cursor:
                cursor.execute(query)
                return cursor.fetchall()

    def execute(self, query: str) -> None:
        with psycopg2.connect(self.conn_string) as conn:
            with conn.cursor() as cursor:
                cursor.execute(query)
                conn.commit()


def get_key(dict, value):
    for k, v in dict.items():
        if len(v) > 0:
            if v[0] == value:
                return k

def med_cent_stolica(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=7)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'Пациент': 'ФИО пациента'}, inplace=True)

    # Renamed columns Дата начала оказания услуги
    Sheet_Name_Here.rename(columns={'Дата': 'Дата начала оказания услуги'}, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'номер страхового полиса (гарантийного письма)': 'Страховой полис'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'наименование услуги': 'Наименование услуги'}, inplace=True)

    # Renamed columns Код услуги
    Sheet_Name_Here.rename(columns={'код услуги': 'Код услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во': 'Количество'}, inplace=True)

    # Deleted columns Стоимость
    Sheet_Name_Here.drop(['Стоимость'], axis=1, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Цена': 'Цена услуги'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз': 'Код МКБ-10'}, inplace=True)

    # Filtered Наименование услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Наименование услуги'].notnull()]

    # Filtered Страховой полис
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Страховой полис'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)


def a_stom(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=12)
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]

    # Deleted columns № п/п
    Sheet_Name_Here.drop(['№ п/п'], axis=1, inplace=True)

    # Renamed columns Страховой полис
    Sheet_Name_Here.rename(columns={'ID': 'Страховой полис'}, inplace=True)

    # Renamed columns ФИО пациента
    Sheet_Name_Here.rename(columns={'ФИО': 'ФИО пациента'}, inplace=True)

    # Renamed columns Код МКБ-10
    Sheet_Name_Here.rename(columns={'Диагноз (код по МКБ-10)': 'Код МКБ-10'}, inplace=True)

    # Renamed columns Цена услуги
    Sheet_Name_Here.rename(columns={'Стоимость услуги': 'Цена услуги'}, inplace=True)

    # Renamed columns Количество
    Sheet_Name_Here.rename(columns={'Кол-во услуг': 'Количество'}, inplace=True)

    # Deleted columns Общая сумма
    Sheet_Name_Here.drop(['Общая сумма'], axis=1, inplace=True)

    # Changed Количество to dtype int
    Sheet_Name_Here['Количество'] = Sheet_Name_Here['Количество'].fillna(0).astype('int')

    # Filled NaN values in 2 columns in Sheet1_1
    columns_to_fill_nan = ['Страховой полис', 'ФИО пациента']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filled NaN values in 1 columns in Sheet1_1
    columns_to_fill_nan = ['Дата услуги']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='ffill')

    # Filtered Код услуги
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Код услуги'].notnull()]

    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)

def allergomed(file_name, file_path):
    wb = load_workbook(file_name, data_only=True)
    sheet_name = wb.sheetnames[0]
    sheet_df_dictonary = pd.read_excel(file_name, engine='openpyxl', sheet_name=[sheet_name], skiprows=3)  # noqa: E501
    Sheet_Name_Here = sheet_df_dictonary[sheet_name]
    # Renamed columns Пропуск
    Sheet_Name_Here.rename(columns={'Застрахованного': 'Пропуск'}, inplace=True)

    # Added column new-column-ztow
    Sheet_Name_Here.insert(3, 'new-column-ztow', 0)

    # Renamed columns ФИО
    Sheet_Name_Here.rename(columns={'new-column-ztow': 'ФИО'}, inplace=True)

    # Set formula of ФИО
    Sheet_Name_Here['ФИО'] = IF(TYPE(Sheet_Name_Here['Пропуск']) != 'NaN', PROPER(Sheet_Name_Here['Пропуск']), None)  # noqa: E501

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'Диагноз': 'МКБ'}, inplace=True)

    # Renamed columns Наименование услуги
    Sheet_Name_Here.rename(columns={'оказанных услуг': 'Наименование услуги'}, inplace=True)

    # Renamed columns Дата услуги
    Sheet_Name_Here.rename(columns={'оказания': 'Дата услуги'}, inplace=True)

    # Renamed columns Кол-во
    Sheet_Name_Here.rename(columns={'услуг': 'Кол-во'}, inplace=True)

    # Renamed columns Цена
    Sheet_Name_Here.rename(columns={'услуги,': 'Цена'}, inplace=True)

    # Added column new-column-hjo7
    Sheet_Name_Here.insert(10, 'new-column-hjo7', 0)

    # Renamed columns ПРОПУСКСКИДКА
    Sheet_Name_Here.rename(columns={'new-column-hjo7': 'ПРОПУСКСКИДКА'}, inplace=True)

    # Added column new-column-doot
    Sheet_Name_Here.insert(11, 'new-column-doot', 0)

    # Renamed columns Скидка
    Sheet_Name_Here.rename(columns={'new-column-doot': 'Скидка'}, inplace=True)

    # Set formula of ПРОПУСКСКИДКА
    Sheet_Name_Here['ПРОПУСКСКИДКА'] = IF(
        AND(TYPE(Sheet_Name_Here['№ полиса']) != 'NaN', FIND(Sheet_Name_Here['№ полиса'], 'Скидка')), SUBSTITUTE(  # noqa: E501
            SUBSTITUTE(Sheet_Name_Here['№ полиса'],
                       LEFT(Sheet_Name_Here['№ полиса'], INT(FIND(Sheet_Name_Here['№ полиса'], '='))), ''), ' руб.',  # noqa: E501
            ''), None)

    # Filled NaN values in 1 columns in Sheet_Name_Here
    columns_to_fill_nan = ['ПРОПУСКСКИДКА']
    Sheet_Name_Here[columns_to_fill_nan] = Sheet_Name_Here[columns_to_fill_nan].fillna(method='bfill')  # noqa: E501

    # Set formula of Скидка
    Sheet_Name_Here['Скидка'] = IF(Sheet_Name_Here['ПРОПУСКСКИДКА'] == '0', None, Sheet_Name_Here['ПРОПУСКСКИДКА'])  # noqa: E501

    # Filtered Кол-во
    Sheet_Name_Here = Sheet_Name_Here[Sheet_Name_Here['Кол-во'].notnull()]

    # Changed Кол-во to dtype int
    Sheet_Name_Here['Кол-во'] = Sheet_Name_Here['Кол-во'].fillna(0).astype('int')

    # Changed Цена to dtype float
    Sheet_Name_Here['Цена'] = to_float_series(Sheet_Name_Here['Цена'])

    # Added column new-column-49l5
    Sheet_Name_Here.insert(5, 'new-column-49l5', 0)

    # Renamed columns ПропускМКБ
    Sheet_Name_Here.rename(columns={'МКБ': 'ПропускМКБ'}, inplace=True)

    # Renamed columns МКБ
    Sheet_Name_Here.rename(columns={'new-column-49l5': 'МКБ'}, inplace=True)

    # Set formula of МКБ
    Sheet_Name_Here['МКБ'] = IF(AND(TYPE(Sheet_Name_Here['ПропускМКБ']) != 'NaN', Sheet_Name_Here['ПропускМКБ'] != ' '),  # noqa: E501
                                Sheet_Name_Here['ПропускМКБ'], None)
    df = pd.DataFrame.from_dict(Sheet_Name_Here)
    df.to_excel(file_name)
    return look_data(file_name, file_path)

def template_qualifier(folder_name, printer=False):
    template_tuple = {
        'Медицинский центр Столица, ООО': med_cent_stolica,
        'А-Стом, ООО': a_stom,
        'Аллергомед Клиника  ООО': allergomed,
    }
    if printer:
        summ = 0
        for i in template_tuple.keys():
            summ += 1
        return print(summ)
    if folder_name in template_tuple:
        sample = template_tuple[folder_name]
    else:
        sample = 'clinic_name'
    return sample


def look_data(local_path, file_path):
    wb = load_workbook(local_path, data_only=True)
    done_2 = False
    skip_row_list = []
    min_row = None
    max_row = None
    sheet_number = None
    for i in range(0, len(wb.sheetnames)):
        sheet_name = wb.sheetnames[i]
        sheet = wb[sheet_name]
        for column_number in range(1, sheet.max_column):
            done = False
            if done_2:
                break
            for row_number in range(1, (sheet.max_row + 2)):
                inform = sheet.cell(row=row_number, column=column_number)
                if not done:
                    for key in headers:
                        if inform.value in headers[key]:
                            min_row = row_number
                            sheet_number = i
                            done = True
                            break
                else:
                    if inform.value is None:
                        try:
                            sheet_df_dictonary = pd.read_excel(local_path, engine='openpyxl', sheet_name=[sheet_name],
                                                               skiprows=min_row - 1)
                            Sheet_Name_Here = sheet_df_dictonary[sheet_name]
#                          test_collumn_name(Sheet_Name_Here)
                        except:
                            pass
                        max_row = row_number - 1
                        done_2 = True
                        break
                    else:
                        continue
    return check_data(local_path, min_row, max_row, sheet_number, file_path, skip_row_list)


def check_data(local_path, min_row, max_row, sheet_number, file_path, skip_row_list=None):
    global key_name, inform_value
    data_list = {
        'policy_number': [],
        'fio': [],
        'guarantee_letter': [],
        'first_name': [],
        'last_name': [],
        'middle_name': [],
        'date': [],
        'end_date': [],
        'doctor_name': [],
        'doctor_last_name': [],
        'doctor_first_name': [],
        'doctor_middle_name': [],
        'tooth_number': [],
        'mkb': [],
        'service_code': [],
        'service_name': [],
        'service_price': [],
        'service_amount': [],
        'total_price': [],
        'payment_type': [],
        'discount_sice': [],
        'diagnosis': [],
        'clinic_code': [],
        'clinic_name': [],
        'doctor_code': [],
        'doctor_speciality': [],
        'doctor_speciality_2': [],
        'branch_code': [],
        'branch_name': [],
        'number_disease_history': [],
        'price-list_id': [],
        'service_type': [],
        'medical_service': [],
    }
    wb = load_workbook(local_path, data_only=True)
    sheet_name = wb.sheetnames[sheet_number]
    sheet = wb[sheet_name]
    for column_number in range(1, sheet.max_column + 1):
        done = False
        for row_number in range(min_row, max_row + 1):
            if skip_row_list is not None:
                if row_number in skip_row_list:
                    continue
            inform = sheet.cell(row=row_number, column=column_number)
            inform_value = inform.value
            if done:
                if inform.value is not None:
                    data_list[key_name].append(str(inform_value))
                else:
                    data_list[key_name].append(None)
            if not done:
                for key in headers:
                    if inform.value in headers[key]:
                        key_name = key
                        done = True
                        break
    try:
        return clear_data(data_list, local_path, file_path)
    except:
        print('Смена формата не прошла', file_path)
        return filling_book(data_list, local_path, file_path)


def clear_data(data_list, local_path, file_path):
    try:
        data_list['policy_number'] = [int(re.sub('^0{2}|^0{3}|^0{4}|^0{5}', '', x)) if x is not None else x for x in
                                      data_list['policy_number']]
    except:
        try:
            data_list['policy_number'] = [x for x in data_list['policy_number']]
        except:
            print('policy_number с ошибкой\n', data_list['policy_number'])
            pass

    try:
        data_list['date'] = [
            datetime.strptime(x, '%Y-%m-%d %H:%M:%S').date().strftime('%d.%m.%Y') if x is not None else x for x in
            data_list['date']]
    except:
        try:
            data_list['date'] = [datetime.strptime(x, '%Y-%m-%d').date().strftime('%d.%m.%Y') if x is not None else x
                                 for x in data_list['date']]
        except:
            try:
                data_list['date'] = [
                    datetime.strptime(x, '%d-%m-%Y').date().strftime('%d.%m.%Y') if x is not None else x for x in
                    data_list['date']]
            except:
                try:
                    for x, y in enumerate(data_list['date']):
                        if y is None:
                            continue
                        try:
                            data_list['date'][x] = datetime.strptime(y, '%Y-%m-%d %H:%M:%S').date().strftime('%d.%m.%Y')
                        except:
                            try:
                                data_list['date'][x] = datetime.strptime(y, '%d.%m.%Y').date().strftime('%d.%m.%Y')
                            except:
                                continue
                except:
                    try:
                        data_list['date'] = [x.split(' ', 2)[0] for x in data_list['date']]
                    except:
                        try:
                            data_list['date'] = [x.split(' ', 2)[0].replace('-', '.').split('.', -1)[-1] + '.' + \
                                                 x.split(' ', 2)[0].replace('-', '.').split('.', -1)[-2] + '.' + \
                                                 x.split(' ', 2)[0].replace('-', '.').split('.', -1)[0] for x in
                                                 data_list['date']]
                        except:
                            print('date с ошибкой\n', data_list['date'])
    try:
        data_list['service_price'] = [float(x) for x in data_list['service_price']]
    except:
        try:
            data_list['service_price'] = [float(x.replace('.', ',')) for x in data_list['service_price']]
        except:
            pass
    try:
        data_list['service_amount'] = [int(x) for x in data_list['service_amount']]
    except:
        pass
    # try:
    #     data_list['guarantee_letter'] = [str(x).replace('ГП № ', '').replace(' от', '') for x in
    #                                      data_list['guarantee_letter']]
    #     if len(data_list['guarantee_letter'][0]) > 15:
    #         data_list['guarantee_letter'] = [x.split(' ', 2)[0] for x in data_list['guarantee_letter']]
    # except:
    #     try:
    #         data_list['guarantee_letter'] = [int(x) for x in data_list['guarantee_letter']]
    #     except:
    #         print('guarantee_letter с ошибкой\n', data_list['guarantee_letter'])
    #         pass

    try:
        data_list['tooth_number'] = [x if x is not None else x for x in data_list['tooth_number']]
    except:
        print('tooth_number с ошибкой\n', data_list['tooth_number'])
        pass
    # sql_to_base(data_list) # генерация запроса для импорта в таблицу данных
    return filling_book(data_list, local_path, file_path)


def filling_book(data_list, file_name, file_path):  # тип оплаты факт
    print("filling book started")
    global maximum_len, value, list_data_2, list_data_3, temp_type
    month_list = ('Ноябрь', 'октябрь', 'Июнь', 'ноябрь', 'июль', '10', 'сентябрь', '11', 'Октябрь', 'май', 'март',
                  'август')
    wb = Workbook()
    ws = wb.active
    ws.append(shablon)
    for month in month_list:
        if month in file_path:
            new_file_folder = file_path.split('/', -1)[-3].replace('"', '').replace("""'""", '')
        else:
            new_file_folder = file_path.split('/', -1)[-2].replace('"', '').replace("""'""", '')
    # new_file_name = str(file_name.replace('.xlsx', '') + '_' + 'morphed.xlsx')
    new_file_name = str(new_file_folder + '_' + file_name.replace('.xlsx', '') + '_' + 'morphed.xlsx')
    if len(data_list['fio']) == 0:
        maximum_len = len(data_list['last_name'])
    else:
        maximum_len = len(data_list['fio'])
    for column_number in range(1, (len(shablon) + 1)):
        for row_number in range(2, maximum_len + 2):
            car = get_column_letter(column_number)
            value = get_key(LOCAL_HEADERS, shablon[column_number - 1])
            if column_number == 24:
                ws[car + str(row_number)] = 'fact'
            if value == 'fio' and len(data_list['fio']) == 0:
                ws[car + str(row_number)] = data_list['last_name'][row_number - 2] \
                                            + ' ' + data_list['first_name'][row_number - 2] \
                                            + ' ' + data_list['middle_name'][row_number - 2]
                continue
            elif value == 'doctor_name' and len(data_list['doctor_name']) == 0 \
                    and (len(data_list['doctor_last_name']) + len(data_list['doctor_first_name'])
                         + len(data_list['doctor_middle_name'])) > 0:
                ws[car + str(row_number)] = data_list['doctor_last_name'][row_number - 2] \
                                            + ' ' + data_list['doctor_first_name'][row_number - 2] \
                                            + ' ' + data_list['doctor_middle_name'][row_number - 2]
                continue
            elif value == 'service_price' and len(data_list['total_price']) > 0:
                ws[car + str(row_number)] = data_list['total_price'][row_number - 2]
            elif value is None or data_list[value] is None or len(data_list[value]) < 1:
                continue
            else:
                ws[car + str(row_number)] = data_list[value][row_number - 2]
    print("before removing")
    wb.save(new_file_name)
    try:
        # os.remove(file_name)
        y.remove(f"{file_name}")
    except:
        pass
    new_file_path = file_path.replace(file_name, '') + new_file_name
    print(new_file_name)
    print(new_file_path)
    try:
        y.upload(f"{new_file_name}", new_file_path)
        print('success')
    except:
        pass


def start(path='disk:/Clinics/morphed', files_path={'path': [], 'file_name': [], 'folder_name': []}):
    for file in y.listdir(path):
        if file.type == 'dir':
            start(file.path, files_path)
        else:
            if '.xls' in file.name and 'morphed' not in file.name:
                files_path['path'].append(file.path)
                files_path['file_name'].append(file.name)
                files_path['folder_name'].append(file.path.split('/', -1)[3])
    return files_path



def start_test(files_path={'path': [], 'file_name': [], 'folder_name': []}):
    list_of_dirs = DWH.select("select name,path from yandex_disk.clinics_files where status is null;")
    print(list_of_dirs)
    for tuple in list_of_dirs:
        try:
            files_path['folder_name'].append(tuple[1].split('/', -1)[4])
            files_path['path'].append(tuple[1])
            files_path['file_name'].append(tuple[0])
        except:
            pass
    return files_path

def main_loop(file_folder_list):
    total_result = {'Успешно': [],
                    'Сломался на этапе смены формата': [],
                    'Сломана функция': [],
                    'Отсутсвует': [],
                    'Другое': [],
                    'Ручное': [],
                    }
    for number, reg_file in enumerate(file_folder_list['path']):
        if 'morphed' in file_folder_list['file_name'][number]:
            continue
        # for i in test:
        #     if i in file_folder_list['folder_name'][number]:
        # if 'ООО "Гранти-мед"'  in file_folder_list['folder_name'][number]:
        if 'test' not in file_folder_list['folder_name'][number]:
            try:
                print(file_folder_list['file_name'][number])
                y.download(reg_file, file_folder_list['file_name'][number])
            except:
                print(file_folder_list['file_name'][number])
                pass
            print(file_folder_list['folder_name'][number])
            clean_file = template_qualifier(file_folder_list['folder_name'][number])
            try:
                if 'clinic_name' in clean_file:
                    total_result['Отсутсвует'].append(file_folder_list['path'][number])
                    continue
                elif 'Ручное' in clean_file:
                    total_result['Ручное'].append(file_folder_list['path'][number])
                    continue
            except:
                pass
            if file_folder_list['file_name'][number].split('.', -1)[-1].lower() == 'xls':
                new_name = file_folder_list['file_name'][number].lower() + 'x'
                try:
                    p.save_book_as(file_name=file_folder_list['file_name'][number], dest_file_name=new_name)
                    os.remove(file_folder_list['file_name'][number])
                    file_folder_list['file_name'][number] = new_name
                except:
                    total_result['Сломался на этапе смены формата'].append(file_folder_list['path'][number])
                    continue
            try:
                clean_file(file_folder_list['file_name'][number], file_folder_list['path'][number])
                total_result['Успешно'].append(file_folder_list['folder_name'][number])
            except:
                total_result['Сломана функция'].append(file_folder_list['path'][number])
                # total_result['Сломана функция'].append(file_folder_list['folder_name'][number])
                traceback.print_exc()
                continue
    return print('Успешно обработано: ', len(set(total_result['Успешно'])), '\n',
                 'Сломался на этапе смены формата: ', len(set(total_result['Сломался на этапе смены формата'])), '\n',
                 'Сломана функция: ', len(set(total_result['Сломана функция'])), '\n',
                 'Отсутсвует', len(set(total_result['Отсутсвует'])), '\n',
                 'Ручное', len(set(total_result['Ручное'])), '\n',
                 'Успешно обработано: ', set(total_result['Успешно']), '\n',
                 'Сломался на этапе смены формата: ', set(total_result['Сломался на этапе смены формата']), '\n',
                 'Сломана функция: ', set(total_result['Сломана функция']), '\n',
                 'Отсутсвует', set(total_result['Отсутсвует']), '\n',
                 'Ручное', set(total_result['Ручное']), )




main_loop(start_test())

#a = start(path='disk:/Clinics/morphed/')

#main_loop(a)

# Lim = Скрипт для большинства ручных клиник ( можно на его базе пробовать обрабатывать остальные таким же образом )



